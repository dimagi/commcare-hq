"""
Read-only profiling for SAAS-20053. Answers "how much data would we delete?"
before running the retire script, since retire()'s Celery fan-out scales with
the number of forms and cases per user, not just the user count.

Deletes nothing. Run in the Django shell on the target environment:

    cchq <env> django-manage shell < scratch_profile_ndoh_wbot_training.py
"""
import math

from openpyxl import load_workbook

from dimagi.utils.chunked import chunked

from corehq.apps.es import CaseES, FormES
from corehq.apps.es.aggregations import TermsAggregation
from corehq.apps.users.dbaccessors import get_user_docs_by_username
from corehq.apps.users.models import CommCareUser, CouchUser
from corehq.apps.users.util import format_username

# ------------------------------- config -------------------------------
DOMAIN = 'ndoh-wbot-training'
USERNAMES_FILE = '/tmp/saas_20053_users.xlsx'  # xlsx with a "username" column
# ----------------------------------------------------------------------


def load_usernames(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    usernames = []
    for (value,) in wb.active.iter_rows(min_col=1, max_col=1, values_only=True):
        if value is None:
            continue
        value = str(int(value)) if isinstance(value, (int, float)) else str(value).strip()
        if value and value.lower() != 'username':
            usernames.append(value)
    wb.close()
    return list(dict.fromkeys(usernames))


def deactivated_user_ids(domain, usernames):
    formatted = {
        (u.lower() if '@' in u else format_username(u, domain)): u
        for u in usernames
    }
    user_ids = []
    for chunk in chunked(list(formatted), 100):
        for doc in get_user_docs_by_username(chunk):
            user = CouchUser.wrap_correctly(doc)
            if (isinstance(user, CommCareUser) and user.domain == domain
                    and not user.is_deleted() and not user.is_active_in_domain(domain)):
                user_ids.append(user._id)
    return user_ids


def counts_by_user(build_query, agg_field, user_ids):
    """Return {user_id: doc_count} using a terms aggregation, chunked."""
    counts = {}
    for chunk in chunked(user_ids, 1000):
        result = (
            build_query(chunk)
            .aggregation(TermsAggregation('by_user', agg_field).size(len(chunk)))
            .size(0)
            .run()
        )
        counts.update(result.aggregations.by_user.counts_by_bucket())
    return counts


def summarize(label, counts, user_ids):
    values = [counts.get(uid, 0) for uid in user_ids]
    total = sum(values)
    nonzero = [v for v in values if v]
    print(f"{label}:")
    print(f"  total:            {total:,}")
    print(f"  users with any:   {len(nonzero):,} / {len(user_ids):,}")
    print(f"  max for one user: {max(values) if values else 0:,}")
    print(f"  avg (nonzero):    {int(total / len(nonzero)) if nonzero else 0:,}")
    print(f"  users > 1000:     {sum(1 for v in values if v > 1000):,}")
    return total


usernames = load_usernames(USERNAMES_FILE)
user_ids = deactivated_user_ids(DOMAIN, usernames)
print(f"\n{len(usernames)} username(s) -> {len(user_ids)} deactivated user(s) to delete in {DOMAIN}\n")

form_counts = counts_by_user(
    lambda ids: FormES().domain(DOMAIN).user_id(ids), 'form.meta.userID', user_ids)
case_counts = counts_by_user(
    lambda ids: CaseES().domain(DOMAIN).owner(ids), 'owner_id', user_ids)

total_forms = summarize("FORMS (submitted by user)", form_counts, user_ids)
total_cases = summarize("CASES (owned by user)", case_counts, user_ids)

# Rough Celery task fan-out estimate (see delete_user_data + users/tasks.py):
#   form tasks   : ceil(forms/50) per user
#   case tasks   : ceil(cases/50) per user, each firing 4 more (x5)
#   system-forms : 1 per user
#   case rebuilds: up to (total cases) tasks
form_tasks = sum(math.ceil(form_counts.get(u, 0) / 50) for u in user_ids)
case_tasks = sum(math.ceil(case_counts.get(u, 0) / 50) for u in user_ids) * 5
tagging_tasks = form_tasks + case_tasks + len(user_ids)
print("\nEstimated background_queue tasks from retiring all of these:")
print(f"  ~{tagging_tasks:,} tagging tasks + up to ~{total_cases:,} case-rebuild tasks")
print("  (tag_cases/tag_forms tasks are rate-limited to 2/sec/worker)")
