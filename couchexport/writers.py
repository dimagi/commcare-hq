import re
import zipfile
from StringIO import StringIO
import csv
import json
from django.template.loader import render_to_string
from openpyxl import style

class UniqueHeaderGenerator(object):
    def __init__(self, max_column_size=None):
        self.used = set()
        self.max_column_size = max_column_size or 2000

    def next_unique(self, header):
        header = self._next_unique(header)
        self.used.add(header)
        return header

    def _next_unique(self, string):
        counter = 1
        if len(string) > self.max_column_size:
            # truncate from the beginning since the end has more specific information
            string = string[-self.max_column_size:]
        orig_string = string
        while string in self.used:
            string = "%s%s" % (orig_string, counter)
            if len(string) > self.max_column_size:
                counterlen = len(str(counter))
                string = "%s%s" % (orig_string[-(self.max_column_size - counterlen):], counter)
            counter += 1

        return string

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

class ExportWriter(object):
    max_table_name_size = 500
    
    def open(self, header_table, file, max_column_size=2000):
        """
        Create any initial files, headings, etc necessary.
        """
        self._isopen = True
        self.max_column_size = max_column_size
        self._current_primary_id = 0
        self.file = file

        self._init()
        self.table_name_generator = UniqueHeaderGenerator(self.max_table_name_size)
        for table_name, table in header_table:
            self.add_table(table_name, table[0])

    def add_table(self, table_name, headers):
        def _clean_name(name):
            return re.sub(r"[[\\?*/:\]]", "-", name)
        table_name_truncated = _clean_name(self.table_name_generator.next_unique(table_name))

        # make sure we trim the headers
        with UniqueHeaderGenerator(self.max_column_size) as g:
            try:
                headers.data = [g.next_unique(header) for header in headers.data]
            except AttributeError:
                headers = [g.next_unique(header) for header in headers]
        self._init_table(table_name, table_name_truncated)
        self.write_row(table_name, headers)

    def write(self, document_table, skip_first=False):
        """
        Given a document that's been parsed into the appropriate
        rows, write those rows to the resulting files.
        """
        assert self._isopen
        for table_name, table in document_table:
            for i, row in enumerate(table):
                if skip_first and i is 0:
                    continue
                # update the primary component of the ID to match
                # how many docs we've seen
                try:
                    row_has_id = row.has_id()
                except AttributeError:
                    row_has_id = False
                if row_has_id:
                    row.id = (self._current_primary_id,) + tuple(row.id[1:])

                self.write_row(table_name, row)
        
        self._current_primary_id += 1

    def write_row(self, table_name, headers):
        """
        Currently just calls the subclass's implementation
        but if we were to add a universal validation step,
        such a thing would happen here.
        """
        return self._write_row(table_name, headers)

    def close(self):
        """
        Close any open file references, do any cleanup.
        """
        assert(self._isopen)
        self._close()
        self._isopen = False

    def _init(self):
        raise NotImplementedError

    def _init_table(self, table_name, table_name_truncated):
        raise NotImplementedError

    def _write_row(self, sheet_index, row):
        raise NotImplementedError

    def _close(self):
        raise NotImplementedError

    @classmethod
    def get_data(cls, row):
        """
        Get around the fact that row can be either an iterable or
        a FormattedRow

        """
        try:
            return row.get_data()
        except AttributeError:
            return row

class CsvExportWriter(ExportWriter):
    
    def _init(self):
        self.tables = {}
        self.table_names = {}
        self.table_files = {}
        
        
    def _init_table(self, table_name, table_name_truncated):
        table_file = StringIO()
        writer = csv.writer(table_file, dialect=csv.excel)
        self.tables[table_name] = writer
        self.table_files[table_name] = table_file
        self.table_names[table_name] = table_name_truncated            
        
    def _write_row(self, sheet_index, row):
        def _encode_if_needed(val):
            return val.encode("utf8") if isinstance(val, unicode) else val

        row = map(_encode_if_needed, self.get_data(row))
        writer = self.tables[sheet_index]
        writer.writerow(row)
        
    def _close(self):
        """
        Close any open file references, do any cleanup.
        """
        archive = zipfile.ZipFile(self.file, 'w', zipfile.ZIP_DEFLATED)
        for index, name in self.table_names.items():
            archive.writestr("%s.csv" % name, self.table_files[index].getvalue())
        archive.close()
        self.file.seek(0)

class Excel2007ExportWriter(ExportWriter):
    max_table_name_size = 31
    
    def _init(self):
        try:
            import openpyxl
        except ImportError:
            raise Exception("It doesn't look like this machine is configured for "
                            "excel export. To export to excel you have to run the "
                            "command:  easy_install openpyxl")
        
        self.book = openpyxl.workbook.Workbook(optimized_write=True)
        self.tables = {}
        self.table_indices = {}
        
        
    def _init_table(self, table_name, table_name_truncated):
        sheet = self.book.create_sheet()
        sheet.title = table_name_truncated
        self.tables[table_name] = sheet
        self.table_indices[table_name] = 0

    
    def _write_row(self, sheet_index, row):
        sheet = self.tables[sheet_index]
        # NOTE: don't touch this. changing anything like formatting in the
        # row by referencing the cells will cause huge memory issues.
        # see: http://packages.python.org/openpyxl/optimized.html
        sheet.append([unicode(v) for v in self.get_data(row)])
        
    def _close(self):
        """
        Close any open file references, do any cleanup.
        """
        self.book.save(self.file)
        

class Excel2003ExportWriter(ExportWriter):
    max_table_name_size = 31
    
    def _init(self):
        try:
            import xlwt
        except ImportError:
            raise Exception("It doesn't look like this machine is configured for "
                            "excel export. To export to excel you have to run the "
                            "command:  easy_install xlutils")
        self.book = xlwt.Workbook()
        self.tables = {}
        self.table_indices = {}
        
    def _init_table(self, table_name, table_name_truncated):
        sheet = self.book.add_sheet(table_name_truncated)
        self.tables[table_name] = sheet
        self.table_indices[table_name] = 0

    def _write_row(self, sheet_index, row):
        row_index = self.table_indices[sheet_index]
        sheet = self.tables[sheet_index]
        # have to deal with primary ids
        for i, val in enumerate(self.get_data(row)):
            sheet.write(row_index,i,unicode(val))
        self.table_indices[sheet_index] = row_index + 1
    
    def _close(self):
        self.book.save(self.file)

class InMemoryExportWriter(ExportWriter):
    """
    Keeps tables in memory. Subclassed by other export writers.
    """
    
    def _init(self):
        self.tables = {}
        self.table_names = {}
    
    def _init_table(self, table_name, table_name_truncated):
        self.table_names[table_name] = table_name_truncated
        self.tables[table_name] = []
        
    def _write_row(self, sheet_index, row):
        table = self.tables[sheet_index]
        # have to deal with primary ids
        row_data = [val for val in self.get_data(row)]
        table.append(row_data)
        
    def _close(self):                
        pass

class JsonExportWriter(InMemoryExportWriter):
    """
    Write tables to JSON
    """

    class ConstantEncoder(json.JSONEncoder):

        def default(self, obj):
            from dimagi.utils.web import json_handler
            from couchexport.export import Constant
            if isinstance(obj, Constant):
                return obj.message
            else:
                return json_handler(obj)

    def _close(self):
        new_tables = {}
        for tablename, data in self.tables.items():
            new_tables[tablename] = {"headers":data[0], "rows": data[1:]}
    
        self.file.write(json.dumps(new_tables, cls=self.ConstantEncoder))

        
class HtmlExportWriter(InMemoryExportWriter):
    """
    Write tables to HTML
    """
    
    def _close(self):
        self.file.write(render_to_string("couchexport/html_export.html", {'tables': self.tables}))
        
        
