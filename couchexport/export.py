from couchdbkit.schema.base import DocumentBase
from couchexport.schema import get_docs, get_schema, get_schema_new
import csv
import json
import zipfile
from StringIO import StringIO
from dimagi.utils.web import json_handler
from django.conf import settings
from couchexport.models import ExportSchema, Format
import logging
import re
from dimagi.utils.mixins import UnicodeMixIn
from django.template.loader import render_to_string
from couchdbkit.consumer import Consumer
from dimagi.utils.couch.database import get_db
from couchexport.writers import Excel2007ExportWriter, CsvExportWriter,\
    Excel2003ExportWriter, JsonExportWriter, HtmlExportWriter
from soil import DownloadBase

class ExportConfiguration(object):
    """
    A representation of the configuration parameters for an export and 
    some functions to actually facilitate the export from this config.
    """
    
    def __init__(self, database, schema_index, previous_export=None, filter=None):
        self.database = database
        if len(schema_index) > 2:
            schema_index = schema_index[0:2]
        self.schema_index = schema_index
        self.previous_export = previous_export
        self.filter = filter
        self.current_seq = self.database.info()["update_seq"]
        self.potentially_relevant_ids = self._potentially_relevant_ids()
        
    def include(self, document):
        """
        Returns True if the document should be included in the results,
        otherwise false
        """
        return self.filter(document) if self.filter else True
    
    def _all_ids(self):
        """
        Gets view results for all documents matching this schema
        """
        return [result['id'] for result in \
                self.database.view("couchexport/schema_index", 
                                   key=self.schema_index).all()]
    
    def _potentially_relevant_ids(self):
        if self.previous_export is not None:
            consumer = Consumer(self.database)
            view_results = consumer.fetch(since=self.previous_export.seq)
            if view_results:
                try:
                    include_ids = set([res["id"] for res in view_results["results"]])
                    possible_ids = set(self._all_ids())
                    return list(include_ids.intersection(possible_ids))
                except TypeError, e:
                    if "string indices must be integers" in str(e):
                        # this is our expected error use case. 
                        raise Exception("Got the string integer thing again during export")
            else:
                # sometimes this comes back empty. I think it might be a bug
                # in couchdbkit, but it's impossible to consistently reproduce.
                # For now, just assume this is fine.
                return []
        else:
            return self._all_ids()
    
    def enum_docs(self):
        for i, doc_id in enumerate(self.potentially_relevant_ids):
            doc = self.database.get(doc_id)
            if self.include(doc):
                yield i, doc

    def get_docs(self):
        for _, doc in self.enum_docs():
            yield doc

    def last_checkpoint(self):
        return self.previous_export or ExportSchema.last(self.schema_index)

class UnsupportedExportFormat(Exception):
    pass

def get_writer(format):
    if format == Format.CSV:
        return CsvExportWriter()
    elif format == Format.HTML:
        return HtmlExportWriter()
    elif format == Format.JSON:
        return JsonExportWriter()
    elif format == Format.XLS:
        return Excel2003ExportWriter()
    elif format == Format.XLS_2007:
        return Excel2007ExportWriter()
    else:
        raise UnsupportedExportFormat("Unsupported export format: %s!" % format)
        
def export_from_tables(tables, file, format, max_column_size=2000):
    if format == Format.CSV:
        _export_csv(tables, file, max_column_size)
    elif format == Format.HTML:
        _export_html(tables, file, max_column_size)
    elif format == Format.JSON:
        _export_json(tables, file)
    elif format == Format.XLS:
        _export_excel(tables, max_column_size).save(file)
    elif format == Format.XLS_2007:
        _export_excel_2007(tables, max_column_size).save(file)
    else:
        raise Exception("Unsupported export format: %s!" % format)
    

def export_raw(headers, data, file, format=Format.XLS_2007, 
               max_column_size=2000, separator='|'):
    """
    Do a raw export from an in-memory representation of headers and data.
    Headers should be a list of (tablename, table) tuples with only one
    row (containing the headers) in the table.
    
    data_table should have the same format but can support multiple rows
    per table if needed.
    
    Example:
    
    headers:
     (("employee", ("id", "name", "gender")),
      ("building", ("id", "name", "address")))
    
    data:
     (("employee", (("1", "cory", "m"),
                    ("2", "christian", "m"),
                    ("3", "amelia", "f"))),
      ("building", (("1", "dimagi", "585 mass ave."),
                    ("2", "old dimagi", "529 main st."))))
    
    """
    # transform docs onto output and save
    writer = get_writer(format)
    
    
    # format the headers the way the export likes them
    headers = map(lambda table_headers: (table_headers[0], 
                                         [FormattedRow(table_headers[1])]),
                  headers)
    writer.open(headers, file)
    
    # do the same for the data
    data = map(lambda table_data: (table_data[0],
                                   [FormattedRow(row) for row in table_data[1]]),
               data)
    writer.write(data)
    writer.close()
    
def export(schema_index, file, format=Format.XLS_2007, 
           previous_export_id=None, filter=None,
           max_column_size=2000, separator='|', export_object=None, process=None):
    """
    Exports data from couch documents matching a given tag to a file. 
    Returns true if it finds data, otherwise nothing
    """

    DownloadBase.set_progress(process, 0, 1)

    config, updated_schema, export_schema_checkpoint = get_export_components(schema_index,
                                                                    previous_export_id, filter)
    # transform docs onto output and save
    if config:
        writer = get_writer(format)

        # open the doc and the headers
        formatted_headers = get_headers(updated_schema, separator=separator)
        writer.open(formatted_headers, file)

        total_docs = len(config.potentially_relevant_ids)
        if process:
            DownloadBase.set_progress(process, 0, total_docs)
        for i, doc in config.enum_docs():
            if export_object and export_object.transform:
                doc = export_object.transform(doc)
            writer.write(format_tables(create_intermediate_tables(doc, updated_schema),
                                       include_headers=False, separator=separator))
            if process:
                DownloadBase.set_progress(process, i + 1, total_docs)
        writer.close()
    return export_schema_checkpoint


def get_export_components(schema_index, previous_export_id=None, filter=None):
    """
    Get all the components needed to build an export file.
    """

    previous_export = ExportSchema.get(previous_export_id)\
        if previous_export_id else None
    database = get_db()
    config = ExportConfiguration(database, schema_index,
        previous_export, filter)

    # handle empty case
    if not config.potentially_relevant_ids:
        return None, None, None


    # get and checkpoint the latest schema
    updated_schema = get_schema_new(config)

    export_schema_checkpoint = ExportSchema(seq=config.current_seq,
        schema=updated_schema,
        index=config.schema_index)

    export_schema_checkpoint.save()

    return config, updated_schema, export_schema_checkpoint


class Constant(UnicodeMixIn):
    def __init__(self, message):
        self.message = message
    def __unicode__(self):
        return self.message

SCALAR_NEVER_WAS = settings.COUCHEXPORT_SCALAR_NEVER_WAS \
                    if hasattr(settings, "COUCHEXPORT_SCALAR_NEVER_WAS") \
                    else "---"

LIST_NEVER_WAS = settings.COUCHEXPORT_LIST_NEVER_WAS \
                    if hasattr(settings, "COUCHEXPORT_LIST_NEVER_WAS") \
                    else "this list never existed"

scalar_never_was = Constant(SCALAR_NEVER_WAS)
list_never_was = Constant(LIST_NEVER_WAS)

def render_never_was(schema):
    if isinstance(schema, dict):
        answ = {}
        for key in schema:
            answ[key] = render_never_was(schema[key])
        return answ

    elif isinstance(schema, list):
        return list_never_was
    else:
        return scalar_never_was

unknown_type = None
def fit_to_schema(doc, schema):

    def log(msg):
        raise Exception("doc-schema mismatch: %s" % msg)

    if schema is None:
        if doc:
            log("%s is not null" % doc)
        return None
    if isinstance(schema, list):
        if not doc:
            doc = []
        if not isinstance(doc, list):
            return fit_to_schema([doc], schema)
        answ = []
        schema_, = schema
        for doc_ in doc:
            answ.append(fit_to_schema(doc_, schema_))
        return answ
    if isinstance(schema, dict):
        if not doc:
            doc = {}
        if not isinstance(doc, dict):
            doc = {'': doc}
        doc_keys = set(doc.keys())
        schema_keys = set(schema.keys())
        if doc_keys - schema_keys:
            log("doc has keys not in schema: '%s'" % ("', '".join(doc_keys - schema_keys)))
        answ = {}
        for key in schema:
            #if schema[key] == unknown_type: continue
            if doc.has_key(key):
                answ[key] = fit_to_schema(doc.get(key), schema[key])
            else:
                answ[key] = render_never_was(schema[key])
        return answ
    if schema == "string":
        if not doc:
            doc = ""
        if not isinstance(doc, basestring):
        #log("%s is not a string" % doc)
            doc = unicode(doc)
        return doc


def get_headers(schema, separator="|"):
    return format_tables(create_intermediate_tables(schema, schema), 
                             include_data=False, separator=separator)

def create_intermediate_tables(docs, schema, integer='#'):
    def lookup(doc, keys):
        for key in keys:
            doc = doc[key]
        return doc
    def split_path(path):
        table = []
        column = []
        id = []
        for k in path:
            if isinstance(k, basestring):
                column.append(k)
            else:
                table.extend(column)
                table.append(integer)
                column = []
                id.append(k)
        return (tuple(table), tuple(column), tuple(id))
    schema = [schema]
    docs = fit_to_schema(docs, schema)
    # first, flatten documents
    queue = [()]
    leaves = []
    while queue:
        path = queue.pop()
        d = lookup(docs, path)
        if isinstance(d, dict):
            for key in d:
                queue.append(path + (key,))
        elif isinstance(d, list):
            for i,_ in enumerate(d):
                queue.append(path + (i,))
        elif d != list_never_was:
            leaves.append((split_path(path), d))
    leaves.sort()
    tables = {}
    for (table_name, column_name, id), val in leaves:
        table = tables.setdefault(table_name, {})
        row = table.setdefault(id, {})
        row[column_name] = val

    return tables

def nice(column_name):
    return "/".join(column_name)

class FormattedRow(object):
    """
    Simple data structure to represent a row of an export. Just 
    a pairing of an id and the data.
    
    The id should be an iterable (compound ids are supported). 
    """
    def __init__(self, data, id=None, separator=".", id_index=0):
        self.data = data
        self.id = id
        self.separator = separator
        self.id_index = id_index
    
    def has_id(self):
        return self.id is not None
    
    @property
    def formatted_id(self):
        if isinstance(self.id, basestring):
            return self.id
        return self.separator.join(map(unicode, self.id))
    
    def get_data(self):
        for i, val in enumerate(self.data):
            if self.has_id() and i == self.id_index:
                yield self.formatted_id
            yield val
        # if the ID is last the condition never gets triggerred
        # during the loop, so do it here
        if self.has_id() and self.id_index >= len(self.data):
            yield self.formatted_id
        
def format_tables(tables, id_label='id', separator='.', include_headers=True,
                      include_data=True):
    answ = []
    assert include_data or include_headers, "This method is pretty useless if you don't include anything!"
    
    for table_name, table in sorted(tables.items()):
        new_table = []
        keys = sorted(table.items()[0][1].keys()) # the keys for every row are the same
        
        if include_headers:
            header_vals = [separator.join(key) for key in keys]
            new_table.append(FormattedRow(header_vals, [id_label], separator))
        
        if include_data:
            for id, row in sorted(table.items()):
                values = [row[key] for key in keys]
                new_table.append(FormattedRow(values, id, separator))
        
        answ.append((separator.join(table_name), new_table))
    return answ

def _export_csv(tables, file, max_column_size):
    #temp = tempfile.TemporaryFile()
    temp = file
    archive = zipfile.ZipFile(temp, 'w', zipfile.ZIP_DEFLATED)
    
    # write forms
    used_names = []
    for table_name, table in tables:
        used_headers = []
        table_name_truncated = _clean_name(_next_unique(table_name, used_names))
        used_names.append(table_name_truncated)
        table_file = StringIO()
        writer = csv.writer(table_file, dialect=csv.excel)
        def _truncate(val):
            ret = _next_unique(val, used_headers, max_column_size)
            used_headers.append(ret)
            return ret
        
        for rowcount, row in enumerate(table):
            if rowcount == 0:
                # make sure we trim the headers
                row = map(_truncate, row)
            for i, val in enumerate(row):
                if isinstance(val, unicode):
                    row[i] = val.encode("utf8")
            writer.writerow(row)
        archive.writestr("%s.csv" % table_name_truncated, table_file.getvalue())
        
    archive.close()
    temp.seek(0)
    return temp


def _export_html(tables, file, max_column_size):
    file.write(render_to_string("couchexport/html_export.html", {'tables': tables}))

class ConstantEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Constant):
            return obj.message
        else:
            return json_handler(obj)

def _export_json(tables, file):
    new_tables = dict()
    for table in tables:
        new_tables[table[0]] = {"headers":table[1][0], "rows": table[1][1:]}

    file.write(json.dumps(new_tables, cls=ConstantEncoder))

def _export_excel(tables, max_column_size):
    try:
        import xlwt
    except ImportError:
        raise Exception("It doesn't look like this machine is configured for "
                        "excel export. To export to excel you have to run the "
                        "command:  easy_install xlutils")
    book = xlwt.Workbook()
    used_names = []
    for table_name, table in tables:
        used_headers = []
        def _truncate(val):
            ret = _next_unique(val, used_headers, max_column_size)
            used_headers.append(ret)
            return ret

        # this is in case the first 20 characters are the same, but we    
        # should do something smarter.    
        table_name_truncated = _next_unique(table_name, used_names, 20)
        used_names.append(table_name_truncated)
        sheet = book.add_sheet(_clean_name(table_name_truncated))
        
        for i,row in enumerate(table):
            if i == 0:
                # make sure we trim the headers
                row = map(_truncate, row)
            for j,val in enumerate(row):
                sheet.write(i,j,unicode(val))
    return book

def _export_excel_2007(tables, max_column_size):
    try:
        import openpyxl
    except ImportError:
        raise Exception("It doesn't look like this machine is configured for "
                        "excel export. To export to excel you have to run the "
                        "command:  easy_install openpyxl")
    book = openpyxl.workbook.Workbook()
    book.remove_sheet(book.worksheets[0])
    used_names = []
    for table_name, table in tables:
        used_headers = []
        def _truncate(val):
            ret = _next_unique(val, used_headers, max_column_size)
            used_headers.append(ret)
            return ret

        # this is in case the first 20 characters are the same, but we    
        # should do something smarter.    
        table_name_truncated = _next_unique(table_name, used_names, 31)
        used_names.append(table_name_truncated)
        sheet = book.create_sheet()
        sheet.title = _clean_name(table_name_truncated)
        for i,row in enumerate(table):
            if i == 0:
                # make sure we trim the headers
                row = map(_truncate, row)
            # the docs claim this should work but the source claims it doesn't 
            #sheet.append(row) 
            for j,val in enumerate(row):
                sheet.cell(row=i,column=j).value = unicode(val)
    return book


def _clean_name(name):
    return re.sub(r"[[\\?*/:\]]", "-", name)
    
def _next_unique(string, reserved_strings, max_len=500):
    counter = 1
    if len(string) > max_len:
        # truncate from the beginning since the end has more specific information
        string = string[-max_len:] 
    orig_string = string
    while string in reserved_strings:
        string = "%s%s" % (orig_string, counter)
        if len(string) > max_len:
            counterlen = len(str(counter))
            string = "%s%s" % (orig_string[-(max_len - counterlen):], counter)
        counter = counter + 1
    return string
