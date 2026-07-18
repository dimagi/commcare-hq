"""
One-off script for SAAS-20053: permanently delete (retire) deactivated mobile
workers -- including those with form submissions -- from a training domain.

retire() is the same operation as the single-user "Delete" button in the UI:
it soft-deletes the user's forms and cases (asynchronously) and soft-deletes
the user document.

Reads the list of usernames from the xlsx attached to the ticket (a single
column with a "username" header, e.g. 0002056027085).

Run in the Django shell on the target environment:

    cchq <env> django-manage shell < scratch_retire_ndoh_wbot_training.py

Run once with DRY_RUN = True, eyeball the counts, then flip DRY_RUN = False and
run it again.
"""
import time

from openpyxl import load_workbook

from dimagi.utils.chunked import chunked

from corehq.apps.users.dbaccessors import get_user_docs_by_username
from corehq.apps.users.models import CommCareUser, CouchUser
from corehq.apps.users.util import SYSTEM_USER_ID, format_username
from corehq.const import USER_CHANGE_VIA_SYSTEM

# ------------------------------- config -------------------------------
DOMAIN = 'ndoh-wbot-training'
USERNAMES_FILE = '/tmp/saas_20053_users.xlsx'  # xlsx with a "username" column
DRY_RUN = True                     # flip to False to actually delete
CHUNK_SIZE = 100                   # users retired between pauses
PAUSE_SECONDS = 5                  # throttle so form/case tasks don't flood
DELETED_BY = SYSTEM_USER_ID 
DELETED_VIA = USER_CHANGE_VIA_SYSTEM
# ----------------------------------------------------------------------


def load_usernames(path):
    """Read the first column of the xlsx, skipping the header row."""
    wb = load_workbook(path, read_only=True, data_only=True)
    usernames = []
    for (value,) in wb.active.iter_rows(min_col=1, max_col=1, values_only=True):
        if value is None:
            continue
        value = str(value).strip()
        if not value or value.lower() == 'username':
            continue
        usernames.append(value)
    wb.close()
    return list(dict.fromkeys(usernames))  # de-dupe, preserve order


def resolve_users(domain, usernames):
    """Return [(username, CommCareUser or None), ...]."""
    # Accept both bare ("0002056027085") and fully-qualified usernames.
    formatted = {
        (u.lower() if '@' in u else format_username(u, domain)): u
        for u in usernames
    }
    by_username = {}
    for chunk in chunked(list(formatted), 100):
        for doc in get_user_docs_by_username(chunk):
            by_username[doc['username']] = CouchUser.wrap_correctly(doc)
    return [(raw, by_username.get(full)) for full, raw in formatted.items()]


def classify(user, domain):
    if user is None:
        return 'not_found'
    if not isinstance(user, CommCareUser) or user.domain != domain:
        return 'wrong_domain'
    if user.is_deleted():
        return 'already_deleted'
    if user.is_active_in_domain(domain):
        return 'still_active'   # skipped: we only delete deactivated users
    return 'to_retire'


usernames = load_usernames(USERNAMES_FILE)
resolved = resolve_users(DOMAIN, usernames)

buckets = {}
for username, user in resolved:
    buckets.setdefault(classify(user, DOMAIN), []).append((username, user))

print(f"\n{len(usernames)} username(s) in {DOMAIN}:")
for status in ('to_retire', 'still_active', 'already_deleted', 'wrong_domain', 'not_found'):
    entries = buckets.get(status, [])
    print(f"  {status:16} {len(entries)}")
    if status != 'to_retire':
        for username, _user in entries[:20]:
            print(f"      {username}")
        if len(entries) > 20:
            print(f"      ... and {len(entries) - 20} more")

to_retire = buckets.get('to_retire', [])

if DRY_RUN:
    print(f"\nDRY RUN: would retire {len(to_retire)} user(s). No changes made.")
else:
    print(f"\nRetiring {len(to_retire)} user(s) from {DOMAIN}...")
    done = 0
    for chunk in chunked(to_retire, CHUNK_SIZE):
        for username, user in chunk:
            user.retire(DOMAIN, deleted_by=DELETED_BY, deleted_via=DELETED_VIA)
            done += 1
        print(f"  {done}/{len(to_retire)}")
        if done < len(to_retire):
            time.sleep(PAUSE_SECONDS)
    print(f"\nDone. Retired {done} user(s).")
