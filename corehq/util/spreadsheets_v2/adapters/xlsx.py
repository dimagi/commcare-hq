from contextlib import contextmanager
from zipfile import BadZipfile
from datetime import datetime, time
import openpyxl
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException
from corehq.util.spreadsheets_v2 import Worksheet, Cell, Workbook, SpreadsheetFileError


@contextmanager
def open_xlsx_workbook(filename):
    with open(filename) as f:
        try:
            openpyxl_workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except (BadZipfile, InvalidFileException) as e:
            raise SpreadsheetFileError(e.message)
        yield _XLSXWorkbookAdaptor(openpyxl_workbook).to_workbook()


class _XLSXWorksheetAdaptor(object):
    def __init__(self, openpyxl_worksheet):
        self._worksheet = openpyxl_worksheet

    def _make_cell_value(self, cell):
        if isinstance(cell.value, datetime):
            if cell._value == 0 and from_excel(0) == datetime(1899, 12, 30, 0, 0):
                # openpyxl has a bug that treats '12:00:00 AM'
                # as 0 seconds from the 'Windows Epoch' of 1899-12-30
                return time(0, 0)
            elif cell.value.time() == time(0, 0):
                return cell.value.date()
            else:
                return cell.value
        return cell.value

    def iter_rows(self):
        for row in self._worksheet.iter_rows():
            yield [Cell(self._make_cell_value(cell)) for cell in row]

    def to_worksheet(self):
        return Worksheet(title=self._worksheet.title, max_row=self._worksheet.max_row,
                         iter_rows=self.iter_rows)


class _XLSXWorkbookAdaptor(object):

    def __init__(self, openpyxl_workbook):
        self._workbook = openpyxl_workbook

    def to_workbook(self):
        return Workbook(
            worksheets=[_XLSXWorksheetAdaptor(worksheet).to_worksheet()
                        for worksheet in self._workbook.worksheets]
        )
