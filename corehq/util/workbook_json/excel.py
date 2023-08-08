import io
from zipfile import BadZipfile
from tempfile import NamedTemporaryFile
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from django.core.files.uploadedfile import UploadedFile
from django.utils.translation import gettext as _


class InvalidExcelFileException(Exception):
    pass


class JSONReaderError(Exception):
    pass


class HeaderValueError(Exception):
    pass


class StringTypeRequiredError(Exception):
    pass


class WorkbookJSONError(Exception):
    pass


class IteratorJSONReader(object):
    """
    >>> def normalize(it):
    ...     r = []
    ...     for row in IteratorJSONReader(it):
    ...         r.append(sorted(row.items()))
    ...     return r
    >>> normalize([])
    []
    >>> normalize([['A', 'B', 'C'], ['1', '2', '3']])
    [[('A', '1'), ('B', '2'), ('C', '3')]]
    >>> normalize([['A', 'data: key', 'user 1', 'user 2', 'is-ok?'],
    ...     ['1', '2', '3', '4', 'yes']])
    [[('A', '1'), ('data', {'key': '2'}), ('is-ok', True), ('user', ['3', '4'])]]
    """

    def __init__(self, rows):
        # you can only call __iter__ once
        self._rows = iter(rows)
        try:
            self.headers = list(next(self._rows))
        except StopIteration:
            self.headers = []
        self.fieldnames = self.get_fieldnames()

    def row_to_json(self, row):
        obj = {}
        for value, header in zip(row, self.headers):
            self.set_field_value(obj, header, value)
        return obj

    def __iter__(self):
        try:
            for row in self._rows:
                yield self.row_to_json(row)
        finally:
            del self._rows

    def get_fieldnames(self):
        obj = {}
        for field, value in zip(self.headers, [''] * len(self.headers)):
            if not isinstance(field, str):
                raise HeaderValueError('Field %s is not a string.' % field)
            self.set_field_value(obj, field, value)
        return list(obj)

    @classmethod
    def set_field_value(cls, obj, field, value):
        if isinstance(value, bytes):
            value = value.decode('utf-8')
        if isinstance(value, str):
            value = value.strip()
        # try dict
        try:
            field, subfield = field.split(':')
        except Exception:
            pass
        else:
            field = field.strip()

            if field not in obj:
                obj[field] = {}

            cls.set_field_value(obj[field], subfield, value)
            return

        # try list
        try:
            field, _ = field.split()
        except Exception:
            pass
        else:
            dud = {}
            cls.set_field_value(dud, field, value)
            (field, value), = list(dud.items())

            if field not in obj:
                obj[field] = []
            elif not isinstance(obj[field], list):
                obj[field] = [obj[field]]
            if value not in (None, ''):
                obj[field].append(value)
            return

        # else flat

        # try boolean
        try:
            field, nothing = field.split('?')
            assert(nothing.strip() == '')
        except Exception:
            pass
        else:
            try:
                value = {
                    'yes': True,
                    'true': True,
                    'no': False,
                    'false': False,
                    '': False,
                    None: False,
                }[value.lower() if hasattr(value, 'lower') else value]
            except KeyError:
                raise JSONReaderError(
                    'Values for field %s must be "yes" or "no", not "%s"' % (
                        field, value)
                )

        # set for any flat type
        field = field.strip()
        if field in obj:
            raise JSONReaderError(
                'You have a repeat field: %s' % field
            )

        obj[field] = value


def get_workbook(file_or_filename):
    try:
        return WorkbookJSONReader(file_or_filename)
    except (HeaderValueError, InvalidExcelFileException) as e:
        raise WorkbookJSONError(_(
            "Upload failed! "
            "Please make sure you are using a valid Excel 2007 or later (.xlsx) file. "
            "Error details: {}."
        ).format(e))
    except JSONReaderError as e:
        raise WorkbookJSONError(_(
            "Upload failed due to a problem with Excel columns. Error details: {}."
        ).format(e))
    except HeaderValueError as e:
        raise WorkbookJSONError(_(
            "Upload encountered a data type error: {}."
        ).format(e))
    except AttributeError as e:
        raise WorkbookJSONError(_(
            "Error processing Excel file: {}."
        ).format(e))


def get_single_worksheet(file_or_filename, title=None):
    workbook = get_workbook(file_or_filename)

    try:
        worksheet = workbook.get_worksheet(title=title)
    except WorksheetNotFound:
        raise WorkbookJSONError(_(
            "Could not find sheet '{title}'."
        ).format(title=title) if title else _("Uploaded file does not contian any sheets."))

    return worksheet


class WorksheetNotFound(Exception):

    def __init__(self, title):
        self.title = title
        super(WorksheetNotFound, self).__init__()


class WorksheetJSONReader(IteratorJSONReader):

    def __init__(self, worksheet, title=None):
        width = 0
        self.title = title
        self.worksheet = worksheet
        try:
            header_row = next(self.worksheet.iter_rows())
        except StopIteration:
            header_row = []
        for cell in header_row:
            if cell.value is None:
                break
            else:
                width += 1
        self.worksheet.calculate_dimension(force=True)

        def iterator():
            def _convert_float(value):
                """
                excel doesn't distinguish between 1 and 1.0
                if it can be an integer assume it is
                """
                if isinstance(value, float) and int(value) == value:
                    return int(value)
                else:
                    # Specifically check for None so that we can allow a value of 0
                    return value if value is not None else ''
            for row in self.worksheet.iter_rows():
                cell_values = [
                    _convert_float(cell.value)
                    for cell in row[:width]
                ]
                if not any(cell != '' for cell in cell_values):
                    break
                yield cell_values
        super(WorksheetJSONReader, self).__init__(iterator())


class WorkbookJSONReader(object):

    def __init__(self, file_or_filename):
        check_types = (UploadedFile, io.RawIOBase, io.BufferedIOBase)
        if isinstance(file_or_filename, check_types):
            tmp = NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
            file_or_filename.seek(0)
            tmp.write(file_or_filename.read())
            file_or_filename.seek(0)
            tmp.close()
            file_or_filename = tmp.name
        try:
            self.wb = openpyxl.load_workbook(file_or_filename, read_only=True, data_only=True)
        except (BadZipfile, InvalidFileException, KeyError) as e:
            raise InvalidExcelFileException(str(e))
        self.worksheets_by_title = {}
        self.worksheets = []

        for worksheet in self.wb.worksheets:
            try:
                ws = WorksheetJSONReader(worksheet, title=worksheet.title)
            except IndexError:
                raise JSONReaderError('This Excel file has unrecognised formatting. Please try downloading '
                                      'the lookup table first, and then add data to it.')
            self.worksheets_by_title[worksheet.title] = ws
            self.worksheets.append(ws)

    def get_worksheet(self, title=None, index=None):
        if title is not None and index is not None:
            raise TypeError("Can only get worksheet by title *or* index")
        if title:
            try:
                return self.worksheets_by_title[title]
            except KeyError:
                raise WorksheetNotFound(title=title)
        elif index:
            try:
                return self.worksheets[index]
            except IndexError:
                raise WorksheetNotFound(title=index)
        else:
            try:
                return self.worksheets[0]
            except IndexError:
                raise WorksheetNotFound(title=0)


def flatten_json_to_path(obj, path=()):
    if isinstance(obj, dict):
        for key, value in obj.items():
            for item in flatten_json_to_path(value, path + (key,)):
                yield item
    elif isinstance(obj, list):
        for key, value in enumerate(obj):
            for item in flatten_json_to_path(value, path + (key,)):
                yield item
    else:
        yield (path, obj)


def format_header(path, value):
    # pretty sure making a string-builder would be slower than concatenation
    s = path[0]
    for p in path[1:]:
        if isinstance(p, str):
            s += f': {p}'
        elif isinstance(p, int):
            s += f' {p + 1}'
    if isinstance(value, bool):
        s += '?'
        value = 'yes' if value else 'no'
    return s, value


def flatten_json(obj):
    for key, value in flatten_json_to_path(obj):
        yield format_header(key, value)


def json_to_headers(obj):
    return [key for key, value in sorted(flatten_json(obj), key=lambda t: alphanumeric_sort_key(t[0]))]


def alphanumeric_sort_key(key):
    """
    Sort the given iterable in the way that humans expect.
    Thanks to http://stackoverflow.com/a/2669120/240553
    """
    import re

    def convert(text):
        return int(text) if text.isdigit() else text

    return [convert(c) for c in re.split('([0-9]+)', key)]


def enforce_string_type(value):
    if isinstance(value, str):
        return value

    if isinstance(value, int):
        return str(value)

    # Don't try to guess for decimal types how they should be converted to string
    raise StringTypeRequiredError()
