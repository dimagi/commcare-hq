from tempfile import NamedTemporaryFile
import openpyxl


class JSONReaderError(Exception):
    pass


class HeaderValueError(Exception):
    pass


class IteratorJSONReader(object):
    """
    >>> def normalize(it):
    ...     r = []
    ...     for row in IteratorJSONReader(it):
    ...         r.append(sorted(row.items()))
    ...     return r
    >>> normalize([])
    []
    >>> normalize([['A', 'B', 'C'], ['1', '2', '3']])
    [[('A', '1'), ('B', '2'), ('C', '3')]]
    >>> normalize([['A', 'data: key', 'user 1', 'user 2', 'is-ok?'],
    ...     ['1', '2', '3', '4', 'yes']])
    [[('A', '1'), ('data', {'key': '2'}), ('is-ok', True), ('user', ['3', '4'])]]
    """

    def __init__(self, rows):
        # you can only call __iter__ once
        self._rows = iter(rows)
        try:
            self.headers = list(self._rows.next())
        except StopIteration:
            self.headers = []
        self.fieldnames = self.get_fieldnames()

    def row_to_json(self, row):
        obj = {}
        for value, header in zip(row, self.headers):
            self.set_field_value(obj, header, value)
        return obj

    def __iter__(self):
        try:
            for row in self._rows:
                yield self.row_to_json(row)
        finally:
            del self._rows

    def get_fieldnames(self):
        obj = {}
        for field, value in zip(self.headers, [''] * len(self.headers)):
            if not isinstance(field, basestring):
                raise HeaderValueError(u'Field %s is not a string.' % field)
            self.set_field_value(obj, field, value)
        return obj.keys()

    @classmethod
    def set_field_value(cls, obj, field, value):
        if isinstance(value, basestring):
            value = value.strip()
        # try dict
        try:
            field, subfield = field.split(':')
        except Exception:
            pass
        else:
            field = field.strip()

            if field not in obj:
                obj[field] = {}

            cls.set_field_value(obj[field], subfield, value)
            return

        # try list
        try:
            field, _ = field.split()
        except Exception:
            pass
        else:
            dud = {}
            cls.set_field_value(dud, field, value)
            (field, value), = dud.items()

            if field not in obj:
                obj[field] = []
            if value not in (None, ''):
                obj[field].append(value)
            return

        # else flat

        # try boolean
        try:
            field, nothing = field.split('?')
            assert(nothing.strip() == '')
        except Exception:
            pass
        else:
            try:
                value = {
                    'yes': True,
                    'true': True,
                    'no': False,
                    'false': False,
                    '': False,
                    None: False,
                }[value.lower() if hasattr(value, 'lower') else value]
            except KeyError:
                raise JSONReaderError(
                    'Values for field %s must be "yes" or "no", not "%s"' % (
                        field, value)
                )

        # set for any flat type
        field = field.strip()
        if field in obj:
            raise JSONReaderError(
                'You have a repeat field: %s' % field
            )

        obj[field] = value


class WorksheetNotFound(Exception):
    def __init__(self, title):
        self.title = title
        super(WorksheetNotFound, self).__init__()


class WorksheetJSONReader(IteratorJSONReader):
    def __init__(self, worksheet, title=None):
        width = 0
        self.title = title
        self.worksheet = worksheet
        try:
            header_row = self.worksheet.iter_rows().next()
        except StopIteration:
            header_row = []
        for cell in header_row:
            if cell.value is None:
                break
            else:
                width += 1
        self.worksheet.calculate_dimension(force=True)

        def iterator():
            def _convert_float(value):
                """
                excel doesn't distinguish between 1 and 1.0
                if it can be an integer assume it is
                """
                if isinstance(value, float) and int(value) == value:
                    return int(value)
                else:
                    return value or ''
            for row in self.worksheet.iter_rows():
                cell_values = [
                    _convert_float(cell.value)
                    for cell in row[:width]
                ]
                if not any(cell_values):
                    break
                yield cell_values
        super(WorksheetJSONReader, self).__init__(iterator())


class WorkbookJSONReader(object):
    def __init__(self, f):
        if isinstance(f, basestring):
            filename = f
        elif not isinstance(f, file):
            tmp = NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
            filename = tmp.name
            tmp.write(f.read())
            tmp.close()
        else:
            filename = f

        self.wb = openpyxl.load_workbook(filename, use_iterators=True)
        self.worksheets_by_title = {}
        self.worksheets = []

        for worksheet in self.wb.worksheets:
            ws = WorksheetJSONReader(worksheet, title=worksheet.title)
            self.worksheets_by_title[worksheet.title] = ws
            self.worksheets.append(ws)

    def get_worksheet(self, title=None, index=None):
        if title is not None and index is not None:
            raise TypeError("Can only get worksheet by title *or* index")
        if title:
            try:
                return self.worksheets_by_title[title]
            except KeyError:
                raise WorksheetNotFound(title=title)
        elif index:
            try:
                return self.worksheets[index]
            except IndexError:
                raise WorksheetNotFound(title=index)
        else:
            try:
                return self.worksheets[0]
            except IndexError:
                raise WorksheetNotFound(title=0)

    def work_book_headers_as_tuples(self):
        """
        Returns raw sheet headers in following format

          (("employee", ("id", "name", "gender")),
           ("building", ("id", "name", "address")))
        """
        all_sheet_headers = []
        for sheet in self.worksheets:
            all_sheet_headers.append(
                (sheet.title, tuple(sheet.headers))
            )
        return tuple(all_sheet_headers)

    def work_book_data_as_tuples(self):
        """
        Note: This is useful to get xlsx file's data into easy readible format
        Exists only to migrate tests using xlsx files to use following format

        Returns raw sheet data in following format

        (("employee", (("1", "cory", "m"),
                       ("2", "christian", "m"),
                       ("3", "amelia", "f"))),
         ("building", (("1", "dimagi", "585 mass ave."),
                       ("2", "old dimagi", "529 main st."))))
        """
        all_sheet_data = []
        for sheet in self.worksheets:
            current_sheet_data = []
            for row in sheet:
                values = [row.get(header) for header in sheet.headers]
                current_sheet_data.append(tuple(values))
            all_sheet_data.append(
                (sheet.title, tuple(current_sheet_data))
            )
        return tuple(all_sheet_data)


def flatten_json_to_path(obj, path=()):
    if isinstance(obj, dict):
        for key, value in obj.items():
            for item in flatten_json_to_path(value, path + (key,)):
                yield item
    elif isinstance(obj, list):
        for key, value in enumerate(obj):
            for item in flatten_json_to_path(value, path + (key,)):
                yield item
    else:
        yield (path, obj)


def format_header(path, value):
    # pretty sure making a string-builder would be slower than concatenation
    s = path[0]
    for p in path[1:]:
        if isinstance(p, basestring):
            s += ': %s' % p
        elif isinstance(p, int):
            s += ' %s' % (p + 1)
    if isinstance(value, bool):
        s += '?'
        value = 'yes' if value else 'no'
    return s, value


def flatten_json(obj):
    for key, value in flatten_json_to_path(obj):
        yield format_header(key, value)


def json_to_headers(obj):
    return [key for key, value in sorted(flatten_json(obj), key=lambda t: alphanumeric_sort_key(t[0]))]


def alphanumeric_sort_key(key):
    """
    Sort the given iterable in the way that humans expect.
    Thanks to http://stackoverflow.com/a/2669120/240553
    """
    import re
    convert = lambda text: int(text) if text.isdigit() else text
    return [convert(c) for c in re.split('([0-9]+)', key)]
