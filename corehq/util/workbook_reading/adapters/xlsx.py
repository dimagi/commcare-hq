from contextlib import contextmanager
from datetime import datetime, time
from zipfile import BadZipfile

import openpyxl
from memoized import memoized
from openpyxl.utils.exceptions import InvalidFileException

from corehq.util.workbook_reading import (
    Cell,
    SpreadsheetFileEncrypted,
    SpreadsheetFileInvalidError,
    SpreadsheetFileNotFound,
    Workbook,
    Worksheet,
)

# Got this from running hexdump and then googling
# which matched something on https://en.wikipedia.org/wiki/List_of_file_signatures:
#     Compound File Binary Format,
#     a container format used for document by older versions of Microsoft Office.
#     It is however an open format used by other programs as well.
# Also checked that it's not used non-encrypted xlsx files, which are just .zip files
XLSX_ENCRYPTED_MARKER = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'


@contextmanager
def open_xlsx_workbook(filename):
    try:
        f = open(filename, 'rb')
    except IOError as e:
        raise SpreadsheetFileNotFound(e)

    with f as f:
        try:
            openpyxl_workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except InvalidFileException as e:
            raise SpreadsheetFileInvalidError(str(e))
        except BadZipfile as e:
            f.seek(0)
            if f.read(8) == XLSX_ENCRYPTED_MARKER:
                raise SpreadsheetFileEncrypted('Workbook is encrypted')
            else:
                raise SpreadsheetFileInvalidError(str(e))
        yield _XLSXWorkbookAdaptor(openpyxl_workbook).to_workbook()


class _XLSXWorksheetAdaptor(object):
    def __init__(self, openpyxl_worksheet):
        self._worksheet = openpyxl_worksheet

    def _make_cell_value(self, cell):
        if isinstance(cell.value, datetime) and cell.value.time() == time(0, 0):
            return cell.value.date()
        return cell.value

    @property
    @memoized
    def _max_row(self):
        # self._worksheet.max_row is the max_row with data OR WITH FORMATTING
        # That means that if formatting is applied, that will be the xlsx row limit
        # Note that this is 1-indexed for consistency with openpyxl
        MAX_BLANK_ROWS = 1000

        max_row = 1
        blank_rows = 0
        for i, row in enumerate(self._worksheet.values, 1):
            if any(v for v in row):
                max_row = i
                blank_rows = 0
            else:
                blank_rows += 1
            if blank_rows >= MAX_BLANK_ROWS:
                # if this threshold is reached, assume there is no more data in the sheet
                return max_row
        return max_row

    def iter_rows(self):
        for i, row in enumerate(self._worksheet.iter_rows(), 1):
            if i > self._max_row:
                break
            yield [Cell(self._make_cell_value(cell)) for cell in row]

    def to_worksheet(self):
        # Note that an empty sheet and a sheet with one row both have max_row = 1
        return Worksheet(title=self._worksheet.title, max_row=self._max_row,
                         iter_rows=self.iter_rows)


class _XLSXWorkbookAdaptor(object):

    def __init__(self, openpyxl_workbook):
        self._workbook = openpyxl_workbook

    def to_workbook(self):
        return Workbook(
            worksheets=[_XLSXWorksheetAdaptor(worksheet).to_worksheet()
                        for worksheet in self._workbook.worksheets]
        )
