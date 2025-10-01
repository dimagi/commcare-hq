import io
import os

from django.test import TestCase
from django.urls import reverse

import openpyxl
from bs4 import BeautifulSoup

from corehq import privileges
from corehq.apps.data_dictionary.models import (
    CaseProperty,
    CasePropertyAllowedValue,
    CaseType,
)
from corehq.apps.data_dictionary.views import (
    ALLOWED_VALUES_SHEET_SUFFIX,
    ExportDataDictionaryView,
    UploadDataDictionaryView,
)
from corehq.apps.domain.shortcuts import create_domain
from corehq.apps.users.models import WebUser
from corehq.util.test_utils import TestFileMixin, privilege_enabled


class DataDictionaryImportTest(TestCase, TestFileMixin):
    domain_name = 'data-dict-import-test'
    file_path = ('data',)
    root = os.path.dirname(__file__)

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.domain = create_domain(cls.domain_name)
        cls.username = f'{cls.domain_name}-user'
        cls.couch_user = WebUser.create(
            None, cls.username, "foobar", None, None,
            email=f'{cls.username}@example.com',
        )
        cls.couch_user.add_domain_membership(cls.domain_name, is_admin=True)
        cls.couch_user.save()

    @classmethod
    def tearDownClass(cls):
        cls.couch_user.delete(cls.domain_name, deleted_by=None)
        cls.domain.delete()
        super().tearDownClass()

    def setUp(self):
        self.url = reverse(UploadDataDictionaryView.urlname, args=[self.domain_name])
        self.client.login(username=self.username, password='foobar')


@privilege_enabled(privileges.DATA_DICTIONARY, privileges.DATA_DICT_TYPES)
class DataDictionaryImportTestWithDataTypes(DataDictionaryImportTest):
    domain_name = 'data-dict-import-test-with-data-types'

    def test_clean_import(self):
        fname = self.get_path('clean_data_dictionary', 'xlsx')
        with open(fname, 'rb') as dd_file:
            response = self.client.post(self.url, {'bulk_upload_file': dd_file})
        self.assertEqual(response.status_code, 200)
        messages = list(response.context['messages'])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), 'Data dictionary import complete')
        case_types = CaseType.objects.filter(domain=self.domain)
        self.assertEqual(2, case_types.count())
        for i in range(1, 3):
            case_type_name = f'caseType{i}'
            case_type = case_types.get(name=case_type_name)
            self.assertEqual(case_type.properties.count(), 3)
            for data_type in ['plain', 'date', 'select']:
                prop = case_type.properties.get(data_type=data_type, name=f'{data_type}prop')
                if data_type == 'select':
                    vals = ['True', 'False', '0', '1'] if i == 1 else ['Yes', 'No']
                    self.assertEqual(len(vals), prop.allowed_values.count())
                    for val in vals:
                        av_qs = prop.allowed_values.filter(
                            allowed_value=val, description=f'{val} description')
                        self.assertEqual(1, av_qs.count())

    def test_broken_import(self):
        fname = self.get_path('broken_data_dictionary', 'xlsx')
        with open(fname, 'rb') as dd_file:
            response = self.client.post(self.url, {'bulk_upload_file': dd_file})
        self.assertEqual(response.status_code, 200)
        messages = list(response.context['messages'])
        self.assertEqual(len(messages), 1)
        expected_errors = {
            'Row 4 in \"caseType1-vl\" sheet is missing a case property field',
            'Error in \"caseType1\" sheet, row 4: Unable to save valid values longer than 255 characters',
            'Missing required valid \"caseType2-vl\" multi-choice sheet for case type \"caseType2\"',
            'Case property \"mistake\" referenced in \"caseType1-vl\" sheet that does not exist in '
            '\"caseType1\" sheet. Row(s) affected: 5, 6',
            "Error in \"caseType1\" sheet, row 5: "
            "{'data_type': [\"Value 'Nonsense' is not a valid choice.\"]}",
            'Error in "caseType1" sheet, row 6: Invalid case property name. It should start with a '
            'letter, and only contain letters, numbers, "-", and "_"',
        }
        message_str = str(messages[0])
        soup = BeautifulSoup(message_str, features="lxml")
        received_errors = {elem.text for elem in soup.find_all('li')}
        self.assertEqual(expected_errors, received_errors)


@privilege_enabled(privileges.DATA_DICTIONARY)
class DataDictionaryImportTestWithoutDataTypes(DataDictionaryImportTest):
    domain_name = 'data-dict-import-test-without-data-types'

    def test_import_with_warnings(self):
        fname = self.get_path('clean_data_dictionary', 'xlsx')
        with open(fname, 'rb') as dd_file:
            response = self.client.post(self.url, {'bulk_upload_file': dd_file})
        self.assertEqual(response.status_code, 200)
        messages = list(response.context['messages'])
        self.assertEqual(len(messages), 2)

        soup = BeautifulSoup(str(messages[1]), features="lxml")
        received_warnings = {elem.text for elem in soup.find_all('li')}
        expected_warnings = {
            'The multi-choice sheets (ending in "-vl") in the uploaded Excel file were ignored '
            'during the import as they are an unsupported format in the Data Dictionary',
            'The "Data Type" column values were ignored during the import as they are '
            'an unsupported format in the Data Dictionary',
        }
        self.assertEqual(received_warnings, expected_warnings)


class DataDictionaryExportTest(DataDictionaryImportTest):
    domain_name = 'data-dict-export-test'

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        case_types = []
        for i in range(1, 3):
            case_type_obj = CaseType(name=f'caseType{i}', domain=cls.domain_name)
            case_type_obj.save()
            case_types.append(case_type_obj)
            for data_type in ['plain', 'date', 'select']:
                prop = CaseProperty.objects.create(
                    case_type=case_type_obj, name=f'{data_type}prop', data_type=data_type)
                if data_type == 'select':
                    vals = ['True', 'False'] if i == 1 else ['Yes', 'No']
                    for val in vals:
                        CasePropertyAllowedValue.objects.create(
                            case_property=prop, allowed_value=val,
                            description=f'{val} description')
        cls.case_types = case_types

    def setUp(self):
        self.url = reverse(ExportDataDictionaryView.urlname, args=[self.domain_name])
        self.client.login(username=self.username, password='foobar')


@privilege_enabled(privileges.DATA_DICTIONARY, privileges.DATA_DICT_TYPES)
class DataDictionaryExportTestWithDataTypes(DataDictionaryExportTest):
    domain_name = 'data-dict-export-test-with-data-types'

    def test_export(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['content-type'], 'application/vnd.ms-excel')
        self.assertEqual(response['Content-Disposition'], 'attachment; filename="data_dictionary.xlsx"')
        workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)
        self.assertEqual(len(workbook.sheetnames), 2 * len(self.case_types))
        for i in range(len(self.case_types)):
            name = self.case_types[i].name
            self.assertEqual(name, workbook.sheetnames[i * 2])
            self.assertEqual(f'{name}{ALLOWED_VALUES_SHEET_SUFFIX}', workbook.sheetnames[i * 2 + 1])


@privilege_enabled(privileges.DATA_DICTIONARY)
class DataDictionaryExportTestWithoutDataTypes(DataDictionaryExportTest):
    domain_name = 'data-dict-export-test-without-data-types'

    def test_export_without_data_type(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)
        self.assertEqual(len(workbook.sheetnames), len(self.case_types))
        for i in range(len(self.case_types)):
            name = self.case_types[i].name
            self.assertEqual(name, workbook.sheetnames[i])
            sheet = workbook[name]
            for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
                self.assertTrue(len(row), 5)
