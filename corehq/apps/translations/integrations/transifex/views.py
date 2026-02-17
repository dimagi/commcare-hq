from io import BytesIO, open
from zipfile import ZipFile

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils.translation import gettext as _
from django.utils.translation import gettext_noop

import openpyxl
import polib
from memoized import memoized

from couchexport.models import Format

from corehq.apps.domain.decorators import login_and_domain_required
from corehq.apps.domain.views.base import BaseDomainView
from corehq.apps.translations.forms import ConvertTranslationsForm
from corehq.apps.translations.generators import PoFileGenerator, Translation
from corehq.apps.translations.integrations.transifex.utils import (
    transifex_details_available_for_domain,
)
from corehq.apps.translations.models import TransifexBlacklist
from corehq.apps.translations.utils import get_file_content_from_workbook
from corehq.util.files import safe_filename_header


class BaseTranslationsView(BaseDomainView):
    section_name = gettext_noop("Translations")

    @property
    def page_context(self):
        context = {
            'transifex_details_available': self.transifex_details_available,
        }
        return context

    @property
    @memoized
    def transifex_details_available(self):
        return transifex_details_available_for_domain(self.domain)

    def transifex_integration_enabled(self, request):
        if not self.transifex_details_available:
            messages.error(request, _('Transifex integration not set for this domain'))
            return False
        return True


class ConvertTranslations(BaseTranslationsView):
    page_title = _('Convert Translations')
    urlname = 'convert_translations'
    template_name = 'translations/bootstrap3/convert_translations.html'

    @property
    @memoized
    def convert_translation_form(self):
        if self.request.POST:
            return ConvertTranslationsForm(self.request.POST, self.request.FILES)
        else:
            return ConvertTranslationsForm()

    @property
    @memoized
    def _uploaded_file_name(self):
        uploaded_file = self.convert_translation_form.cleaned_data.get('upload_file')
        return uploaded_file.name

    @staticmethod
    def _parse_excel_sheet(worksheet):
        """
        :return: the rows and the index for expected columns
        """
        rows = [row for row in worksheet.iter_rows()]
        headers = [cell.value for cell in rows[0]]
        source = headers.index('source')
        translation = headers.index('translation')
        occurrence = headers.index('occurrence') if 'occurrence' in headers else None
        context = headers.index('context') if 'context' in headers else None
        return rows, source, translation, occurrence, context

    def _generate_translations_for_po(self, worksheet):
        """
        :return: a list of Translation objects
        """
        rows, source, translation, occurrence, context = self._parse_excel_sheet(worksheet)
        translations = {worksheet.title: []}
        for row in rows[1:]:
            _occurrence = row[occurrence].value or '' if occurrence is not None else ''
            _context = row[context].value if context is not None else ''
            translations[worksheet.title].append(
                Translation(
                    row[source].value,
                    row[translation].value,
                    [(_occurrence, '')],
                    _context)
            )
        return translations

    def _generate_po_content(self, worksheet):
        """
        extract translations from worksheet and converts to a po file

        :return: content of file generated
        """
        translations = self._generate_translations_for_po(worksheet)
        with PoFileGenerator(translations, {}) as po_file_generator:
            generated_files = po_file_generator.generate_translation_files()
            with open(generated_files[0].path, 'rb') as f:
                return f.read()

    def _generate_excel_file(self, uploaded_file):
        """
        extract translations from po file and converts to a xlsx file

        :return: Workbook object
        """
        po_file = polib.pofile(uploaded_file.read().decode('utf-8'))
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = "Translations"
        ws.append(['source', 'translation', 'context', 'occurrence'])
        for po_entry in po_file:
            ws.append([po_entry.msgid, po_entry.msgstr, po_entry.msgctxt,
                       po_entry.occurrences[0][0] if po_entry.occurrences else ''])
        return wb

    def _po_file_response(self):
        uploaded_file = self.convert_translation_form.cleaned_data.get('upload_file')
        worksheet = openpyxl.load_workbook(uploaded_file).worksheets[0]
        content = self._generate_po_content(worksheet)
        response = HttpResponse(content, content_type="text/plain; charset=utf-8")
        response['Content-Disposition'] = safe_filename_header(worksheet.title, 'po')
        return response

    def _excel_file_response(self):
        wb = self._generate_excel_file(self.convert_translation_form.cleaned_data.get('upload_file'))
        content = get_file_content_from_workbook(wb)
        response = HttpResponse(content, content_type=Format.from_format('xlsx').mimetype)
        response['Content-Disposition'] = safe_filename_header(self._uploaded_file_name.split('.po')[0], 'xlsx')
        return response

    def _zip_file_response(self):
        uploaded_file = self.convert_translation_form.cleaned_data.get('upload_file')
        uploaded_zipfile = ZipFile(uploaded_file)
        mem_file = BytesIO()
        with ZipFile(mem_file, 'w') as zipfile:
            for file_info in uploaded_zipfile.filelist:
                filename = file_info.filename
                if filename.endswith('.po'):
                    po_file = BytesIO(uploaded_zipfile.read(filename))
                    wb = self._generate_excel_file(po_file)
                    result_filename = filename.split('.po')[0]
                    zipfile.writestr(result_filename + '.xlsx', get_file_content_from_workbook(wb))
                elif filename.endswith('.xls') or filename.endswith('.xlsx'):
                    worksheet = openpyxl.load_workbook(BytesIO(uploaded_zipfile.read(filename))).worksheets[0]
                    po_file_content = self._generate_po_content(worksheet)
                    result_filename = filename.split('.xls')[0]
                    zipfile.writestr(result_filename + '.po', po_file_content)
                else:
                    assert False, "unexpected filename: {}".format(filename)
        mem_file.seek(0)
        response = HttpResponse(mem_file, content_type='application/zip')
        zip_filename = 'Converted-' + uploaded_zipfile.filename.split('.zip')[0]
        response['Content-Disposition'] = safe_filename_header(zip_filename, "zip")
        return response

    def post(self, request, *args, **kwargs):
        if self.convert_translation_form.is_valid():
            uploaded_filename = self._uploaded_file_name
            if uploaded_filename.endswith('.xls') or uploaded_filename.endswith('.xlsx'):
                return self._po_file_response()
            elif uploaded_filename.endswith('.po'):
                return self._excel_file_response()
            elif uploaded_filename.endswith('.zip'):
                return self._zip_file_response()
        return self.get(request, *args, **kwargs)

    def section_url(self):
        return self.page_url

    @property
    def page_context(self):
        context = super(ConvertTranslations, self).page_context
        context['convert_translations_form'] = self.convert_translation_form
        return context


@login_and_domain_required
def delete_translation_blacklist(request, domain, pk):
    TransifexBlacklist.objects.filter(domain=domain, pk=pk).delete()
    return redirect(ConvertTranslations.urlname, domain=domain)
