from __future__ import absolute_import
from __future__ import unicode_literals
import openpyxl
import polib

from io import open
from memoized import memoized
from openpyxl import Workbook

from django.http import HttpResponse
from django.utils.translation import ugettext as _, ugettext_noop
from django.contrib import messages
from django.utils.decorators import method_decorator

from corehq import toggles
from corehq.apps.hqwebapp.decorators import use_select2
from corehq.apps.translations.forms import (
    ConvertTranslationsForm,
    PullResourceForm,
)
from corehq.apps.app_manager.app_translations.generators import Translation, PoFileGenerator
from corehq.util.files import safe_filename_header
from corehq.apps.domain.views import BaseDomainView
from custom.icds.translations.integrations.exceptions import ResourceMissing
from custom.icds.translations.integrations.transifex import Transifex
from custom.icds.translations.integrations.utils import get_file_content_from_workbook, \
    transifex_details_available_for_domain


class BaseTranslationsView(BaseDomainView):
    @property
    def page_context(self):
        context = {
            'transifex_details_available': self.transifex_details_available,
        }
        return context

    @property
    @memoized
    def transifex_details_available(self):
        return transifex_details_available_for_domain(self.domain)

    def transifex_integration_enabled(self, request):
        if not self.transifex_details_available:
            messages.error(request, _('Transifex integration not set for this domain'))
            return False
        return True


class ConvertTranslations(BaseTranslationsView):
    page_title = _('Convert Translations')
    urlname = 'convert_translations'
    template_name = 'convert_translations.html'
    section_name = ugettext_noop("Translations")

    @property
    @memoized
    def convert_translation_form(self):
        if self.request.POST:
            return ConvertTranslationsForm(self.request.POST, self.request.FILES)
        else:
            return ConvertTranslationsForm()

    @property
    @memoized
    def _uploaded_file_name(self):
        uploaded_file = self.convert_translation_form.cleaned_data.get('upload_file')
        return uploaded_file.name

    @staticmethod
    def _parse_excel_sheet(worksheet):
        """
        :return: the rows and the index for expected columns
        """
        rows = [row for row in worksheet.iter_rows()]
        headers = [cell.value for cell in rows[0]]
        source = headers.index('source')
        translation = headers.index('translation')
        occurrence = headers.index('occurrence') if 'occurrence' in headers else None
        context = headers.index('context') if 'context' in headers else None
        return rows, source, translation, occurrence, context

    def _generate_translations_for_po(self, worksheet):
        """
        :return: a list of Translation objects
        """
        rows, source, translation, occurrence, context = self._parse_excel_sheet(worksheet)
        translations = {worksheet.title: []}
        for row in rows[1:]:
            _occurrence = row[occurrence].value if occurrence is not None else ''
            _context = row[context].value if context is not None else ''
            translations[worksheet.title].append(
                Translation(
                    row[source].value,
                    row[translation].value,
                    [(_occurrence, None)],
                    _context)
            )
        return translations

    def _generate_po_file(self, worksheet):
        """
        extract translations from worksheet and converts to a po file
        :return: list of files generated
        """
        translations = self._generate_translations_for_po(worksheet)
        with PoFileGenerator(translations, {}) as po_file_generator:
            return po_file_generator.generated_files[0][1]

    def _generate_excel_file(self):
        """
        extract translations from po file and converts to a xlsx file

        :return: Workbook object
        """
        uploaded_file = self.convert_translation_form.cleaned_data.get('upload_file')
        po_file = polib.pofile(uploaded_file.read())
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = "Translations"
        ws.append(['source', 'translation', 'context', 'occurrence'])
        for po_entry in po_file:
            ws.append([po_entry.msgid, po_entry.msgstr, po_entry.msgctxt,
                       po_entry.occurrences[0][0] if po_entry.occurrences else ''])
        return wb

    def _po_file_response(self):
        uploaded_file = self.convert_translation_form.cleaned_data.get('upload_file')
        worksheet = openpyxl.load_workbook(uploaded_file).worksheets[0]
        generated_file = self._generate_po_file(worksheet)
        with open(generated_file, 'r', encoding="utf-8") as f:
            content = f.read()
        response = HttpResponse(content, content_type="text/html; charset=utf-8")
        response['Content-Disposition'] = safe_filename_header(worksheet.title, 'po')
        return response

    def _excel_file_response(self):
        wb = self._generate_excel_file()
        content = get_file_content_from_workbook(wb)
        response = HttpResponse(content, content_type="text/html; charset=utf-8")
        response['Content-Disposition'] = safe_filename_header(self._uploaded_file_name.split('.po')[0], 'xlsx')
        return response

    def post(self, request, *args, **kwargs):
        if self.convert_translation_form.is_valid():
            uploaded_filename = self._uploaded_file_name
            if uploaded_filename.endswith('.xls') or uploaded_filename.endswith('.xlsx'):
                return self._po_file_response()
            elif uploaded_filename.endswith('.po'):
                return self._excel_file_response()
        return self.get(request, *args, **kwargs)

    def section_url(self):
        return self.page_url

    @property
    def page_context(self):
        context = super(ConvertTranslations, self).page_context
        context['convert_translations_form'] = self.convert_translation_form
        return context


@method_decorator([toggles.APP_TRANSLATIONS_WITH_TRANSIFEX.required_decorator()], name='dispatch')
class PullResource(BaseTranslationsView):
    page_title = _('Pull Resource')
    urlname = 'pull_resource'
    template_name = 'pull_resource.html'
    section_name = ugettext_noop("Translations")

    @use_select2
    def dispatch(self, request, *args, **kwargs):
        return super(PullResource, self).dispatch(request, *args, **kwargs)

    def section_url(self):
        return self.page_url

    @property
    @memoized
    def pull_resource_form(self):
        if self.request.POST:
            return PullResourceForm(self.domain, self.request.POST)
        else:
            return PullResourceForm(self.domain)

    @property
    def page_context(self):
        context = super(PullResource, self).page_context
        if context['transifex_details_available']:
            context['pull_resource_form'] = self.pull_resource_form
        return context

    def _generate_excel_file(self, domain, resource_slug):
        """
        extract translations from po file pulled from transifex and converts to a xlsx file

        :return: Workbook object
        """
        target_lang = self.pull_resource_form.cleaned_data['target_lang']
        transifex = Transifex(domain=domain, app_id=None,
                              source_lang=target_lang,
                              project_slug=self.pull_resource_form.cleaned_data['transifex_project_slug'],
                              version=None)
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title='translations')
        ws.append(['context', 'source', 'translation', 'occurrence'])
        for po_entry in transifex.client.get_translation(resource_slug, target_lang, False):
            ws.append([po_entry.msgctxt, po_entry.msgid, po_entry.msgstr,
                       po_entry.occurrences[0][0] if po_entry.occurrences else ''])
        return wb

    def _pull_resource(self, request):
        resource_slug = self.pull_resource_form.cleaned_data['resource_slug']
        wb = self._generate_excel_file(request.domain, resource_slug)
        content = get_file_content_from_workbook(wb)
        response = HttpResponse(content, content_type="text/html; charset=utf-8")
        response['Content-Disposition'] = safe_filename_header(resource_slug, 'xlsx')
        return response

    def post(self, request, *args, **kwargs):
        if self.transifex_integration_enabled(request):
            if self.pull_resource_form.is_valid():
                try:
                    return self._pull_resource(request)
                except ResourceMissing:
                    messages.add_message(request, messages.ERROR, 'Resource not found')
        return self.get(request, *args, **kwargs)
