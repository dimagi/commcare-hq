from django.core.management.base import BaseCommand

from openpyxl import load_workbook

from corehq.apps.locations.models import SQLLocation


class Command(BaseCommand):
    help = """
    Uses the Excel spreadsheet from "Download Organization Structure" to
    update the location IDs of locations. It uses "site_code" to match
    the locations in the spreadsheet with the locations in the database.
    """

    def add_arguments(self, parser):
        parser.add_argument("domain", type=str)
        parser.add_argument(
            "input_file",
            type=str,
            help="Path to the Excel file containing the location data.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Perform a dry run without making any changes.",
        )

    def handle(self, domain, input_file, *args, **options):
        dry_run = options["dry_run"]

        # Load the workbook and select the active worksheet
        workbook = load_workbook(filename=input_file)
        for sheet in workbook.worksheets:
            if sheet.title == "types":
                continue

            for row in sheet.iter_rows(min_row=2, values_only=True):
                location_id, site_code = row[0], row[1]
                try:
                    location = SQLLocation.objects.get(domain=domain, site_code=site_code)
                except SQLLocation.DoesNotExist:
                    self.stdout.write(f"Location with site code {site_code} does not exist.")
                if location.location_id != location_id:
                    self.stdout.write(f"Updating {location} from {location.location_id} to {location_id}")
                    if not dry_run:
                        location.location_id = location_id
                        location.save()
