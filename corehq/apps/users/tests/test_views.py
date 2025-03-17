import json
from contextlib import contextmanager
from copy import deepcopy
from io import BytesIO
from openpyxl import Workbook
from unittest.mock import patch, Mock
import re

from django.http import Http404, HttpResponseRedirect
from django.test import TestCase, Client
from django.test.client import RequestFactory
from django.urls import reverse

from corehq import privileges
from corehq.apps.accounting.utils import domain_has_privilege
from corehq.apps.domain.shortcuts import create_domain
from corehq.apps.es.tests.utils import es_test, populate_user_index
from corehq.apps.es.users import user_adapter
from corehq.apps.events.models import (
    AttendanceTrackingConfig,
    AttendeeModel,
)
from corehq.apps.hqcase.case_helper import CaseHelper
from corehq.apps.locations.models import LocationType
from corehq.apps.locations.tests.util import delete_all_locations, make_loc
from corehq.apps.user_importer.exceptions import UserUploadError
from corehq.apps.users.audit.change_messages import UserChangeMessage
from corehq.apps.users.dbaccessors import delete_all_users
from corehq.apps.users.exceptions import InvalidRequestException
from corehq.apps.users.models import (
    CommCareUser,
    CouchUser,
    HqPermissions,
    UserHistory,
    UserRole,
    WebUser, HQApiKey,
)
from corehq.apps.users.views import _delete_user_role, _update_role_from_view, BaseUploadUser
from corehq.apps.users.views.mobile.users import MobileWorkerListView
from corehq.const import USER_CHANGE_VIA_WEB
from corehq.util.test_utils import (
    flag_enabled,
    generate_cases,
    privilege_enabled,
)
from corehq.util.workbook_json.excel import WorkbookJSONError


def get_default_available_permissions(**kwargs):
    permissions = HqPermissions(**kwargs).to_json()
    permissions.pop('manage_domain_alerts')
    return permissions


class TestMobileWorkerListView(TestCase):
    domain = 'mobile-worker-list-view'
    web_username = 'test-webuser'
    password = '***'

    def setUp(self):
        super().setUp()
        self.project = create_domain(self.domain)
        self.web_user = WebUser.create(self.domain, self.web_username, self.password, None, None)

        # We aren't testing permissions for this test
        self.web_user.is_superuser = True
        self.web_user.save()

        self.role = UserRole.create(self.domain, 'default mobile use role', is_commcare_user_default=True)

    def tearDown(self):
        self.project.delete()
        delete_all_users()
        super().tearDown()

    def _remote_invoke(self, route, data):
        self.client.login(username=self.web_username, password=self.password)
        return self.client.post(
            reverse(MobileWorkerListView.urlname, args=[self.domain]),
            json.dumps(data),
            content_type='application/json;charset=UTF-8',
            HTTP_DJNG_REMOTE_METHOD=route,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

    def test_create_mobile_worker(self):
        resp = self._remote_invoke('create_mobile_worker', {
            "user": {
                "first_name": "Test",
                "last_name": "Test",
                "username": "test.test",
                "password": "123"
            }
        })
        content = json.loads(resp.content)
        self.assertEqual(content['success'], True)
        user = CouchUser.get_by_username(
            '{}@{}.commcarehq.org'.format(
                'test.test',
                self.domain,
            )
        )
        self.assertIsNotNone(user)
        self.assertEqual(user.get_role(self.domain).id, self.role.id)

    @flag_enabled('ATTENDANCE_TRACKING')
    @privilege_enabled(privileges.ATTENDANCE_TRACKING)
    def test_commcare_attendee_case_created(self):
        """An attendance tracking case should be created for a mobile worker on creation"""
        with self.enable_mobile_worker_attendees(), \
                self.get_mobile_worker() as user:
            self.assertAttendeeCreatedForUser(user)

    @flag_enabled('ATTENDANCE_TRACKING')
    def test_commcare_attendee_case_not_created_due_to_privilege(self):
        """This tests the case where a domain was on a higher plan and used the attendance tracking, but have
        downgraded ever since and now creates a new mobile worker"""
        with self.enable_mobile_worker_attendees(), \
                self.get_mobile_worker():
            self.assertAttendeeNotCreated()

    @flag_enabled('ATTENDANCE_TRACKING')
    @privilege_enabled(privileges.ATTENDANCE_TRACKING)
    def test_commcare_attendee_case_not_created_due_to_config(self):
        # AttendanceTrackingConfig does not exist for this domain
        self.assertNoAttendanceTrackingConfig()
        with self.get_mobile_worker():
            self.assertAttendeeNotCreated()

    def assertAttendeeCreatedForUser(self, user):
        models = AttendeeModel.objects.by_domain(self.domain)
        self.assertEqual(len(models), 1)
        self.assertEqual(models[0].user_id, user.user_id)

    def assertAttendeeNotCreated(self):
        self.assertEqual(AttendeeModel.objects.by_domain(self.domain), [])

    def assertNoAttendanceTrackingConfig(self):
        self.assertIsNone(
            AttendanceTrackingConfig.objects.filter(domain=self.domain).first()
        )

    @contextmanager
    def enable_mobile_worker_attendees(self):
        config, __ = AttendanceTrackingConfig.objects.update_or_create(
            domain=self.domain,
            defaults={'mobile_worker_attendees': True},
        )
        try:
            yield
        finally:
            config.delete()

    @contextmanager
    def get_mobile_worker(self):
        username = 'test.test'
        self._remote_invoke('create_mobile_worker', {
            "user": {
                "first_name": "Test",
                "last_name": "Test",
                "username": username,
                "password": "123"
            }
        })
        user = CouchUser.get_by_username(f'{username}@{self.domain}.commcarehq.org')
        try:
            yield user
        finally:
            close_user_attendee(self.domain, user.user_id)
            user.delete(None, None)


def close_user_attendee(domain, user_id):
    for model in AttendeeModel.objects.by_domain(domain):
        if model.user_id == user_id:
            helper = CaseHelper(case_id=model.case_id, domain=domain)
            helper.close()


@generate_cases((
    ('jmoney', False),
    ('jmoney91', False),
    ('j+money', False),
    ('j.money', False),
    ('j_money', False),
    ('j-money', False),
    ('j$', True),
    ('jmoney@something', True),
    ('jmoney...', True),
    ('.jmoney', True),
), TestMobileWorkerListView)
def test_check_usernames_for_invalid_characters(self, username, error):
    resp = self._remote_invoke('check_username', {
        'username': username
    })
    content = json.loads(resp.content)
    self.assertIs('error' in content, error)


class TestUpdateRoleFromView(TestCase):
    domain = "test_update_role"

    BASE_JSON = {
        'domain': domain,
        'name': None,
        'default_landing_page': 'webapps',
        'is_non_admin_editable': False,
        'is_archived': False,
        'upstream_id': None,
        'permissions': get_default_available_permissions(edit_web_users=True),
        'assignable_by': []
    }

    @classmethod
    def setUpTestData(cls):
        cls.role = UserRole(
            domain=cls.domain,
            name="role1",
        )
        cls.role.save()

    @classmethod
    def tearDownClass(cls):
        cls.role.delete()
        super().tearDownClass()

    def tearDown(self):
        for role in UserRole.objects.all():
            if role.id != self.role.id:
                role.delete()

    def test_create_role(self):
        role_data = deepcopy(self.BASE_JSON)
        role_data["name"] = "role2"
        role_data["assignable_by"] = [self.role.couch_id]
        role = _update_role_from_view(self.domain, role_data)
        self.assertEqual(role.name, "role2")
        self.assertEqual(role.default_landing_page, "webapps")
        self.assertFalse(role.is_non_admin_editable)
        self.assertEqual(role.assignable_by, [self.role.couch_id])
        self.assertEqual(role.permissions.to_json(), role_data['permissions'])
        return role

    def test_create_role_duplicate_name(self):
        role_data = deepcopy(self.BASE_JSON)
        role_data["name"] = "role1"
        with self.assertRaises(ValueError):
            _update_role_from_view(self.domain, role_data)

    def test_update_role(self):
        role_data = deepcopy(self.BASE_JSON)
        role_data["_id"] = self.role.get_id
        role_data["name"] = "role1"  # duplicate name during update is OK for now
        role_data["default_landing_page"] = None
        role_data["is_non_admin_editable"] = True
        role_data["permissions"] = get_default_available_permissions(
            edit_reports=True, view_report_list=["report1"]
        )
        updated_role = _update_role_from_view(self.domain, role_data)
        self.assertEqual(updated_role.name, "role1")
        self.assertIsNone(updated_role.default_landing_page)
        self.assertTrue(updated_role.is_non_admin_editable)
        self.assertEqual(updated_role.assignable_by, [])
        self.assertEqual(updated_role.permissions.to_json(), role_data['permissions'])

    def test_update_role_for_manage_domain_alerts(self):
        def patch_privilege_check(_domain, privilege_slug):
            if privilege_slug == privileges.CUSTOM_DOMAIN_ALERTS:
                return True
            return domain_has_privilege(_domain, privilege_slug)

        role_data = deepcopy(self.BASE_JSON)
        role_data['_id'] = self.role.get_id
        role_data['permissions']['manage_domain_alerts'] = True
        self.assertFalse(self.role.permissions.to_json()['manage_domain_alerts'])

        role_data['permissions']['manage_domain_alerts'] = True
        with patch('corehq.apps.users.views.domain_has_privilege', side_effect=patch_privilege_check):
            _update_role_from_view(self.domain, role_data)
        self.role.refresh_from_db()
        self.assertTrue(self.role.permissions.to_json()['manage_domain_alerts'])

    def test_landing_page_validation(self):
        role_data = deepcopy(self.BASE_JSON)
        role_data["default_landing_page"] = "bad value"
        with self.assertRaises(ValueError):
            _update_role_from_view(self.domain, role_data)


class TestDeleteRole(TestCase):
    domain = 'test-role-delete'

    def test_delete_role(self):
        role = UserRole.create(self.domain, 'test-role')
        _delete_user_role(self.domain, {"_id": role.get_id})
        self.assertFalse(UserRole.objects.filter(pk=role.id).exists())

    def test_delete_role_not_exist(self):
        with self.assertRaises(Http404):
            _delete_user_role(self.domain, {"_id": "mising"})

    def test_delete_role_with_users(self):
        self.user_count_mock.return_value = 1
        role = UserRole.create(self.domain, 'test-role')
        with self.assertRaisesRegex(InvalidRequestException, "It has one user"):
            _delete_user_role(self.domain, {"_id": role.get_id, 'name': role.name})

    def test_delete_commcare_user_default_role(self):
        role = UserRole.create(self.domain, 'test-role', is_commcare_user_default=True)
        with self.assertRaisesRegex(InvalidRequestException, "default role for Mobile Users"):
            _delete_user_role(self.domain, {"_id": role.get_id, 'name': role.name})

    def test_delete_role_wrong_domain(self):
        role = UserRole.create("other-domain", 'test-role')
        with self.assertRaises(Http404):
            _delete_user_role(self.domain, {"_id": role.get_id})

    def setUp(self):
        user_count_patcher = patch('corehq.apps.users.views.get_role_user_count', return_value=0)
        self.user_count_mock = user_count_patcher.start()
        self.addCleanup(user_count_patcher.stop)


class TestDeletePhoneNumberView(TestCase):
    domain = 'test'

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.webuser_username = f"webuser@{cls.domain}.commcarehq.org"
        cls.dummy_password = "******"
        cls.project = create_domain(cls.domain)
        cls.web_user = WebUser.create(cls.domain, cls.webuser_username, cls.dummy_password,
                                      None, None)
        cls.web_user.set_role(cls.domain, 'admin')
        cls.web_user.save()
        cls.commcare_user = CommCareUser.create(cls.domain, "test-user", cls.dummy_password, None, None)

    @classmethod
    def tearDownClass(cls):
        cls.commcare_user.delete(cls.domain, deleted_by=None)
        cls.web_user.delete(cls.domain, deleted_by=None)
        cls.project.delete()
        super().tearDownClass()

    def setUp(self):
        self.client.login(username=self.webuser_username, password=self.dummy_password)

    def test_no_phone_number(self):
        response = self.client.post(
            reverse('delete_phone_number', args=[self.domain, self.commcare_user.get_id]),
            {'phone_number': ''}
        )
        self.assertEqual(response.status_code, 404)

    def test_delete_phone_number(self):
        phone_number = '99999999'
        self.client.post(
            reverse('delete_phone_number', args=[self.domain, self.commcare_user.get_id]),
            {'phone_number': phone_number}
        )

        user_history_log = UserHistory.objects.get(user_id=self.commcare_user.get_id)
        self.assertIsNone(user_history_log.message)
        self.assertEqual(user_history_log.change_messages, UserChangeMessage.phone_numbers_removed([phone_number]))
        self.assertEqual(user_history_log.changed_by, self.web_user.get_id)
        self.assertEqual(user_history_log.by_domain, self.domain)
        self.assertEqual(user_history_log.for_domain, self.domain)
        self.assertEqual(user_history_log.changed_via, USER_CHANGE_VIA_WEB)


@es_test(requires=[user_adapter], setup_class=True)
@patch('corehq.apps.users.decorators.can_use_filtered_user_download', return_value=True)
class TestCountWebUsers(TestCase):

    view = 'count_web_users'

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        cls.domain = 'test'
        cls.domain_obj = create_domain(cls.domain)

        location_type = LocationType(domain=cls.domain, name='phony')
        location_type.save()

        cls.some_location = make_loc('1', 'some_location', type=location_type, domain=cls.domain_obj.name)

        cls.admin_user = WebUser.create(
            cls.domain_obj.name,
            'edith1@wharton.com',
            'badpassword',
            None,
            None,
            email='edith@wharton.com',
            first_name='Edith',
            last_name='Wharton',
            is_admin=True,
        )

        cls.admin_user_with_location = WebUser.create(
            cls.domain_obj.name,
            'edith2@wharton.com',
            'badpassword',
            None,
            None,
            email='edith2@wharton.com',
            first_name='Edith',
            last_name='Wharton',
            is_admin=True,
        )
        cls.admin_user_with_location.set_location(cls.domain, cls.some_location)

        populate_user_index([
            cls.admin_user_with_location,
            cls.admin_user,
        ])

    @classmethod
    def tearDownClass(cls):
        delete_all_locations()

        cls.admin_user_with_location.delete(cls.domain_obj.name, deleted_by=None)
        cls.admin_user.delete(cls.domain_obj.name, deleted_by=None)

        cls.domain_obj.delete()
        super().tearDownClass()

    def test_admin_user_sees_all_web_users(self, _):
        self.client.login(
            username=self.admin_user.username,
            password='badpassword',
        )
        result = self.client.get(reverse(self.view, kwargs={'domain': self.domain}))
        self.assertEqual(json.loads(result.content)['user_count'], 2)

    def test_admin_location_user_sees_all_web_users(self, _):
        self.client.login(
            username=self.admin_user_with_location.username,
            password='badpassword',
        )
        result = self.client.get(reverse(self.view, kwargs={'domain': self.domain}))
        self.assertEqual(json.loads(result.content)['user_count'], 2)


class BulkUserUploadAPITest(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.domain_name = 'bulk-user-upload-domain'
        cls.domain = create_domain(cls.domain_name)
        cls.domain.save()
        cls.addClassCleanup(cls.domain.delete)
        cls.user = WebUser.create(cls.domain_name, 'test@test.com', 'password', created_by=None, created_via=None)
        cls.addClassCleanup(cls.user.delete, cls.domain_name, deleted_by=None)
        cls.api_key = HQApiKey.objects.create(user=cls.user.get_django_user(), domain=cls.domain_name)

    def setUp(self):
        self.client = Client()
        self.url = reverse('bulk_user_upload_api', args=[self.domain_name])

    @staticmethod
    def _create_valid_workbook():
        workbook = Workbook()
        users_sheet = workbook.create_sheet(title='users')
        users_sheet.append(['username', 'email', 'password'])
        users_sheet.append(['test_user', 'test@example.com', 'password'])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)
        file.name = 'users.xlsx'

        return file

    def _make_post_request(self, file):
        return self.client.post(
            self.url,
            {'bulk_upload_file': file},
            HTTP_AUTHORIZATION=f'ApiKey {self.user.username}:{self.api_key.plaintext_key}',
            format='multipart'
        )

    def test_success(self):
        file = self._create_valid_workbook()

        with patch('corehq.apps.users.views.mobile.users.BaseUploadUser.upload_users'):
            response = self._make_post_request(file)
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json(), {'success': True})

    def test_api_no_authentication(self):
        response = self.client.post(self.url, {'bulk_upload_file': 'mock_file'})
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.content.decode(), 'Authorization Required')

    def test_api_invalid_authentication(self):
        response = self.client.post(
            self.url,
            {'bulk_upload_file': 'mock_file'},
            HTTP_AUTHORIZATION=f'ApiKey {self.user.username}:invalid_key'
        )
        self.assertEqual(response.status_code, 401)

    def test_no_file_uploaded(self):
        response = self.client.post(
            self.url,
            HTTP_AUTHORIZATION=f'ApiKey {self.user.username}:{self.api_key.plaintext_key}',
            format='multipart'
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json(), {'success': False, 'message': 'no file uploaded'})

    @patch('corehq.apps.users.views.mobile.users.get_workbook')
    def test_invalid_file_format(self, mock_get_workbook):
        mock_get_workbook.side_effect = WorkbookJSONError('Invalid file format')
        file = BytesIO(b'invalid file content')
        file.name = 'invalid_file.txt'
        response = self._make_post_request(file)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json(), {'success': False, 'message': 'Invalid file format'})

    def test_invalid_workbook_headers(self):
        workbook = Workbook()
        users_sheet = workbook.create_sheet(title='users')
        users_sheet.append(['invalid_header', 'email', 'password'])
        users_sheet.append(['test_user', 'test@example.com', 'password'])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)
        file.name = 'users.xlsx'

        response = self._make_post_request(file)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json(), {
            'message': 'The following are required column headers: username.\n'
                       'The following are illegal column headers: invalid_header.',
            'success': False
        })

    @flag_enabled('TABLEAU_USER_SYNCING')
    def test_tableau_role_and_groups_headers(self):
        workbook = Workbook()
        users_sheet = workbook.create_sheet(title='users')
        users_sheet.append(['username', 'email', 'password', 'tableau_role', 'tableau_groups'])
        users_sheet.append(['test_user', 'test@example.com', 'password', 'fakerole', 'fakegroup'])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)
        file.name = 'users.xlsx'

        # Test user with permission to edit Tableau Configs
        self.user.is_superuser = False
        role_with_upload_and_edit_tableau_permission = UserRole.create(
            self.domain, 'edit-tableau', permissions=HqPermissions(edit_web_users=True,
                                                                   edit_user_tableau_config=True)
        )
        self.user.set_role(self.domain_name,
                        role_with_upload_and_edit_tableau_permission.get_qualified_id())
        self.user.save()

        with patch('corehq.apps.users.views.mobile.users.BaseUploadUser.upload_users'):
            response = self._make_post_request(file)
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json(), {'success': True})

        # Test user without permission to edit Tableau Configs
        role_with_upload_permission = UserRole.create(
            self.domain, 'edit-web-users', permissions=HqPermissions(edit_web_users=True)
        )
        self.user.set_role(self.domain_name, role_with_upload_permission.get_qualified_id())
        self.user.save()

        file.seek(0)
        response = self._make_post_request(file)
        self.assertEqual(response.status_code, 400)

        expected_pattern = re.compile(
            r"Only users with 'Manage Tableau Configuration' edit permission in domains "
            r"where Tableau User Syncing is enabled can upload files with 'Tableau Role' "
            r"and/or 'Tableau Groups' fields\.\nThe following are illegal column headers: "
            r"(?:tableau_groups, tableau_role|tableau_role, tableau_groups)\.",
        )
        self.assertRegex(response.json()['message'], expected_pattern)

    @patch('corehq.apps.users.views.mobile.users.BaseUploadUser.upload_users')
    def test_user_upload_error(self, mock_upload_users):
        mock_upload_users.side_effect = UserUploadError('User upload error')
        file = self._create_valid_workbook()

        response = self._make_post_request(file)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json(), {'success': False, 'message': 'User upload error'})

    @patch('corehq.apps.users.views.mobile.users.notify_exception')
    @patch('corehq.apps.users.views.mobile.users.BaseUploadUser.upload_users')
    def test_exception(self, mock_upload_users, mock_notify_exception):
        mock_upload_users.side_effect = Exception('Unexpected error')
        file = self._create_valid_workbook()

        response = self._make_post_request(file)

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.json(), {'success': False, 'message': 'Unexpected error'})
        mock_notify_exception.assert_called_once_with(None, message='Unexpected error')

    def test_cant_upload_multiple_files(self):
        file1 = self._create_valid_workbook()
        file2 = self._create_valid_workbook()

        response = self.client.post(
            self.url,
            {'bulk_upload_file': file1, 'another_file': file2},
            HTTP_AUTHORIZATION=f'ApiKey {self.user.username}:{self.api_key.plaintext_key}',
            format='multipart'
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json(), {
            'success': False,
            'message': 'only one file can be uploaded at a time'
        })


class BaseUploadUserTest(TestCase):

    mock_couch_user = WebUser(
        username="testuser",
        _id="user123",
        domain="test-domain",
    )

    def setUp(self):
        self.domain = 'test-domain'
        self.factory = RequestFactory()
        self.view = BaseUploadUser()
        self.view.request = self.factory.get('/')
        self.view.args = []
        self.view.kwargs = {'domain': self.domain}
        self.view._domain = self.domain
        self.view.is_web_upload = True

    @patch('corehq.apps.users.views.reverse')
    @patch('corehq.apps.users.views.BaseUploadUser.upload_users')
    @patch('corehq.apps.users.views.BaseUploadUser.process_workbook')
    @patch('corehq.apps.users.views.get_workbook')
    def test_post_success(self, mock_get_workbook, mock_process_workbook, mock_upload_users, mock_reverse):
        mock_get_workbook.return_value = Mock()
        mock_process_workbook.return_value = (Mock(), Mock())
        mock_task_ref = Mock()
        mock_upload_users.return_value = mock_task_ref
        mock_reverse.return_value = '/success/'

        request = self.factory.post('/', {'bulk_upload_file': Mock()})
        request.couch_user = self.mock_couch_user
        response = self.view.post(request)

        mock_reverse.assert_called_once_with(
            'web_user_upload_status',
            args=[self.domain, mock_task_ref.download_id]
        )
        self.assertIsInstance(response, HttpResponseRedirect)
        self.assertEqual(response.url, '/success/')
