from collections import defaultdict
import openpyxl

from django.urls import reverse
from django.utils.translation import ugettext as _

from lxml import etree

from corehq.apps.app_manager.views.media_utils import interpolate_media_path
from corehq.apps.translations.app_translations.download import get_bulk_app_single_sheet_by_name
from corehq.apps.translations.app_translations.utils import get_bulk_app_sheet_headers
from corehq.apps.translations.const import SINGLE_SHEET_NAME


def download_multimedia_paths_rows(app, only_missing=False):
    paths = defaultdict(list)
    for ref in app.all_media():
        paths[ref.path].append(ref)

    module_index_by_unique_id = {m.unique_id: m.id for m in app.get_modules()}

    def _readable_ref(ref):
        module_index = module_index_by_unique_id[ref.module_unique_id]
        readable = _("Menu {index}: {name}").format(index=module_index, name=ref.get_module_name())
        if ref.form_unique_id is not None:
            readable += _(" > Form {index}: {name}").format(index=ref.form_order, name=ref.get_form_name())
        return readable

    rows = []
    for path, refs in paths.items():
        if not only_missing or path not in app.multimedia_map:
            rows.append((_("Paths"), [path, ''] + [_readable_ref(r) for r in refs]))

    return rows


def validate_multimedia_paths_rows(app, rows):
    old_paths_last_seen = {i.path: None for i in app.all_media()}
    new_paths_last_seen = defaultdict(lambda: None)

    errors = []
    warnings = []
    for i, row in enumerate(rows):
        (old_path, new_path) = row

        if old_path not in old_paths_last_seen:
            errors.append(_("Path in row {} could not be found in application: "
                            "<code>{}</code>").format(i, old_path))
        elif old_path == new_path:
            errors.append(_("In row {}, old and new paths are both <code>{}</code>. Please provide "
                            "an updated path or remove this row").format(i, old_path))
        elif old_paths_last_seen[old_path] is not None:
            # Duplicate old paths is an error: can't rename to two different new values
            errors.append(_("Path in row {} was already renamed in row {}: "
                            "<code>{}</code>").format(i, old_paths_last_seen[old_path], old_path))
        old_paths_last_seen[old_path] = i

        interpolated_new_path = interpolate_media_path(new_path)    # checks for jr://
        if interpolated_new_path != new_path:
            warnings.append(_("Path <code>{}</code> in row {} was replaced with "
                              "<code>{}</code>").format(new_path, i, interpolated_new_path))
        else:
            # It's usually a bad idea to change file extensions, since the file itself isn't changing
            old_extension = old_path.split(".")[-1].lower()
            new_extension = new_path.split(".")[-1].lower()
            if old_extension != new_extension:
                warnings.append(_("File extension in row {} changed "
                                  "from {} to {}".format(i, old_extension, new_extension)))

            # Duplicate new paths is a warning: will combine what were previously different items
            if new_path in new_paths_last_seen:
                warnings.append(_("New path in row {} was already used to rename row {}: "
                                  "<code>{}</code>").format(i, new_paths_last_seen[new_path], new_path))
        new_paths_last_seen[new_path] = i

    return errors, warnings


def update_multimedia_paths(app, paths):
    # Update module and form references
    success_counts = defaultdict(lambda: 0)
    for old_path, new_path in paths.items():
        for module in app.modules:
            success_counts[module.unique_id] += module.rename_media(old_path, new_path)
            for form in module.get_forms():
                update_count = form.rename_media(old_path, new_path)
                if update_count:
                    success_counts[form.unique_id] += update_count

    # Update app's upstream map of multimedia
    for old_path, new_path in paths.items():
        if old_path in app.multimedia_map:  # path will not be present if file is missing from app
            app.multimedia_map.update({
                new_path: app.multimedia_map[old_path],
            })

    # Put together success messages
    successes = []
    for module in app.modules:
        if success_counts[module.unique_id]:
            successes.append(_("{} item(s) updated in <a href='{}' target='_blank'>{}</a>").format(
                             success_counts[module.unique_id],
                             reverse("view_module", args=[app.domain, app.id, module.unique_id]),
                             module.default_name()))
        for form in module.forms:
            if success_counts[form.unique_id]:
                successes.append(_("{} item(s) updated in <a href='{}' target='_blank'>{}</a>").format(
                                 success_counts[form.unique_id],
                                 reverse("view_form", args=[app.domain, app.id, form.unique_id]),
                                 "{} > {}".format(module.default_name(), form.default_name())))

    return successes


def download_audio_translator_files(domain, app, lang, eligible_for_transifex_only=True):
    # Get bulk app translation single sheet data
    headers = get_bulk_app_sheet_headers(app, single_sheet=True, lang=lang,
                                         eligible_for_transifex_only=eligible_for_transifex_only)
    headers = headers[0]    # There's only one row since these are the headers for the single-sheet format
    headers = headers[1]    # Drop the first element (sheet name), leaving the second (list of header names)
    audio_text_index = headers.index('default_' + lang)
    audio_path_index = headers.index('audio_' + lang)
    sheets = get_bulk_app_single_sheet_by_name(app, lang, eligible_for_transifex_only=True)
    audio_rows = [row for row in sheets[SINGLE_SHEET_NAME] if row[audio_path_index]]

    # Create file for re-upload to HQ's bulk app translations
    upload_workbook = openpyxl.Workbook()
    upload_sheet = upload_workbook.worksheets[0]
    upload_sheet.title = SINGLE_SHEET_NAME
    upload_sheet.append(headers)

    # Create dict of audio path to text, and disambiguate any missing path that points to multiple texts
    rows_by_audio = {}
    for row in audio_rows:
        audio_path = row[audio_path_index]
        text = row[audio_text_index]
        if audio_path in rows_by_audio and audio_path not in app.multimedia_map:
            if rows_by_audio[audio_path] != text:
                extension = "." + audio_path.split(".")[-1]
                not_extension = audio_path[:-len(extension)]
                suffix = 1
                while audio_path in rows_by_audio and rows_by_audio[audio_path] != text:
                    suffix += 1
                    audio_path = "{}_{}{}".format(not_extension, suffix, extension)
                row[audio_path_index] = audio_path
                upload_sheet.append(row)    # add new path to sheet for re-upload to HQ
        rows_by_audio[audio_path] = text

    # Create dict of rows, keyed by label text to de-duplicate paths
    rows_by_text = defaultdict(list)
    for row in audio_rows:
        rows_by_text[row[audio_text_index]].append(row)

    def _get_filename_from_duplicate_rows(rows):
        return rows[0][audio_path_index]

    # Add a row to upload sheet for each filename being eliminated because the text was duplicated
    for text, rows in rows_by_text.items():
        filename = _get_filename_from_duplicate_rows(rows)
        for row in rows:
            if row[audio_path_index] != filename:
                row[audio_path_index] = filename
                upload_sheet.append(row)

    # Create file for translato, with a row for each unique text label
    translator_workbook = openpyxl.Workbook()
    sheet0 = translator_workbook.worksheets[0]
    sheet0.title = "filepaths"
    sheet0.append([lang, "audio"])
    sheet1 = translator_workbook.create_sheet("verification")
    sheet1.append(headers)
    for text, rows in rows_by_text.items():
        if not any([row[audio_path_index] in app.multimedia_map for row in rows]):
            filename = _get_filename_from_duplicate_rows(rows)
            sheet0.append([text, filename])
            sheet1.append(rows[0])

    return {
        "bulk_upload.xlsx": upload_workbook,
        "excel_for_translator.xlsx": translator_workbook,
    }
