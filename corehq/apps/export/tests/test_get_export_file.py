import json
import re
from unittest.mock import patch

from django.core.cache import cache
from django.test import SimpleTestCase

from openpyxl import load_workbook

from couchexport.export import get_writer
from couchexport.models import Format
from couchexport.transforms import couch_to_excel_datetime

from corehq.apps.es.cases import case_adapter
from corehq.apps.es.tests.utils import es_test
from corehq.apps.export.const import (
    CASE_NAME_TRANSFORM,
    DEID_DATE_TRANSFORM,
    EMPTY_VALUE,
    MISSING_VALUE,
)
from corehq.apps.export.export import (
    ExportFile,
    _ExportWriter,
    get_export_file,
    get_export_writer,
    write_export_instance,
)
from corehq.apps.export.models import (
    MAIN_TABLE,
    CaseExportInstance,
    ExportColumn,
    ExportItem,
    FormExportInstance,
    MultipleChoiceItem,
    Option,
    PathNode,
    ScalarItem,
    SplitExportColumn,
    StockFormExportColumn,
    StockItem,
    TableConfiguration,
)
from corehq.apps.export.tests.util import (
    DEFAULT_CASE_TYPE,
    DOMAIN,
    get_export_json,
    new_case,
)
from corehq.util.files import TransientTempfile
from corehq.util.test_utils import flag_enabled


def assert_instance_gives_results(docs, export_instance, expected_result):
    with TransientTempfile() as temp_path:
        writer = get_export_writer([export_instance], temp_path)
        with writer.open([export_instance]):
            write_export_instance(writer, export_instance, docs)

        with ExportFile(writer.path, writer.format) as export:
            assert json.loads(export.read()) == expected_result


class WriterTest(SimpleTestCase):

    docs = [
        {
            'domain': 'my-domain',
            '_id': '1234',
            "form": {
                "q1": "foo",
                "q2": {
                    "q4": "bar",
                },
                "q3": "baz",
                "mc": "two extra"
            }
        },
        {
            'domain': 'my-domain',
            '_id': '12345',
            "form": {
                "q1": "bip",
                "q2": {
                    "q4": "boop",
                },
                "q3": "bop",
                "mc": "one two",
                "date": "2015-07-22T14:16:49.584880Z",
            }
        },
    ]

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_simple_table(self, export_save):
        """
        Confirm that some simple documents and a simple FormExportInstance
        are writtern with _write_export_file() correctly
        """

        export_instance = FormExportInstance(
            export_format=Format.JSON,
            tables=[
                TableConfiguration(
                    label="My table",
                    selected=True,
                    columns=[
                        ExportColumn(
                            label="Q3",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='q3')],
                            ),
                            selected=True
                        ),
                        ExportColumn(
                            label="Q1",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='q1')],
                            ),
                            selected=True
                        ),
                    ]
                )
            ]
        )

        assert_instance_gives_results(self.docs, export_instance, {
            'My table': {
                'headers': ['Q3', 'Q1'],
                'rows': [['baz', 'foo'], ['bop', 'bip']],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    @patch('corehq.apps.export.export.MAX_NORMAL_EXPORT_SIZE', 2)
    @flag_enabled('PAGINATED_EXPORTS')
    def test_paginated_table(self, export_save):
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            tables=[
                TableConfiguration(
                    label="My table",
                    selected=True,
                    columns=[
                        ExportColumn(
                            label="Q3",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='q3')],
                            ),
                            selected=True
                        ),
                        ExportColumn(
                            label="Q1",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='q1')],
                            ),
                            selected=True
                        ),
                    ]
                )
            ]
        )

        assert_instance_gives_results(self.docs + self.docs, export_instance, {
            'My table_000': {
                'headers': ['Q3', 'Q1'],
                'rows': [['baz', 'foo'], ['bop', 'bip']],
            },
            'My table_001': {
                'headers': ['Q3', 'Q1'],
                'rows': [['baz', 'foo'], ['bop', 'bip']],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_split_questions(self, export_save):
        """Ensure columns are split when `split_multiselects` is set to True"""
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            domain=DOMAIN,
            case_type=DEFAULT_CASE_TYPE,
            split_multiselects=True,
            tables=[TableConfiguration(
                label="My table",
                selected=True,
                path=[],
                columns=[
                    SplitExportColumn(
                        label="MC",
                        item=MultipleChoiceItem(
                            path=[PathNode(name='form'), PathNode(name='mc')],
                            options=[
                                Option(value='one'),
                                Option(value='two'),
                            ]
                        ),
                        selected=True,
                    )
                ]
            )]
        )

        assert_instance_gives_results(self.docs, export_instance, {
            'My table': {
                'headers': ['MC | one', 'MC | two', 'MC | extra'],
                'rows': [[EMPTY_VALUE, 1, 'extra'], [1, 1, '']],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_array_data_in_scalar_question(self, export_save):
        '''
        This test ensures that when a question id has array data
        that we return still return a string for scalar data.
        This happens rarely
        '''
        doc = {
            'domain': 'my-domain',
            '_id': '12345',
            "form": {
                "array": ["one", "two"],
            }
        }

        export_instance = FormExportInstance(
            export_format=Format.JSON,
            domain=DOMAIN,
            xmlns='xmlns',
            tables=[TableConfiguration(
                label="My table",
                selected=True,
                path=[],
                columns=[
                    ExportColumn(
                        label="Scalar Array",
                        item=ScalarItem(path=[PathNode(name='form'), PathNode(name='array')]),
                        selected=True,
                    )
                ]
            )]
        )

        assert_instance_gives_results([doc], export_instance, {
            'My table': {
                'headers': ['Scalar Array'],
                'rows': [['one two']],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_form_stock_columns(self, export_save):
        """Ensure that we can export stock properties in a form export"""
        docs = [{
            '_id': 'simone-biles',
            'domain': DOMAIN,
            'form': {
                'balance': [
                    {
                        '@type': 'question-id',
                        'entry': {
                            '@quantity': '2',
                        }
                    }, {
                        '@type': 'other-question-id',
                        'entry': {
                            '@quantity': '3',
                        }
                    }]
            },
        }, {
            '_id': 'sam-mikulak',
            'domain': DOMAIN,
            'form': {
                'balance': {
                    '@type': 'question-id',
                    'entry': {
                        '@quantity': '2',
                    }
                },
            },
        }, {
            '_id': 'kerri-walsh',
            'domain': DOMAIN,
            'form': {
                'balance': {
                    '@type': 'other-question-id',
                    'entry': {
                        '@quantity': '2',
                    }
                },
            },
        }, {
            '_id': 'april-ross',
            'domain': DOMAIN,
            'form': {},
        }]
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            domain=DOMAIN,
            tables=[TableConfiguration(
                label="My table",
                selected=True,
                path=[],
                columns=[
                    StockFormExportColumn(
                        label="StockItem @type",
                        item=StockItem(
                            path=[
                                PathNode(name='form'),
                                PathNode(name='balance:question-id'),
                                PathNode(name='@type'),
                            ],
                        ),
                        selected=True,
                    ),
                    StockFormExportColumn(
                        label="StockItem @quantity",
                        item=StockItem(
                            path=[
                                PathNode(name='form'),
                                PathNode(name='balance:question-id'),
                                PathNode(name='entry'),
                                PathNode(name='@quantity'),
                            ],
                        ),
                        selected=True,
                    ),
                ]
            )]
        )

        assert_instance_gives_results(docs, export_instance, {
            'My table': {
                'headers': ['StockItem @type', 'StockItem @quantity'],
                'rows': [
                    ['question-id', '2'],
                    ['question-id', '2'],
                    [MISSING_VALUE, MISSING_VALUE],
                    [MISSING_VALUE, MISSING_VALUE],
                ],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_transform_dates(self, export_save):
        """Ensure dates are transformed for excel when `transform_dates` is set to True"""
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            domain=DOMAIN,
            case_type=DEFAULT_CASE_TYPE,
            transform_dates=True,
            tables=[TableConfiguration(
                label="My table",
                selected=True,
                path=[],
                columns=[
                    ExportColumn(
                        label="Date",
                        item=MultipleChoiceItem(
                            path=[PathNode(name='form'), PathNode(name='date')],
                        ),
                        selected=True,
                    )
                ]
            )]
        )

        assert_instance_gives_results(self.docs, export_instance, {
            'My table': {
                'headers': ['Date'],
                'rows': [[MISSING_VALUE], [couch_to_excel_datetime('2015-07-22T14:16:49.584880Z', None)]],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_split_questions_false(self, export_save):
        """Ensure multiselects are not split when `split_multiselects` is set to False"""
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            domain=DOMAIN,
            case_type=DEFAULT_CASE_TYPE,
            split_multiselects=False,
            tables=[TableConfiguration(
                label="My table",
                selected=True,
                path=[],
                columns=[
                    SplitExportColumn(
                        label="MC",
                        item=MultipleChoiceItem(
                            path=[PathNode(name='form'), PathNode(name='mc')],
                            options=[
                                Option(value='one'),
                                Option(value='two'),
                            ]
                        ),
                        selected=True,
                    )
                ]
            )]
        )

        assert_instance_gives_results(self.docs, export_instance, {
            'My table': {
                'headers': ['MC'],
                'rows': [['two extra'], ['one two']],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_multi_table(self, export_save):
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            tables=[
                TableConfiguration(
                    label="My table",
                    selected=True,
                    path=[],
                    columns=[
                        ExportColumn(
                            label="Q3",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='q3')],
                            ),
                            selected=True,
                        ),
                    ]
                ),
                TableConfiguration(
                    label="My other table",
                    selected=True,
                    path=[PathNode(name='form', is_repeat=False), PathNode(name="q2", is_repeat=False)],
                    columns=[
                        ExportColumn(
                            label="Q4",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='q2'), PathNode(name='q4')],
                            ),
                            selected=True,
                        ),
                    ]
                )
            ]
        )

        assert_instance_gives_results(self.docs, export_instance, {
            'My table': {
                'headers': ['Q3'],
                'rows': [['baz'], ['bop']],
            },
            'My other table': {
                'headers': ['Q4'],
                'rows': [['bar'], ['boop']],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_multi_table_order(self, export_save):
        tables = [
            TableConfiguration(
                label="My table {}".format(i),
                selected=True,
                path=[],
                columns=[
                    ExportColumn(
                        label="Q{}".format(i),
                        item=ScalarItem(
                            path=[PathNode(name='form'), PathNode(name='q{}'.format(i))],
                        ),
                        selected=True,
                    ),
                ]
            )
            for i in range(10)
        ]
        export_instance = FormExportInstance(
            export_format=Format.HTML,
            tables=tables
        )

        docs = [
            {
                'domain': 'my-domain',
                '_id': '1234',
                "form": {'q{}'.format(i): 'value {}'.format(i) for i in range(10)}
            }
        ]

        with TransientTempfile() as temp_path:
            writer = get_export_writer([export_instance], temp_path)
            with writer.open([export_instance]):
                write_export_instance(writer, export_instance, docs)
            with ExportFile(writer.path, writer.format) as export:
                exported_tables = [table for table in re.findall(b'<table>', export.read())]

        expected_tables = [t.label for t in tables]
        self.assertEqual(len(expected_tables), len(exported_tables))
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_multiple_write_export_instance_calls(self, export_save):
        """
        Confirm that calling _write_export_instance() multiple times
        (as part of a bulk export) works as expected.
        """
        export_instances = [
            FormExportInstance(
                tables=[
                    TableConfiguration(
                        label="My table",
                        selected=True,
                        path=[],
                        columns=[
                            ExportColumn(
                                label="Q3",
                                item=ScalarItem(
                                    path=[PathNode(name='form'), PathNode(name='q3')],
                                ),
                                selected=True,
                            ),
                        ]
                    ),
                ]
            ),
            FormExportInstance(
                tables=[
                    TableConfiguration(
                        label="My other table",
                        selected=True,
                        path=[PathNode(name="form", is_repeat=False), PathNode(name="q2", is_repeat=False)],
                        columns=[
                            ExportColumn(
                                label="Q4",
                                item=ScalarItem(
                                    path=[PathNode(name='form'), PathNode(name='q2'), PathNode(name='q4')],
                                ),
                                selected=True,
                            ),
                        ]
                    )
                ]
            ),
            FormExportInstance(
                tables=[
                    TableConfiguration(
                        label="My other table",
                        selected=True,
                        path=[PathNode(name="form", is_repeat=False), PathNode(name="q2", is_repeat=False)],
                        columns=[
                            ExportColumn(
                                label="Q4",
                                item=ScalarItem(
                                    path=[PathNode(name='form'), PathNode(name='q2'), PathNode(name='q4')],
                                ),
                                selected=True,
                            ),
                        ]
                    )
                ]
            )

        ]

        with TransientTempfile() as temp_path:
            writer = _ExportWriter(get_writer(Format.JSON), temp_path)
            with writer.open(export_instances):
                write_export_instance(writer, export_instances[0], self.docs)
                write_export_instance(writer, export_instances[1], self.docs)
                write_export_instance(writer, export_instances[2], self.docs)

            with ExportFile(writer.path, writer.format) as export:
                self.assertEqual(
                    json.loads(export.read()),
                    {
                        'My table': {
                            'headers': ['Q3'],
                            'rows': [['baz'], ['bop']],

                        },
                        'Export2-My other table': {
                            'headers': ['Q4'],
                            'rows': [['bar'], ['boop']],
                        },
                        'Export3-My other table': {
                            'headers': ['Q4'],
                            'rows': [['bar'], ['boop']],
                        },
                    }
                )
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_empty_location(self, export_save):
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            tables=[
                TableConfiguration(
                    label="My table",
                    selected=True,
                    columns=[
                        ExportColumn(
                            label="location",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='meta'), PathNode(name='location')],
                            ),
                            selected=True
                        ),
                    ]
                )
            ]
        )

        docs = [
            {
                'domain': 'my-domain',
                '_id': '1234',
                'form': {
                    'meta': {
                        'location': {'xmlns': 'abc'},
                    }
                }
            }
        ]

        assert_instance_gives_results(docs, export_instance, {
            'My table': {
                'headers': ['location'],
                'rows': [[EMPTY_VALUE]],
            }
        })
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_empty_table_label(self, export_save):
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            domain=DOMAIN,
            case_type=DEFAULT_CASE_TYPE,
            split_multiselects=True,
            tables=[TableConfiguration(
                label="",
                selected=True,
                path=[],
                columns=[
                    ExportColumn(
                        label="Q1",
                        item=ScalarItem(
                            path=[PathNode(name='form'), PathNode(name='q1')],
                        ),
                        selected=True
                    ),
                ]
            )]
        )

        assert_instance_gives_results(self.docs, export_instance, {
            'Sheet1': {
                'headers': ['Q1'],
                'rows': [['foo'], ['bip']],
            }
        })
        self.assertTrue(export_save.called)


@es_test(requires=[case_adapter], setup_class=True)
class ExportTest(SimpleTestCase):

    @classmethod
    def setUpClass(cls):
        super(ExportTest, cls).setUpClass()
        with patch('corehq.pillows.utils.get_user_type', return_value='CommCareUser'):
            cases = [
                new_case(
                    case_id='robin',
                    name='batman',
                    case_json={"foo": "apple", "bar": "banana", "date": '2016-4-24'},
                ),
                new_case(
                    owner_id="some_other_owner",
                    case_json={"foo": "apple", "bar": "banana", "date": '2016-4-04'},
                ),
                new_case(type="some_other_type", case_json={"foo": "apple", "bar": "banana"}),
                new_case(closed=True, case_json={"foo": "apple", "bar": "banana"})
            ]
            case_adapter.bulk_index(cases, refresh=True)

            cache.clear()

    @classmethod
    def tearDownClass(cls):
        cache.clear()
        super(ExportTest, cls).tearDownClass()

    @patch('corehq.apps.export.models.CaseExportInstance.save')
    def test_get_export_file(self, export_save):
        export_json = get_export_json(
            CaseExportInstance(
                export_format=Format.JSON,
                domain=DOMAIN,
                case_type=DEFAULT_CASE_TYPE,
                tables=[TableConfiguration(
                    label="My table",
                    selected=True,
                    path=[],
                    columns=[
                        ExportColumn(
                            label="Foo column",
                            item=ExportItem(
                                path=[PathNode(name="foo")]
                            ),
                            selected=True,
                        ),
                        ExportColumn(
                            label="Bar column",
                            item=ExportItem(
                                path=[PathNode(name="bar")]
                            ),
                            selected=True,
                        )
                    ]
                )]
            ),
        )

        self.assertEqual(
            export_json,
            {
                'My table': {
                    'headers': [
                        'Foo column',
                        'Bar column'],
                    'rows': [
                        ['apple', 'banana'],
                        ['apple', 'banana'],
                        ['apple', 'banana'],
                    ],
                }
            }
        )
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.FormExportInstance.save')
    def test_case_name_transform(self, export_save):
        docs = [
            {
                'domain': 'my-domain',
                '_id': '1234',
                "form": {
                    "caseid": "robin",
                },
            },
            {
                'domain': 'my-domain',
                '_id': '1234',
                "form": {
                    "caseid": "i-do-not-exist",
                },
            }
        ]
        export_instance = FormExportInstance(
            export_format=Format.JSON,
            tables=[
                TableConfiguration(
                    label="My table",
                    selected=True,
                    columns=[
                        ExportColumn(
                            label="case_name",
                            item=ScalarItem(
                                path=[PathNode(name='form'), PathNode(name='caseid')],
                                transform=CASE_NAME_TRANSFORM,
                            ),
                            selected=True
                        ),
                    ]
                )
            ]
        )

        assert_instance_gives_results(docs, export_instance, {
            'My table': {
                'headers': ['case_name'],
                'rows': [['batman'], [MISSING_VALUE]],
            }
        })
        self.assertTrue(export_save.called)

    @patch('couchexport.deid.DeidGenerator.random_number', return_value=3)
    @patch('corehq.apps.export.models.CaseExportInstance.save')
    def test_export_transforms(self, export_save, _):
        export_json = get_export_json(
            CaseExportInstance(
                export_format=Format.JSON,
                domain=DOMAIN,
                case_type=DEFAULT_CASE_TYPE,
                tables=[TableConfiguration(
                    label="My table",
                    selected=True,
                    path=[],
                    columns=[
                        ExportColumn(
                            label="DEID Date Transform column",
                            item=ExportItem(
                                path=[PathNode(name="date")]
                            ),
                            selected=True,
                            deid_transform=DEID_DATE_TRANSFORM,
                        )
                    ]
                )]
            ),
        )

        export_json['My table']['rows'].sort()
        self.assertEqual(
            export_json,
            {
                'My table': {
                    'headers': [
                        'DEID Date Transform column *sensitive*',
                    ],
                    'rows': [
                        [MISSING_VALUE],
                        ['2016-04-07'],
                        ['2016-04-27'],  # offset by 3 since that's the mocked random offset
                    ],
                }
            }
        )
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.CaseExportInstance.save')
    def test_selected_false(self, export_save):
        export_json = get_export_json(
            CaseExportInstance(
                export_format=Format.JSON,
                domain=DOMAIN,
                case_type=DEFAULT_CASE_TYPE,
                tables=[TableConfiguration(
                    label="My table",
                    selected=False,
                    path=[],
                    columns=[]
                )]
            )
        )
        self.assertEqual(export_json, {})
        self.assertTrue(export_save.called)

    @patch('corehq.apps.export.models.CaseExportInstance.save')
    def test_simple_bulk_export(self, export_save):

        with TransientTempfile() as temp_path:
            export_file = get_export_file(
                [
                    CaseExportInstance(
                        export_format=Format.JSON,
                        domain=DOMAIN,
                        case_type=DEFAULT_CASE_TYPE,
                        tables=[TableConfiguration(
                            selected=True,
                            label="My table",
                            path=MAIN_TABLE,
                            columns=[
                                ExportColumn(
                                    label="Foo column",
                                    item=ExportItem(
                                        path=[PathNode(name="foo")]
                                    ),
                                    selected=True,
                                ),
                            ]
                        )]
                    ),
                    CaseExportInstance(
                        export_format=Format.JSON,
                        domain=DOMAIN,
                        case_type=DEFAULT_CASE_TYPE,
                        tables=[TableConfiguration(
                            label="My table",
                            selected=True,
                            path=MAIN_TABLE,
                            columns=[
                                ExportColumn(
                                    label="Bar column",
                                    item=ExportItem(
                                        path=[PathNode(name="bar")]
                                    ),
                                    selected=True,
                                )
                            ]
                        )]
                    ),
                ],
                [],  # No filters
                temp_path,
            )

            expected = {
                'Export1-My table': {
                    "A1": "Foo column",
                    "A2": "apple",
                    "A3": "apple",
                    "A4": "apple",
                },
                "Export2-My table": {
                    "A1": "Bar column",
                    "A2": "banana",
                    "A3": "banana",
                    "A4": "banana",
                },
            }

            with export_file as export:
                wb = load_workbook(export)
                self.assertEqual(wb.sheetnames, ["Export1-My table", "Export2-My table"])

                for sheet in expected.keys():
                    for cell in expected[sheet].keys():
                        self.assertEqual(
                            wb[sheet][cell].value,
                            expected[sheet][cell],
                            'AssertionError: Sheet "{}", cell "{}" expected: "{}", got "{}"'.format(
                                sheet, cell, expected[sheet][cell], wb[sheet][cell].value
                            )
                        )
        self.assertTrue(export_save.called)


class TableHeaderTest(SimpleTestCase):

    def test_deid_column_headers(self):
        col = ExportColumn(
            label="my column",
            deid_transform="deid_id",
        )
        self.assertEqual(col.get_headers(), ["my column *sensitive*"])
