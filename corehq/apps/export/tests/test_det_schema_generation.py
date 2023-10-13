import os
import tempfile
from io import BytesIO
from unittest.mock import patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from corehq.apps.export.det.exceptions import DETConfigError
from corehq.apps.export.det.schema_generator import (
    FormDETSchemaHelper,
    generate_from_form_export_instance,
    generate_from_case_export_instance,
    generate_from_datasource_export_instance,
)
from corehq.apps.export.models import FormExportInstance, CaseExportInstance
from corehq.util.test_utils import TestFileMixin
from corehq.apps.userreports.models import DataSourceConfiguration
from corehq.apps.export.models.new import datasource_export_instance


class TestDETFCaseInstance(SimpleTestCase, TestFileMixin):
    file_path = ['data', 'det']
    root = os.path.dirname(__file__)

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.export_instance = CaseExportInstance.wrap(cls.get_json('case_export_instance'))

    @patch('corehq.apps.export.det.schema_generator._get_dd_property_types', lambda domain, case_type: {})
    def test_empty(self):
        old_tables = self.export_instance.tables
        self.export_instance.tables = []
        with self.assertRaises(DETConfigError):
            generate_from_case_export_instance(self.export_instance, BytesIO())
        self.export_instance.tables = old_tables

    @patch('corehq.apps.export.det.schema_generator._get_dd_property_types', lambda domain, case_type: {})
    def test_generate_from_case_schema(self):
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx') as tmp:
            generate_from_case_export_instance(self.export_instance, tmp)
            wb = load_workbook(filename=tmp.name)
            ws = wb.worksheets[0]
            all_data = list(ws.values)
            headings = all_data.pop(0)
            data_by_headings = [dict(zip(headings, row)) for row in all_data]
            id_row = data_by_headings.pop(0)
            self.assertEqual('case_id', id_row['Source Field'])
            self.assertEqual('id', id_row['Field'])
            domain_row = data_by_headings.pop(0)
            self.assertEqual('domain', domain_row['Source Field'])
            self.assertEqual('domain', domain_row['Field'])
            main_table = self.export_instance.selected_tables[0]
            self.assertEqual(len(main_table.selected_columns), len(data_by_headings))

            expected_paths = {
                "_id": "id",
                "activity_name": "properties.activity_name",
                "closed": "closed",
                "closed_by": "properties.closed_by",
                "closed_on": "date_closed",
                "event_date": "properties.event_date",
                "event_duration": "properties.event_duration",
                "event_score": "properties.event_score",
                "indices.activity": "indices.parent.case_id",
                "indices.foo": "indices.sibling.case_id",
                "location": "properties.location",
                "modified_on": "properties.modified_on",
                "name": "properties.case_name",
                "number": "properties.number",
                "opened_by": "opened_by",
                "opened_on": "properties.date_opened",
                "owner_id": "properties.owner_id",
                "user_id": "user_id"
            }
            for i, input_column in enumerate(main_table.selected_columns):
                self.assertEqual(input_column.label, data_by_headings[i]['Field'])
                self.assertEqual(expected_paths[input_column.item.readable_path],
                                 data_by_headings[i]['Source Field'])

            # test individual fields / types
            data_by_headings_by_source_field = {
                row['Source Field']: row for row in data_by_headings
            }
            self.assertEqual('str2date', data_by_headings_by_source_field['date_closed']['Map Via'])
            self.assertEqual(None, data_by_headings_by_source_field['properties.event_date']['Map Via'])

            # note: subtables not supported

    @patch('corehq.apps.export.det.schema_generator._get_dd_property_types')
    def test_case_schema_data_dictionary_support(self, dd_property_type_mock):
        dd_property_type_mock.return_value = {
            'event_date': 'date',
            'event_duration': 'number',
        }
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx') as tmp:
            generate_from_case_export_instance(self.export_instance, tmp)
            wb = load_workbook(filename=tmp.name)
            ws = wb.worksheets[0]
            all_data = list(ws.values)
            headings = all_data.pop(0)
            data_by_headings = [dict(zip(headings, row)) for row in all_data]
            # test individual fields / types
            data_by_headings_by_source_field = {
                row['Source Field']: row for row in data_by_headings
            }
            self.assertEqual('str2date', data_by_headings_by_source_field['properties.event_date']['Map Via'])
            self.assertEqual('str2num', data_by_headings_by_source_field['properties.event_duration']['Map Via'])
            # ensure defaults still work
            self.assertEqual(None, data_by_headings_by_source_field['properties.event_score']['Map Via'])
            self.assertEqual('str2date', data_by_headings_by_source_field['date_closed']['Map Via'])


class TestDETFormInstance(SimpleTestCase, TestFileMixin):
    file_path = ['data', 'det']
    root = os.path.dirname(__file__)

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.export_instance = FormExportInstance.wrap(cls.get_json('form_export_instance'))

    def test_empty(self):
        old_tables = self.export_instance.tables

        def _cleanup():
            self.export_instance.tables = old_tables

        self.addCleanup(_cleanup)

        self.export_instance.tables = []
        with self.assertRaises(DETConfigError):
            generate_from_form_export_instance(self.export_instance, BytesIO())

    def test_generate_from_form_schema(self):
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx') as tmp:
            generate_from_form_export_instance(self.export_instance, tmp)
            wb = load_workbook(filename=tmp.name)
            ws = wb.worksheets[0]
            all_data = list(ws.values)
            headings = all_data.pop(0)
            data_by_headings = [dict(zip(headings, row)) for row in all_data]
            id_row = data_by_headings.pop(0)
            self.assertEqual('id', id_row['Source Field'])
            self.assertEqual('id', id_row['Field'])
            domain_row = data_by_headings.pop(0)
            self.assertEqual('domain', domain_row['Source Field'])
            self.assertEqual('domain', domain_row['Field'])
            main_table = self.export_instance.selected_tables[0]
            self.assertEqual(len(main_table.selected_columns), len(data_by_headings))
            # basic sanity check
            for i, input_column in enumerate(main_table.selected_columns):
                self.assertEqual(input_column.label, data_by_headings[i]['Field'])
                self.assertEqual(FormDETSchemaHelper.transform_path(input_column.item.readable_path),
                                 data_by_headings[i]['Source Field'])

            # test individual fields / types
            data_by_headings_by_field = {
                row['Field']: row for row in data_by_headings
            }
            self.assertEqual('form.@xmlns', data_by_headings_by_field['@xmlns']['Source Field'])

            data_by_headings_by_source_field = {
                row['Source Field']: row for row in data_by_headings
            }
            self.assertEqual('str2date', data_by_headings_by_source_field['received_on']['Map Via'])
            self.assertEqual('str2date', data_by_headings_by_source_field['form.event_date']['Map Via'])
            self.assertEqual('str2num', data_by_headings_by_source_field['form.event_duration_minutes']['Map Via'])


class TestDETFormInstanceWithRepeat(SimpleTestCase, TestFileMixin):
    file_path = ['data', 'det']
    root = os.path.dirname(__file__)

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.export_instance = FormExportInstance.wrap(cls.get_json('form_export_instance_with_repeat'))

    def test_main_table_not_selected(self):
        self.export_instance.tables[0].selected = False

        def _cleanup():
            self.export_instance.tables[0].selected = True

        self.addCleanup(_cleanup)
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx') as tmp:
            generate_from_form_export_instance(self.export_instance, tmp)
            wb = load_workbook(filename=tmp.name)
            self.assertEqual(1, len(wb.worksheets))
            repeat_ws = wb.worksheets[0]
            self._check_repeat_worksheet(repeat_ws)

    def test_generate_from_form_schema(self):
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx') as tmp:
            generate_from_form_export_instance(self.export_instance, tmp)
            wb = load_workbook(filename=tmp.name)
            repeat_ws = wb.worksheets[1]
            self._check_repeat_worksheet(repeat_ws)

    def _check_repeat_worksheet(self, repeat_ws):
        repeat_data = list(repeat_ws.values)
        headings = repeat_data.pop(0)
        rows_by_headings = [dict(zip(headings, row)) for row in repeat_data]
        self.assertEqual('form.form.children[*]', rows_by_headings[0]['Data Source'])
        self.assertEqual('child_name', rows_by_headings[1]['Source Field'])
        self.assertEqual('child_dob', rows_by_headings[2]['Source Field'])


class TestDatasourceInstance(SimpleTestCase, TestFileMixin):

    domain = "test-domain"

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.config = DataSourceConfiguration(**cls._sample_data_source_dict())

    def test_generate_from_datasource_export_instance(self):
        export_instance = datasource_export_instance(self.config)

        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx') as tmp:
            generate_from_datasource_export_instance(export_instance, tmp)
            wb = load_workbook(filename=tmp.name)

            self.assertEqual(len(wb.worksheets), 1)
            ws = wb.worksheets[0]
            ws_data = list(ws.values)
            headings = ws_data.pop(0)
            data_by_headings = [dict(zip(headings, row)) for row in ws_data]

            self.assertEqual(data_by_headings[0]['Data Source'], "ucr")
            self.assertEqual(data_by_headings[1]['Source Field'], "doc_id")
            self.assertEqual(data_by_headings[2]['Source Field'], "inserted_at")
            self.assertEqual(data_by_headings[2]['Map Via'], "str2date")
            self.assertEqual(data_by_headings[3]['Source Field'], "name_string")
            self.assertEqual(data_by_headings[4]['Source Field'], "computed_owner_name")
            self.assertEqual(data_by_headings[5]['Source Field'], "age_string")
            self.assertEqual(data_by_headings[6]['Source Field'], "closed_string")

    @classmethod
    def _sample_data_source_dict(cls):
        return {
            'domain': cls.domain,
            'table_id': 'table_id',
            'display_name': 'Test datasource',
            'referenced_doc_type': 'CommCareCase',
            'configured_filter': {
                'type': 'boolean_expression',
                'operator': 'eq',
                'expression': {
                    'type': 'property_name',
                    'property_name': 'type',
                    'datatype': None
                },
                'property_value': 'case',
                'comment': None
            },
            'configured_indicators': [
                {
                    'type': 'expression',
                    'column_id': 'name_string',
                    'datatype': 'string',
                    'display_name': 'name',
                    'expression': {
                        'type': 'property_name',
                        'property_name': 'name',
                        'datatype': None
                    },
                    'is_nullable': True,
                    'is_primary_key': False,
                    'create_index': False,
                    'transform': {},
                    'comment': None
                },
                {
                    'datatype': 'string',
                    'type': 'expression',
                    'column_id': 'computed_owner_name',
                    'expression': {
                        'type': 'property_name',
                        'property_name': 'owner_id',
                        'datatype': 'string'
                    },
                    'is_nullable': True,
                    'is_primary_key': False,
                    'create_index': False,
                    'transform': {},
                    'display_name': None,
                    'comment': None
                },
                {
                    'type': 'expression',
                    'column_id': 'age_string',
                    'datatype': 'string',
                    'display_name': 'age',
                    'expression': {
                        'type': 'property_name',
                        'property_name': 'age',
                        'datatype': None
                    },
                    'is_nullable': True,
                    'is_primary_key': False,
                    'create_index': False,
                    'transform': {},
                    'comment': None
                },
                {
                    'type': 'expression',
                    'column_id': 'closed_string',
                    'datatype': 'string',
                    'display_name': 'closed',
                    'expression': {
                        'type': 'property_name',
                        'property_name': 'closed',
                        'datatype': None
                    },
                    'is_nullable': True,
                    'is_primary_key': False,
                    'create_index': False,
                    'transform': {},
                    'comment': None
                }
            ],
        }
