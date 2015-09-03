from collections import namedtuple
from couchdbkit import ResourceNotFound
from openpyxl.utils.exceptions import InvalidFileException
from corehq.apps.fixtures.exceptions import FixtureUploadError, ExcelMalformatException, \
    DuplicateFixtureTagException, FixtureAPIException
from django.core.validators import ValidationError
from django.utils.translation import ugettext as _, ugettext_noop
from corehq.apps.fixtures.models import FixtureTypeField, FixtureDataType, FixtureDataItem, FixtureItemField, \
    FieldList
from corehq.apps.users.bulkupload import GroupMemoizer
from corehq.apps.users.models import CommCareUser
from corehq.apps.users.util import normalize_username
from corehq.util.spreadsheets.excel import WorksheetNotFound, \
    WorkbookJSONReader
from dimagi.utils.couch.bulk import CouchTransaction
from soil import DownloadBase
from corehq.apps.locations.models import SQLLocation


DELETE_HEADER = "Delete(Y/N)"
FAILURE_MESSAGES = {
    "has_no_column": ugettext_noop(
        "Workbook 'types' has no column '{column_name}'."
    ),
    "neither_fields_nor_attributes": ugettext_noop(
        "Lookup-tables can not have empty fields and empty properties on items. table_id '{tag}' has no fields and no properties"
    ),
    "duplicate_tag": ugettext_noop(
        "Lookup-tables should have unique 'table_id'. There are two rows with table_id '{tag}' in 'types' sheet."
    ),
    "has_no_field_column": ugettext_noop(
        "Excel-sheet '{tag}' does not contain the column '{field}' "
        "as specified in its 'types' definition"
    ),
    "has_extra_column": ugettext_noop(
        "Excel-sheet '{tag}' has an extra column"
        "'{field}' that's not defined in its 'types' definition"
    ),
    "wrong_property_syntax": ugettext_noop(
        "Properties should be specified as 'field 1: property 1'. In 'types' sheet, "
        "'{prop_key}' is not correctly formatted"
    ),
    "sheet_has_no_property": ugettext_noop(
        "Excel-sheet '{tag}' does not contain property "
        "'{property}' of the field '{field}' as specified in its 'types' definition"
    ),
    "sheet_has_extra_property": ugettext_noop(
        "Excel-sheet '{tag}'' has an extra property "
        "'{property}' for the field '{field}' that's not defined in its 'types' definition. "
        "Re-check the formatting"
    ),
    "invalid_field_with_property": ugettext_noop(
        "Fields with attributes should be numbered as 'field: {field} integer"
    ),
    "invalid_property": ugettext_noop(
        "Attribute should be written as '{field}: {prop} interger'"
    ),
    "wrong_field_property_combos": ugettext_noop(
        "Number of values for field '{field}' and attribute '{prop}' should be same"
    ),
    "replace_with_UID": ugettext_noop(
        "Rows shouldn't contain UIDs while using replace option. Excel sheet '{tag}' contains UID in a row."
    ),
}


class FixtureUploadResult(object):
    """
    Helper structure for handling the results of a fixture upload.
    """
    def __init__(self):
        self.success = True
        self.unknown_groups = []
        self.unknown_users = []
        self.messages = []
        self.errors = []
        self.number_of_fixtures = 0


class FixtureTableDefinition(object):

    def __init__(self, table_id, fields, item_attributes, is_global, uid, delete):
        self.table_id = table_id
        self.fields = fields
        self.item_attributes = item_attributes
        self.is_global = is_global
        self.uid = uid
        self.delete = delete

    @classmethod
    def from_row(cls, row_dict):
        tag = row_dict.get('table_id') or row_dict.get('tag')
        if tag is None:
            raise ExcelMalformatException(_(FAILURE_MESSAGES['has_no_column']).format(column_name='table_id'))

        field_names = row_dict.get('field')
        item_attributes = row_dict.get('property')

        if field_names is None and item_attributes is None:
            raise ExcelMalformatException(_(FAILURE_MESSAGES['neither_fields_nor_attributes']).format(tag=tag))

        field_names = [] if field_names is None else field_names
        item_attributes = [] if item_attributes is None else item_attributes

        def _get_field_properties(prop_key):
            if row_dict.has_key(prop_key):
                try:
                    properties = row_dict[prop_key]["property"]
                    assert isinstance(properties, list)
                except (KeyError, AssertionError):
                    error_message = _(FAILURE_MESSAGES["wrong_property_syntax"]).format(
                        prop_key=prop_key,
                    )
                    raise ExcelMalformatException(error_message)
                else:
                    return properties
            else:
                return []

        fields = [
            FixtureTypeField(
                field_name=field,
                properties=_get_field_properties('field {count}'.format(count=i + 1))
            ) for i, field in enumerate(field_names)
        ]

        return FixtureTableDefinition(
            table_id=tag,
            fields=fields,
            item_attributes=item_attributes,
            is_global=row_dict.get('is_global', False),
            uid=row_dict.get('UID'),
            delete=(row_dict.get(DELETE_HEADER) or '').lower() == 'y',
        )


class FixtureWorkbook(object):
    """
    Helper class for working with the fixture workbook
    """

    def __init__(self, file_or_filename):
        try:
            self.workbook = WorkbookJSONReader(file_or_filename)
        except AttributeError:
            raise FixtureUploadError(_("Error processing your Excel (.xlsx) file"))
        except InvalidFileException:
            raise FixtureUploadError(_("Invalid file-format. Please upload a valid xlsx file."))

    def get_types_sheet(self):
        try:
            return self.workbook.get_worksheet(title='types')
        except WorksheetNotFound as e:
            raise FixtureUploadError(_("Workbook does not contain a sheet called '%(title)s'") % {'title': e.title})

    def get_data_sheet(self, data_type):
        return self.workbook.get_worksheet(data_type.tag)

    def get_all_type_sheets(self):
        type_sheets = []
        seen_tags = set()
        for number_of_fixtures, dt in enumerate(self.get_types_sheet()):
            table_definition = FixtureTableDefinition.from_row(dt)
            if table_definition.table_id in seen_tags:
                raise DuplicateFixtureTagException(
                    _(FAILURE_MESSAGES['duplicate_tag']).format(tag=table_definition.table_id))

            seen_tags.add(table_definition.table_id)
            type_sheets.append(table_definition)
        return type_sheets

    def validate(self):
        self.get_types_sheet()
        self.get_all_type_sheets()


def safe_fixture_upload(domain, file_ref, replace, task=None):
    try:
        return do_fixture_upload(domain, file_ref, replace, task)
    except FixtureUploadError as e:
        result = FixtureUploadResult()
        result.success = False
        result.errors.append(unicode(e))
        return result

def do_fixture_upload(domain, file_ref, replace, task=None):
    workbook = get_workbook(file_ref.get_filename())
    try:
        return run_upload(domain, workbook, replace=replace, task=task)
    except WorksheetNotFound as e:
        raise FixtureUploadError(_("Workbook does not contain a sheet called '%(title)s'") % {'title': e.title})
    except ExcelMalformatException as e:
        raise FixtureUploadError(_("Uploaded excel file has following formatting-problems: '%(e)s'") % {'e': e})
    except FixtureAPIException as e:
        raise FixtureUploadError(unicode(e))
    except Exception:
        raise FixtureUploadError(_("Fixture upload failed for some reason and we have noted this failure. "
                                   "Please make sure the excel file is correctly formatted and try again."))


def validate_file_format(file_or_filename):
    """
    Does basic validation on the uploaded file. Raises a FixtureUploadError if
    something goes wrong.
    """
    workbook = FixtureWorkbook(file_or_filename)
    workbook.validate()


def get_workbook(file_or_filename):
    return FixtureWorkbook(file_or_filename)


LocationCache = namedtuple("LocationCache", "is_error location message")


def get_memoized_location(domain):
    """
    Returns a memoized location getter containing error information.
    """
    locations = {}
    def get_location(user_input):
        user_input = user_input.lower()
        if user_input not in locations:
            try:
                loc = SQLLocation.objects.get_from_user_input(domain, user_input)
                locations[user_input] = LocationCache(False, loc, None)
            except SQLLocation.DoesNotExist:
                locations[user_input] = LocationCache(True, None, _(
                    "Unknown location: '%(name)s'. But the row is "
                    "successfully added"
                ) % {'name': user_input})
            except SQLLocation.MultipleObjectsReturned:
                locations[user_input] = LocationCache(True, None, _(
                    "Multiple locations found with the name: '%(name)s'.  "
                    "Try using site code. But the row is successfully added"
                ) % {'name': user_input})
        return locations[user_input]
    return get_location


def run_upload(domain, workbook, replace=False, task=None):
    return_val = FixtureUploadResult()
    group_memoizer = GroupMemoizer(domain)
    get_location = get_memoized_location(domain)

    def diff_lists(list_a, list_b):
        set_a = set(list_a)
        set_b = set(list_b)
        not_in_b = set_a.difference(set_b)
        not_in_a = set_a.difference(set_a)
        return list(not_in_a), list(not_in_b)

    with CouchTransaction() as transaction:
        type_sheets = workbook.get_all_type_sheets()
        total_tables = len(type_sheets)
        return_val.number_of_fixtures = total_tables

        def _update_progress(table_count, item_count, items_in_table):
            if task:
                processed = table_count * 10 + (10. * item_count / items_in_table)
                DownloadBase.set_progress(task, processed, 10 * total_tables)

        for table_number, table_def in enumerate(type_sheets):
            tag = table_def.table_id
            new_data_type = FixtureDataType(
                domain=domain,
                is_global=table_def.is_global,
                tag=table_def.table_id,
                fields=table_def.fields,
                item_attributes=table_def.item_attributes
            )
            try:
                tagged_fdt = FixtureDataType.fixture_tag_exists(domain, tag)
                if tagged_fdt:
                    data_type = tagged_fdt
                # support old usage with 'UID'
                elif table_def.uid:
                    data_type = FixtureDataType.get(table_def.uid)
                else:
                    data_type = new_data_type

                if replace and data_type != new_data_type:
                    data_type.recursive_delete(transaction)
                    data_type = new_data_type

                data_type.fields = table_def.fields
                data_type.item_attributes = table_def.item_attributes
                data_type.is_global = table_def.is_global
                assert data_type.doc_type == FixtureDataType._doc_type
                if data_type.domain != domain:
                    data_type = new_data_type
                    return_val.errors.append(_("'%(UID)s' is not a valid UID. But the new type is created.") % {'UID': table_def.uid})
                if table_def.delete:
                    data_type.recursive_delete(transaction)
                    continue
            except (ResourceNotFound, KeyError) as e:
                data_type = new_data_type
            transaction.save(data_type)

            data_items = list(workbook.get_data_sheet(data_type))
            items_in_table = len(data_items)
            for sort_key, di in enumerate(data_items):
                _update_progress(table_number, sort_key, items_in_table)
                # Check that type definitions in 'types' sheet vs corresponding columns in the item-sheet MATCH
                item_fields_list = di['field'].keys() if 'field' in di else []
                not_in_sheet, not_in_types = diff_lists(item_fields_list, data_type.fields_without_attributes)
                if len(not_in_sheet) > 0:
                    error_message = _(FAILURE_MESSAGES["has_no_field_column"]).format(tag=tag, field=not_in_sheet[0])
                    raise ExcelMalformatException(error_message)
                if len(not_in_types) > 0:
                    error_message = _(FAILURE_MESSAGES["has_extra_column"]).format(tag=tag, field=not_in_types[0])
                    raise ExcelMalformatException(error_message)

                # check that this item has all the properties listed in its 'types' definition
                item_attributes_list = di['property'].keys() if 'property' in di else []
                not_in_sheet, not_in_types = diff_lists(item_attributes_list, data_type.item_attributes)
                if len(not_in_sheet) > 0:
                    error_message = _(FAILURE_MESSAGES["has_no_field_column"]).format(tag=tag, field=not_in_sheet[0])
                    raise ExcelMalformatException(error_message)
                if len(not_in_types) > 0:
                    error_message = _(FAILURE_MESSAGES["has_extra_column"]).format(tag=tag, field=not_in_types[0])
                    raise ExcelMalformatException(error_message)

                # check that properties in 'types' sheet vs item-sheet MATCH
                for field in data_type.fields:
                    if len(field.properties) > 0:
                        sheet_props = di.get(field.field_name, {})
                        sheet_props_list = sheet_props.keys()
                        type_props = field.properties
                        not_in_sheet, not_in_types = diff_lists(sheet_props_list, type_props)
                        if len(not_in_sheet) > 0:
                            error_message = _(FAILURE_MESSAGES["sheet_has_no_property"]).format(
                                tag=tag,
                                property=not_in_sheet[0],
                                field=field.field_name
                            )
                            raise ExcelMalformatException(error_message)
                        if len(not_in_types) > 0:
                            error_message = _(FAILURE_MESSAGES["sheet_has_extra_property"]).format(
                                tag=tag,
                                property=not_in_types[0],
                                field=field.field_name
                            )
                            raise ExcelMalformatException(error_message)
                        # check that fields with properties are numbered
                        if type(di['field'][field.field_name]) != list:
                            error_message = _(FAILURE_MESSAGES["invalid_field_with_property"]).format(field=field.field_name)
                            raise ExcelMalformatException(error_message)
                        field_prop_len = len(di['field'][field.field_name])
                        for prop in sheet_props:
                            if type(sheet_props[prop]) != list:
                                error_message = _(FAILURE_MESSAGES["invalid_property"]).format(
                                    field=field.field_name,
                                    prop=prop
                                )
                                raise ExcelMalformatException(error_message)
                            if len(sheet_props[prop]) != field_prop_len:
                                error_message = _(FAILURE_MESSAGES["wrong_field_property_combos"]).format(
                                    field=field.field_name,
                                    prop=prop
                                )
                                raise ExcelMalformatException(error_message)

                # excel format check should have been covered by this line. Can make assumptions about data now
                type_fields = data_type.fields
                item_fields = {}
                for field in type_fields:
                    # if field doesn't have properties
                    if len(field.properties) == 0:
                        item_fields[field.field_name] = FieldList(
                            field_list=[FixtureItemField(
                                # using unicode here, to cast ints, and multi-language strings
                                field_value=unicode(di['field'][field.field_name]),
                                properties={}
                            )]
                        )
                    else:
                        field_list = []
                        field_prop_combos = di['field'][field.field_name]
                        prop_combo_len = len(field_prop_combos)
                        prop_dict = di[field.field_name]
                        for x in range(0, prop_combo_len):
                            fix_item_field = FixtureItemField(
                                field_value=unicode(field_prop_combos[x]),
                                properties={prop: unicode(prop_dict[prop][x]) for prop in prop_dict}
                            )
                            field_list.append(fix_item_field)
                        item_fields[field.field_name] = FieldList(
                            field_list=field_list
                        )

                item_attributes = di.get('property', {})
                new_data_item = FixtureDataItem(
                    domain=domain,
                    data_type_id=data_type.get_id,
                    fields=item_fields,
                    item_attributes=item_attributes,
                    sort_key=sort_key
                )
                try:
                    if di['UID'] and not replace:
                        old_data_item = FixtureDataItem.get(di['UID'])
                    else:
                        old_data_item = new_data_item
                        pass
                    old_data_item.fields = item_fields
                    old_data_item.item_attributes = item_attributes
                    if old_data_item.domain != domain or not old_data_item.data_type_id == data_type.get_id:
                        old_data_item = new_data_item
                        return_val.errors.append(_("'%(UID)s' is not a valid UID. But the new item is created.") % {'UID': di['UID']})
                    assert old_data_item.doc_type == FixtureDataItem._doc_type
                    if di[DELETE_HEADER] == "Y" or di[DELETE_HEADER] == "y":
                        old_data_item.recursive_delete(transaction)
                        continue
                except (ResourceNotFound, KeyError):
                    old_data_item = new_data_item
                transaction.save(old_data_item)

                old_groups = old_data_item.groups
                for group in old_groups:
                    old_data_item.remove_group(group)
                old_users = old_data_item.users
                for user in old_users:
                    old_data_item.remove_user(user)
                old_locations = old_data_item.locations
                for location in old_locations:
                    old_data_item.remove_location(location)

                for group_name in di.get('group', []):
                    group = group_memoizer.by_name(group_name)
                    if group:
                        old_data_item.add_group(group, transaction=transaction)
                    else:
                        return_val.errors.append(_("Unknown group: '%(name)s'. But the row is successfully added") % {'name': group_name})

                for raw_username in di.get('user', []):
                    try:
                        username = normalize_username(str(raw_username), domain)
                    except ValidationError:
                        return_val.errors.append(_("Invalid username: '%(name)s'. Row is not added") % {'name': raw_username})
                        continue
                    user = CommCareUser.get_by_username(username)
                    if user:
                        old_data_item.add_user(user)
                    else:
                        return_val.errors.append(_("Unknown user: '%(name)s'. But the row is successfully added") % {'name': raw_username})

                for name in di.get('location', []):
                    location_cache = get_location(name)
                    if location_cache.is_error:
                        return_val.errors.append(location_cache.message)
                    else:
                        old_data_item.add_location(location_cache.location,
                                                   transaction=transaction)

    return return_val
