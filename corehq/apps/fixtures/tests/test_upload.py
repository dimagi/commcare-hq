from contextlib import contextmanager
from datetime import timedelta
from io import BytesIO
from unittest.mock import patch

from django.test import SimpleTestCase, TestCase

import openpyxl

from couchexport.export import export_raw
from couchexport.models import Format

from corehq.apps.domain.shortcuts import create_domain
from corehq.apps.fixtures.exceptions import FixtureUploadError
from corehq.apps.fixtures.models import (
    Field,
    LookupTable,
    LookupTableRow,
    LookupTableRowOwner,
    OwnerType,
    TypeField,
)
from corehq.apps.fixtures.upload import validate_fixture_file_format
from corehq.apps.fixtures.upload.failure_messages import FAILURE_MESSAGES
from corehq.apps.fixtures.upload.run_upload import _run_upload
from corehq.apps.fixtures.upload.workbook import get_workbook
from corehq.apps.groups.models import Group
from corehq.apps.locations.models import LocationType, SQLLocation
from corehq.apps.users.models import CommCareUser
from corehq.util.test_utils import generate_cases, make_make_path, new_db_connection

from dimagi.utils.couch.database import iter_docs

from ..upload import run_upload as mod

_make_path = make_make_path(__file__)


# (slug (or filename), [expected errors], file contents (None if excel file exists in repository))
validation_test_cases = [
    ('duplicate_tag', [
        "Lookup-tables should have unique 'table_id'. "
        "There are two rows with table_id 'things' in 'types' sheet.",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name'), (None, 'N', 'apple')],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1'),
            ('N', 'things', 'yes', 'name'), ('N', 'things', 'yes', 'name')
        ]
    }),
    ("invalid_table_id", [
        "table_id 'invalid table_id' should not contain spaces or special characters, or start with a number."
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name'), (None, 'N', 'apple')],
        'types': [('Delete(Y/N)', 'table_id', 'is_global?', 'field 1'), ('N', 'invalid table_id', 'yes', 'name')]
    }),
    ('multiple_errors', [
        "Excel worksheet 'level_1' does not contain the column "
        "'field: fun_fact' as specified in its 'types' definition",
        "Excel worksheet 'level_1' does not contain the column "
        "'field: other' as specified in its 'types' definition",
        "Excel worksheet 'level_2' does not contain the column "
        "'field: other' as specified in its 'types' definition",
        "There's no sheet for type 'level_3' in 'types' sheet. "
        "There must be one sheet per row in the 'types' sheet.",
    ], {
        'level_names': [
            ('UID', 'Delete(Y/N)', 'field: level_1', 'field: level_2'),
            (None, 'N', 'State', 'County')
        ],
        'level_2': [
            ('UID', 'Delete(Y/N)', 'field: id', 'name: lang 1', 'field: name 1',
                                                'name: lang 2', 'field: name 2', 'field: level_1'),
            (None, 'N', 'barnstable', 'en', 'Barnstable', 'fra', 'Barnstable', 'MA'),
            (None, 'N', 'berkshire', 'en', 'Berkshire', 'fra', 'Berkshire', 'MA'),
            (None, 'N', 'bristol', 'en', 'Bristol', 'fra', 'Bristol', 'MA'),
            (None, 'N', 'dukes', 'en', 'Dukes', 'fra', 'Dukes', 'MA')
        ],
        'level_1': [
            ('UID', 'Delete(Y/N)', 'field: id', 'name: lang 1', 'field: name 1',
                                                'name: lang 2', 'field: name 2', 'field: country'),
            (None, 'N', 'MA', 'en', 'Massachusetts', 'fra', 'Massachusetts', 'USA')
        ],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?',
             'field 1', 'field 2', 'field 3', 'field 4', 'field 5', 'field 2 : property 1'),
            ('N', 'level_1', 'yes', 'id', 'name', 'country', 'other', 'fun_fact', 'lang'),
            ('N', 'level_2', 'yes', 'id', 'name', 'level_1', 'other', None, 'lang'),
            ('N', 'level_3', 'yes', 'id', 'name', 'level_2', 'other', None, 'lang')
        ]
    }),
    ('type_has_no_sheet', [
        "There's no sheet for type 'things' in 'types' sheet. "
        "There must be one sheet per row in the 'types' sheet.",
    ], {
        'types': [('Delete(Y/N)', 'table_id', 'is_global?', 'field 1'), ('N', 'things', 'yes', 'name')]
    }),
    ('has_no_field_column', [
        "Excel worksheet 'things' does not contain the column 'field: name' "
        "as specified in its 'types' definition",
    ], {
        'things': [('UID', 'Delete(Y/N)'), (None, 'N')],
        'types': [('Delete(Y/N)', 'table_id', 'is_global?', 'field 1'), ('N', 'things', 'yes', 'name')]
    }),
    ('has_no_field_column_extra_rows', [
        "Excel worksheet 'things' does not contain the column 'field: name' "
        "as specified in its 'types' definition",
    ], {
        'things': [('UID', 'Delete(Y/N)'), (None, 'N')],
        'types': [('Delete(Y/N)', 'table_id', 'is_global?', 'field 1'), ('N', 'things', 'yes', 'name')]
    }),
    ('has_extra_column', [
        "Excel worksheet 'things' has an extra column"
        "'field: fun_fact' that's not defined in its 'types' definition",
    ], {
        'things': [
            ('UID', 'Delete(Y/N)', 'field: name', 'field: fun_fact'),
            (None, 'N', 'apple', 'a day keeps the doctor away')
        ],
        'types': [('Delete(Y/N)', 'table_id', 'is_global?', 'field 1'), ('N', 'things', 'yes', 'name')]
    }),
    ('sheet_has_no_property', [
        "Excel worksheet 'things' does not contain property "
        "'lang' of the field 'name' as specified in its 'types' definition",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name 1'), (None, 'N', 'apple')],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('sheet_has_extra_property', [
        "Excel worksheet 'things' has an extra property "
        "'style' for the field 'name' that's not defined in its 'types' definition. "
        "Re-check the formatting",
    ], {
        'things': [
            ('UID', 'Delete(Y/N)', 'field: name 1', 'name: lang 1', 'name: style 1'),
            (None, 'N', 'apple', 'en', 'lowercase')
        ],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('invalid_field_with_property', [
        "Fields with attributes should be numbered as 'field: name integer'",
        # also triggers wrong_field_property_combos
        "Number of values for field 'name' and attribute 'lang' should be same",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name', 'name: lang 1'), (None, 'N', 'apple', 'en')],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('invalid_property', [
        "Attribute should be written as 'name: lang integer'",
        # also triggers wrong_field_property_combos
        "Number of values for field 'name' and attribute 'lang' should be same",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name 1', 'name: lang'), (None, 'N', 'apple', 'en')],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('wrong_field_property_combos', [
        "Number of values for field 'name' and attribute 'lang' should be same",
    ], {
        'things': [
            ('UID', 'Delete(Y/N)', 'field: name 1', 'name: lang 1', 'field: name 2'),
            (None, 'N', 'apple', 'en', 'malum')
        ],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('has_no_column', [
        "Workbook 'types' has no column 'table_id'.",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name'), (None, 'N', 'apple')],
        'types': [('Delete(Y/N)', 'is_global?', 'field 1'), ('N', 'yes', 'name')]}),
    ('neither_fields_nor_attributes', [
        "Lookup-tables can not have empty fields and empty properties on items. "
        "table_id 'things' has no fields and no properties",
    ], {
        'things': [('UID', 'Delete(Y/N)'), (None, 'N')],
        'types': [('Delete(Y/N)', 'table_id', 'is_global?'), ('N', 'things', 'yes')]}),
    ('invalid_field_syntax', [
        "In Excel worksheet 'things', field 'name' should be numbered as 'field: name integer",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name', 'name'), (None, 'N', 'apple', 'en')],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('wrong_property_syntax', [
        "Properties should be specified as 'field 1: property 1'. In 'types' sheet, "
        "'field 1' is not correctly formatted"
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name'), (None, 'N', 'apple')],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('invalid_field_name_numerical', [
        "Error in 'types' sheet for 'field 1', '100'. "
        "Field names should be strings, not numbers",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name'), (None, 'N', 'apple')],
        'types': [('Delete(Y/N)', 'table_id', 'is_global?', 'field 1'), ('N', 'things', 'yes', 100)]
    }),
    ('not_excel_file', [
        "Upload failed! Please make sure you are using a valid Excel 2007 or later (.xlsx) file. " \
        "Error details: \"There is no item named '[Content_Types].xml' in the archive\".",
    ], None),
    ('no_types_sheet', [
        "Workbook does not contain a sheet called types",
    ], {'things': [('UID', 'Delete(Y/N)', 'field: name'), (None, 'N', 'apple')]}),
    ('wrong_index_syntax', [
        "'field 1' is not correctly formatted in 'types' sheet. Whether a field is indexed should be specified "
        "as 'field 1: is_indexed?'. Its value should be 'yes' or 'no'.",
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name', 'name'), (None, 'N', 'apple', 'en')],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1', 'field 1: is_indexed'),
            ('N', 'things', 'yes', 'name', 'lang', 'a')
        ]
    }),
    ('field_type_error', [
        "Fields with attributes should be numbered as 'field: name integer'"
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name', 'name: lang 1'), (None, 'N', 1, 1)],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    }),
    ('property_type_error', [
        "Attribute should be written as 'name: lang integer'"
    ], {
        'things': [('UID', 'Delete(Y/N)', 'field: name 1', 'name: lang'), (None, 'N', 1, 1)],
        'types': [
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: property 1'),
            ('N', 'things', 'yes', 'name', 'lang')
        ]
    })
]


class TestValidation(SimpleTestCase):
    maxDiff = None

    @generate_cases(validation_test_cases)
    def test_validation(self, filename, error_messages, file_contents):
        if file_contents:
            workbook = openpyxl.Workbook()
            for title, rows in file_contents.items():
                if title == 'types':
                    sheet = workbook.create_sheet(title, 0)
                else:
                    sheet = workbook.create_sheet(title)
                for row in rows:
                    sheet.append(row)
            upload_file = BytesIO()
            workbook.save(upload_file)
            upload_file.seek(0)
        else:
            upload_file = _make_path('test_upload', '{}.xlsx'.format(filename))
        if error_messages:
            with self.assertRaises(FixtureUploadError) as context:
                validate_fixture_file_format(upload_file)
            self.assertEqual(context.exception.errors, error_messages)
        else:
            # assert doesn't raise anything
            validate_fixture_file_format(upload_file)

    def test_comprehensiveness(self):
        to_test = set(FAILURE_MESSAGES.keys())
        tested = set([validation for validation, _, _ in validation_test_cases])
        untested = to_test - tested
        self.assertTrue(
            not untested,
            "Some fixture upload errors are still untested.\n\n"
            "You have to write a test for the following fixture upload errors:\n{}"
            .format('\n'.join(untested)))


class Args(tuple):
    def __repr__(self):
        return f'({self[0]})'


class TestFixtureWorkbook(SimpleTestCase):

    def test_indexed_field(self):
        workbook = get_workbook(_make_path('test_upload', 'ok.xlsx'))
        type_sheets = workbook.get_all_type_sheets()
        indexed_field = type_sheets[0].fields[0]
        self.assertEqual(indexed_field.field_name, 'name')
        self.assertTrue(indexed_field.is_indexed)

    @generate_cases([Args(a) for a in [
        (
            'add to beginning',
            [(None, 'N', 'apple'), ('b', 'N', 'banana'), ('c', 'N', 'coconut')],
            {'b': 0, 'c': 1},
            [(0, 'apple'), (1, 'banana'), (2, 'coconut')],
        ), (
            'add to end',
            [('a', 'N', 'apple'), ('b', 'N', 'banana'), (None, 'N', 'coconut')],
            {'a': 0, 'b': 1},
            [(0, 'apple'), (1, 'banana'), (2, 'coconut')],
        ), (
            'new first item',
            [(None, 'N', 'peach'), ('b', 'N', 'banana'), ('c', 'N', 'coconut')],
            {'a': 0, 'b': 1, 'c': 2},
            [(0, 'peach'), (1, 'banana'), (2, 'coconut')],
        ), (
            'no change',
            [('a', 'N', 'apple'), ('b', 'N', 'banana'), ('c', 'N', 'coconut')],
            {'a': 0, 'b': 1, 'c': 2},
            [(0, 'apple'), (1, 'banana'), (2, 'coconut')],
        ), (
            'rearrange',
            [('b', 'N', 'banana'), ('c', 'N', 'coconut'), ('a', 'N', 'apple')],
            {'a': 0, 'b': 1, 'c': 2},
            [(1, 'banana'), (2, 'coconut'), (3, 'apple')],
        ), (
            'add and rearrange',
            [(x, 'N', x) for x in 'abcdefghi'],
            {'b': 2, 'd': 5, 'f': 3, 'h': 8, 'i': 16},
            [
                (0, 'a'),
                (2, 'b'),
                (3, 'c'),
                (5, 'd'),
                (6, 'e'),
                (7, 'f'),
                (8, 'g'),
                (9, 'h'),
                (16, 'i'),
            ],
        )
    ]])
    def test_iter_rows(self, name, sheet_rows, old_keys, expected_rows):
        book = self.get_workbook(sheet_rows)
        table, = book.iter_tables('test')
        rows = book.iter_rows(table, old_keys)
        actual_rows = [(r.sort_key, row_name(r)) for r in rows]
        self.assertEqual(actual_rows, expected_rows)

    def get_workbook(self, rows):
        headers = TestFixtureUpload.headers
        data = TestFixtureUpload.make_rows(rows)
        return TestFixtureUpload.get_workbook_from_data(headers, data)


def row_name(item):
    return item.fields['name'][0].value


class TestFixtureUpload(TestCase):
    do_upload = _run_upload

    headers = (
        (
            'types',
            ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1', 'field 1: is_indexed?')
        ),
        (
            'things',
            ('UID', 'Delete(Y/N)', 'field: name')
        )
    )

    @staticmethod
    def make_rows(item_rows):
        # given a list of fixture-items, return formatted excel rows
        return (
            (
                'types',
                [('N', 'things', 'yes', 'name', 'yes')]
            ),
            (
                'things',
                item_rows
            )
        )

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.domain = 'fixture-upload-test'
        cls.project = create_domain(cls.domain)
        cls.addClassCleanup(cls.project.delete)

    @staticmethod
    def get_workbook_from_data(headers, rows):
        file = BytesIO()
        export_raw(headers, rows, file, format=Format.XLS_2007)
        return get_workbook(file)

    def upload(self, rows_or_workbook, **kw):
        if isinstance(rows_or_workbook, (list, tuple)):
            data = self.make_rows(rows_or_workbook)
            workbook = self.get_workbook_from_data(self.headers, data)
        else:
            workbook = rows_or_workbook
        domain = kw.pop("domain", self.domain)
        return type(self).do_upload(domain, workbook, **kw)

    def get_table(self, domain=None):
        try:
            return LookupTable.objects.by_domain_tag(domain or self.domain, 'things')
        except LookupTable.DoesNotExist:
            return None

    def get_rows(self, transform=row_name, *, domain=None):
        # return list of field values of fixture table 'things'
        def sort_key(item):
            return item.sort_key, transform(item)

        items = LookupTableRow.objects.iter_rows(domain or self.domain, tag='things')
        if transform is None:
            return list(items)
        return [transform(item) for item in sorted(items, key=sort_key)]

    def test_row_addition(self):
        # upload and then reupload with addition of a new fixture-item should create new items
        self.upload([(None, 'N', 'apple')])
        self.assertEqual(self.get_rows(), ['apple'])

        # reupload with additional row
        apple_id = self.get_rows(None)[0].id.hex
        self.upload([(apple_id, 'N', 'apple'), (None, 'N', 'orange')])
        self.assertEqual(self.get_rows(), ['apple', 'orange'])

    def test_replace_rows(self):
        self.upload([(None, 'N', 'apple')])

        self.upload([(None, 'N', 'orange'), (None, 'N', 'banana')], replace=True)
        self.assertEqual(self.get_rows(), ['orange', 'banana'])

    def test_rows_with_no_changes(self):
        self.upload([(None, 'N', 'apple')])
        rows = self.get_rows(None)
        self.assertEqual(len(rows), 1, rows)
        table_id = self.get_table().id
        apple_id = rows[0].id

        self.upload([(apple_id, 'N', 'apple')])
        self.assertEqual(self.get_table().id, table_id)
        self.assertEqual(self.get_rows(None)[0].id, apple_id)

    def test_replace_duplicate_rows(self):
        self.upload([
            (None, 'N', 'apple'),
            (None, 'N', 'apple'),
            (None, 'N', 'apple'),
        ])
        self.assertEqual(self.get_rows(), ['apple', 'apple', 'apple'])

        self.upload([(None, 'N', 'apple')], replace=True)
        self.assertEqual(self.get_rows(), ['apple'])

    def test_rearrange_rows(self):
        self.upload([
            (None, 'N', 'apple'),
            (None, 'N', 'banana'),
            (None, 'N', 'coconut'),
        ])
        self.assertEqual(self.get_rows(), ['apple', 'banana', 'coconut'])

        ids = {row_name(r): r.id.hex for r in self.get_rows(None)}
        self.upload([
            (ids['banana'], 'N', 'banana'),
            (ids['coconut'], 'N', 'coconut'),
            (ids['apple'], 'N', 'apple'),
        ])
        self.assertEqual(self.get_rows(), ['banana', 'coconut', 'apple'])

    def test_add_rows_without_replace(self):
        self.upload([(None, 'N', 'orange')])

        self.upload([(None, 'N', 'apple'), (None, 'N', 'banana')])
        self.assertEqual(self.get_rows(), ['apple', 'orange', 'banana'])

    def test_delete_table(self):
        self.upload([(None, 'N', 'apple')])
        row_ids = {r.id for r in self.get_rows(None)}

        data = [
            ('types', [('Y', 'things', 'yes', 'name', 'yes')]),
            ('things', [(None, 'N', 'apple')]),
        ]
        self.upload(self.get_workbook_from_data(self.headers, data))

        self.assertIsNone(self.get_table())
        overlap = LookupTableRow.objects.filter(id__in=list(row_ids)).count()
        self.assertFalse(overlap)

    def test_delete_missing_table(self):
        data = [
            ('types', [('Y', 'things', 'yes', 'name', 'yes')]),
            ('things', [(None, 'N', 'apple')]),
        ]
        self.upload(self.get_workbook_from_data(self.headers, data))

        self.assertIsNone(self.get_table())

    def test_update_table(self):
        def part(item):
            return item.fields['part'][0].value

        self.upload([(None, 'N', 'apple')])
        apple_id = {row_name(r): r.id.hex for r in self.get_rows(None)}["apple"]

        headers = (
            self.headers[0],
            ('things', ('UID', 'Delete(Y/N)', 'field: part')),
        )
        data = [
            ('types', [('N', 'things', 'yes', 'part', 'yes')]),
            ('things', [(apple_id, 'N', 'branch')]),
        ]
        self.upload(self.get_workbook_from_data(headers, data))
        self.assertEqual(self.get_rows(part), ['branch'])

    def test_upload_sheet_with_missing_UID_column(self):
        self.upload([(None, 'N', 'apple')])

        headers = (
            self.headers[0],
            ('things', ('Delete(Y/N)', 'field: name')),
        )
        data = [
            ('types', [('N', 'things', 'yes', 'name', 'yes')]),
            ('things', [('N', 'apple')]),
            ('things', [('N', 'orange')]),
        ]
        self.upload(self.get_workbook_from_data(headers, data))
        self.assertEqual(self.get_rows(), ['apple', 'orange'])

    def test_delete_row(self):
        self.upload([(None, 'N', 'apple'), (None, 'N', 'orange')])
        ids = {row_name(r): r.id.hex for r in self.get_rows(None)}

        self.upload([
            (ids['apple'], 'Y', 'apple'),
            (ids['orange'], 'N', 'orange'),
        ])
        self.assertEqual(self.get_rows(), ['orange'])

    def test_upload_progress_and_result(self):
        task = FakeTask()
        with patch.object(mod, "timedelta", lambda **k: timedelta()):
            result = self.upload([
                (None, 'N', 'apple'),
                (None, 'N', 'banana'),
                (None, 'N', 'coconut'),
                (None, 'N', 'doughnut'),
            ], task=task)
        self.assertEqual(task.states, [
            {'state': 'PROGRESS', 'meta': {'current': 0.0, 'total': 10}},
            {'state': 'PROGRESS', 'meta': {'current': 2.5, 'total': 10}},
            {'state': 'PROGRESS', 'meta': {'current': 5.0, 'total': 10}},
            {'state': 'PROGRESS', 'meta': {'current': 7.5, 'total': 10}},
            {'state': 'PROGRESS', 'meta': {'current': 7.5, 'total': 10}},
        ])
        self.assertEqual(result.number_of_fixtures, 1)
        self.assertTrue(result.success)

    def test_upload_progress_with_zero_tables(self):
        workbook = self.get_workbook_from_data(self.headers, [])
        task = FakeTask()
        with patch.object(mod, "timedelta", lambda **k: timedelta()):
            result = self.upload(workbook, task=task)
        self.assertEqual(result.errors, [])
        self.assertIsNone(self.get_table())

    def test_table_uid_conflict(self):
        result = self.upload_table_with_uid_conflict()
        self.assertFalse(result.errors)

    def upload_table_with_uid_conflict(self):
        self.upload([(None, 'N', 'apple')], domain=self.domain + "2")
        bad_uid = self.get_table(self.domain + "2").id

        headers = (
            ('types', ('UID', 'Delete(Y/N)', 'table_id', 'is_global?', 'field 1')),
            ('things', ('UID', 'Delete(Y/N)', 'field: name')),
        )
        data = [
            ('types', [(bad_uid, 'N', 'things', 'yes', 'name')]),
            ('things', [(None, 'N', 'apple')]),
        ]
        result = self.upload(self.get_workbook_from_data(headers, data))
        self.assertNotEqual(self.get_table().id, bad_uid)
        return result

    def test_row_uid_conflict(self):
        result = self.upload_row_with_uid_conflict()
        self.assertFalse(result.errors)

    def upload_row_with_uid_conflict(self):
        domain2 = self.domain + "2"
        self.upload([(None, 'N', 'apple')], domain=domain2)
        bad_uid, = [r.id.hex for r in self.get_rows(None, domain=domain2)]

        result = self.upload([(bad_uid, 'N', 'apple')])
        new_uid, = [r.id.hex for r in self.get_rows(None)]
        self.assertNotEqual(new_uid, bad_uid)
        return result

    def test_error_on_non_global_table(self):
        data = [
            ('types', [('N', 'things', 'no', 'name', 'yes')]),
            ('things', [(None, 'N', 'apple')]),
        ]
        workbook = self.get_workbook_from_data(self.headers, data)
        result = self.upload(workbook, skip_orm=True)
        self.assertEqual(result.errors, ["type things is not defined as global"])
        self.assertIsNone(self.get_table())

    def test_ownerships_ignored_on_skip_orm(self):
        user = CommCareUser.create(
            self.domain, f"user@{self.domain}.commcarehq.org", "pass", None, None)
        self.addCleanup(user.delete, self.domain, deleted_by=None)
        region = LocationType(domain=self.domain, name="region", code="region")
        region.save()
        SQLLocation(domain=self.domain, name="loc", location_type=region).save()

        headers = TestLookupTableOwnershipUpload.headers
        data = [
            ('types', [('N', 'things', 'yes', 'name')]),
            ('things', [(None, 'N', 'apple', 'user', None, 'loc')]),
        ]
        workbook = self.get_workbook_from_data(headers, data)
        result = self.upload(workbook, skip_orm=True)
        apple_id, = [r.id for r in self.get_rows(None)]

        ownerships = list(LookupTableRowOwner.objects.filter(domain=self.domain, row_id=apple_id))
        self.assertFalse(ownerships)
        self.assertFalse(result.errors)

    def test_sql_transaction(self):
        @contextmanager
        def checked_tx():
            with atomic():
                yield
                with new_db_connection(), self.assertRaises(LookupTable.DoesNotExist):
                    get_table()
                did_check.append(True)

        def get_table():
            return LookupTable.objects.get(domain=self.domain, tag='things')

        atomic = mod.atomic
        did_check = []
        with patch.object(mod, "atomic", checked_tx):
            self.upload([(None, 'N', 'apple'), (None, 'N', 'orange')])
        self.assertIsNotNone(get_table())
        self.assertTrue(did_check)

    def test_upload_should_clear_cache_on_error(self):
        def error_tx():
            raise Exception("cannot save")

        upload_error = patch.object(mod, "atomic", error_tx)
        clear_patch = patch.object(mod, "clear_fixture_cache")
        with upload_error, clear_patch as clear_cache, self.assertRaises(Exception):
            self.upload([(None, 'N', 'orange')])
        clear_cache.assert_called_once()


class TestLookupTableOwnershipUpload(TestCase):
    do_upload = _run_upload

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        publish = patch("corehq.apps.locations.document_store.publish_location_saved")
        publish.start()
        cls.addClassCleanup(publish.stop)

        cls.domain = 'fixture-upload-test'
        cls.project = create_domain(cls.domain)
        cls.addClassCleanup(cls.project.delete)

        cls.user1 = CommCareUser.create(
            cls.domain, f"user1@{cls.domain}.commcarehq.org", "pass", None, None)
        cls.addClassCleanup(cls.user1.delete, cls.domain, deleted_by=None)
        cls.user2 = CommCareUser.create(
            cls.domain, f"user2@{cls.domain}.commcarehq.org", "pass", None, None)
        cls.addClassCleanup(cls.user2.delete, cls.domain, deleted_by=None)
        cls.user3 = CommCareUser.create(
            cls.domain, f"3@{cls.domain}.commcarehq.org", "pass", None, None)
        cls.addClassCleanup(cls.user3.delete, cls.domain, deleted_by=None)

        # group names are case sensitive, user and location names are not
        cls.group1 = Group(domain=cls.domain, name="G1", users=[cls.user1._id])
        cls.group1.save()
        cls.addClassCleanup(cls.group1.delete)
        cls.group2 = Group(domain=cls.domain, name="g2", users=[cls.user1._id])
        cls.group2.save()
        cls.addClassCleanup(cls.group2.delete)
        cls.group3 = Group(domain=cls.domain, name="3", users=[cls.user1._id])
        cls.group3.save()
        cls.addClassCleanup(cls.group3.delete)

        cls.region = LocationType(domain=cls.domain, name="region", code="region")
        cls.region.save()
        cls.loc1 = SQLLocation(domain=cls.domain, name="loc1", location_type=cls.region)
        cls.loc1.save()
        cls.loc2 = SQLLocation(domain=cls.domain, name="loc2", location_type=cls.region)
        cls.loc2.save()
        cls.loc3 = SQLLocation(domain=cls.domain, name="3", location_type=cls.region)
        cls.loc3.save()

    def test_row_ownership(self):
        self.upload([(None, 'N', 'apple', 'user1', 'G1', 'loc1')])
        self.assertEqual(self.get_rows(), [('apple', {'user1'}, {'G1'}, {'loc1'})])

    def test_replace_row_ownership(self):
        self.upload([(None, 'N', 'apple', 'user1', 'G1', 'loc1')])
        apple_id, = [r.id.hex for r in self.get_rows(None)]

        self.upload([(apple_id, 'N', 'apple', 'user2', 'g2', 'loc2')])
        self.assertEqual(self.get_rows(), [('apple', {'user2'}, {'g2'}, {'loc2'})])

    def test_delete_ownership(self):
        self.upload([(None, 'N', 'apple', 'user1', 'G1', 'loc1')])
        apple_id, = [r.id.hex for r in self.get_rows(None)]

        self.upload([(apple_id, 'N', 'apple', None, None, None)])
        self.assertEqual(self.get_rows(), [('apple', set(), set(), set())])

    def test_ownerships_deleted_with_table(self):
        self.upload([(None, 'N', 'apple', 'user1', 'G1', 'loc1')])
        apple_id, = [r.id.hex for r in self.get_rows(None)]

        data = [
            ('types', [('Y', 'things', 'no', 'name')]),
            ('things', [(apple_id, 'N', 'apple', 'user1', 'G1', 'loc1')]),
        ]
        workbook = TestFixtureUpload.get_workbook_from_data(self.headers, data)
        type(self).do_upload(self.domain, workbook)

        ownerships = list(LookupTableRowOwner.objects.filter(domain=self.domain, row_id=apple_id))
        self.assertEqual(ownerships, [])

    def test_unknown_owners(self):
        result = self.upload([(None, 'N', 'apple', 'who', 'what', 'where')], check_result=False)
        self.assertEqual(self.get_rows(), [('apple', set(), set(), set())])
        self.assertEqual(set(result.errors), {
            "Unknown group: 'what'. But the row is successfully added",
            "Unknown user: 'who'. But the row is successfully added",
            "Unknown location: 'where'. But the row is successfully added",
        })

    def test_multiple_locations(self):
        loc_dup = SQLLocation(domain=self.domain, name="d u p", location_type=self.region)
        loc_dup.save()
        loc_dup2 = SQLLocation(domain=self.domain, name="d u p", location_type=self.region)
        loc_dup2.save()
        assert SQLLocation.objects.filter(name="d u p").count() == 2

        result = self.upload([(None, 'N', 'apple', None, None, 'd u p')], check_result=False)
        self.assertEqual(self.get_rows(), [('apple', set(), set(), set())])
        self.assertEqual(result.errors, [
            "Multiple locations found with the name: 'd u p'.  "
            "Try using site code. But the row is successfully added"
        ])

    def test_case_insensitive_ownership_matching(self):
        loc = SQLLocation(domain=self.domain, name="L O C", location_type=self.region)
        loc.save()
        assert loc.name != loc.site_code, loc.site_code
        self.upload([(None, 'N', 'apple', 'User1', None, 'L o c')])
        self.assertEqual(self.get_rows(), [('apple', {'user1'}, set(), {'L O C'})])

    def test_case_sensitive_group_match(self):
        assert self.group1.name == 'G1', self.group1.name
        result = self.upload([(None, 'N', 'apple', None, 'g1', None)], check_result=False)
        self.assertEqual(self.get_rows(), [('apple', set(), set(), set())])
        self.assertEqual(result.errors, ["Unknown group: 'g1'. But the row is successfully added"])

    def test_invalid_username(self):
        result = self.upload([(None, 'N', 'apple', 'n@pe', None, None)], check_result=False)
        self.assertEqual(self.get_rows(), [('apple', set(), set(), set())])
        self.assertEqual(result.errors, ["Invalid username: 'n@pe'. But the row is successfully added"])

    def test_non_string_owner_names(self):
        result = self.upload([(None, 'N', 'apple', 3, 3, 3)], check_result=False)
        self.assertEqual(self.get_rows(), [('apple', {'3'}, {'3'}, {'3'})])
        self.assertFalse(result.errors)

    def upload(self, rows, *, check_result=True, **kw):
        data = self.make_rows(rows)
        workbook = TestFixtureUpload.get_workbook_from_data(self.headers, data)
        result = type(self).do_upload(self.domain, workbook, **kw)
        if check_result:
            self.assertFalse(result.errors)
            self.assertFalse(result.messages)
        return result

    def get_rows(self, transform=row_name):
        def sort_key(item):
            return item.sort_key, transform(item)

        def get_owner_ids(row, owner_type):
            return [o.owner_id for o in LookupTableRowOwner.objects.filter(
                row_id=row.id,
                owner_type=owner_type,
            )]

        def get_users(row):
            doc_ids = get_owner_ids(row, OwnerType.User)
            users = iter_docs(CommCareUser.get_db(), doc_ids)
            return [CommCareUser.wrap(d) for d in users]

        def get_groups(row):
            doc_ids = get_owner_ids(row, OwnerType.Group)
            return [Group.wrap(d) for d in iter_docs(Group.get_db(), doc_ids)]

        def get_locations(row):
            loc_ids = get_owner_ids(row, OwnerType.Location)
            return SQLLocation.active_objects.filter(location_id__in=loc_ids)

        items = LookupTableRow.objects.iter_rows(self.domain, tag='things')
        if transform is None:
            return list(items)
        return [
            (
                transform(item),
                {u.raw_username for u in get_users(item)},
                {g.name for g in get_groups(item)},
                {x.name for x in get_locations(item)},
            )
            for item in sorted(items, key=sort_key)
        ]

    headers = (
        ('types', ('Delete(Y/N)', 'table_id', 'is_global?', 'field 1')),
        ('things', ('UID', 'Delete(Y/N)', 'field: name', 'user 1', 'group 1', 'location 1')),
    )

    @staticmethod
    def make_rows(item_rows):
        return (
            ('types', [('N', 'things', 'no', 'name')]),
            ('things', item_rows),
        )


class TestLocationLookups(TestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        publish = patch("corehq.apps.locations.document_store.publish_location_saved")
        publish.start()
        cls.addClassCleanup(publish.stop)

        cls.domain = 'fixture-upload-test'
        cls.project = create_domain(cls.domain)
        cls.addClassCleanup(cls.project.delete)

        cls.region = LocationType(domain=cls.domain, name="region", code="region")
        cls.region.save()
        cls.loc = SQLLocation(domain=cls.domain, name="L O C", location_type=cls.region)
        cls.loc.save()
        assert cls.loc.name != cls.loc.site_code, cls.loc.site_code

        # duplicates with matching site code
        cls.loc_dup1 = SQLLocation(domain=cls.domain, name="DUP", location_type=cls.region)
        cls.loc_dup1.save()
        cls.loc_dup2 = SQLLocation(domain=cls.domain, name="dup", location_type=cls.region)
        cls.loc_dup2.save()
        dup_names = SQLLocation.objects.filter(name__iexact="dup").count()
        assert dup_names == 2, dup_names
        assert cls.loc_dup1.site_code == "dup", cls.loc_dup1.site_code
        assert cls.loc_dup2.site_code != "dup", cls.loc_dup2.site_code

        # duplicates with no matching site code
        cls.loc_dup3 = SQLLocation(domain=cls.domain, name="D U P", location_type=cls.region)
        cls.loc_dup3.save()
        cls.loc_dup4 = SQLLocation(domain=cls.domain, name="d u p", location_type=cls.region)
        cls.loc_dup4.save()
        dup_names = SQLLocation.objects.filter(name__iexact="d u p").count()
        dup_sites = SQLLocation.objects.filter(site_code="d u p").count()
        assert dup_names == 2, dup_names
        assert dup_sites == 0, dup_sites

    def test_match_site_code(self):
        self.assertEqual(self.lookup(self.loc.site_code), self.loc.location_id)

    def test_case_insensitive_name_match(self):
        self.assertEqual(self.lookup("l o c"), self.loc.location_id)

    def test_site_code_overrides_duplicate_name(self):
        self.assertEqual(self.lookup("dup"), self.loc_dup1.location_id)

    def test_site_code_match(self):
        self.assertEqual(self.lookup(self.loc_dup4.site_code), self.loc_dup4.location_id)

    def test_duplicate_name(self):
        self.assertEqual(self.lookup('d u p'), mod.MULTIPLE)
        # User-friendly error is added at another place in the uploader

    def lookup(self, name_or_site_code):
        assert name_or_site_code.islower(), """
            Location names are converted to lower case by
            `_FixtureWorkbook.get_owners()` before being passed to
            `_load_location_ids_by_name()`, so it does not make sense to
            pass names with upper case letters here."""
        result = mod._load_location_ids_by_name([name_or_site_code], self.domain)
        return result.get(name_or_site_code.lower())


class FakeTask:
    def __init__(self):
        self.states = []

    def update_state(self, **kw):
        self.states.append(kw)


class TestTableKey(SimpleTestCase):

    def test_tag(self):
        t1 = LookupTable(tag="this")
        t2 = LookupTable(tag="this")
        self.assertEqual(mod.table_key(t1), mod.table_key(t2))

        t3 = LookupTable(tag="that")
        self.assertNotEqual(mod.table_key(t1), mod.table_key(t3))

    def test_is_global(self):
        t1 = LookupTable(is_global=True)
        t2 = LookupTable(is_global=True)
        self.assertEqual(mod.table_key(t1), mod.table_key(t2))

        t3 = LookupTable(is_global=False)
        self.assertNotEqual(mod.table_key(t1), mod.table_key(t3))

    def test_fields(self):
        t1 = LookupTable(fields=[TypeField(name="name", properties=["color"])])
        t2 = LookupTable(fields=[TypeField(name="name", properties=["color"])])
        self.assertEqual(mod.table_key(t1), mod.table_key(t2))

        t3 = LookupTable(fields=[TypeField(name="name", properties=["hue"])])
        self.assertNotEqual(mod.table_key(t1), mod.table_key(t3))

    def test_item_attributes(self):
        t1 = LookupTable(item_attributes=["name"])
        t2 = LookupTable(item_attributes=["name"])
        self.assertEqual(mod.table_key(t1), mod.table_key(t2))

        t3 = LookupTable(item_attributes=["origin"])
        self.assertNotEqual(mod.table_key(t1), mod.table_key(t3))

    def test_description(self):
        t1 = LookupTable(description="Hello old friend")
        t2 = LookupTable(description="Hello old friend")
        self.assertEqual(mod.table_key(t1), mod.table_key(t2))

        t3 = LookupTable(description="It's good to see you")
        self.assertNotEqual(mod.table_key(t1), mod.table_key(t3))


class TestRowKey(SimpleTestCase):

    def test_fields(self):
        t1 = LookupTableRow(fields={"state": [Field(
            value="name",
            properties={"color": "blue"},
        )]})
        t2 = LookupTableRow(fields={"state": [Field(
            value="name",
            properties={"color": "blue"},
        )]})
        self.assertEqual(mod.row_key(t1), mod.row_key(t2))

        t3 = LookupTableRow(fields={"state": [Field(
            value="name",
            properties={"color": "red"},
        )]})
        self.assertNotEqual(mod.row_key(t1), mod.row_key(t3))

    def test_fields_order(self):
        t1 = LookupTableRow(fields={
            "state": [Field(value="name")],
            "temp": [Field(value="value")],
        })
        t2 = LookupTableRow(fields={
            "temp": [Field(value="value")],
            "state": [Field(value="name")],
        })
        self.assertEqual(mod.row_key(t1), mod.row_key(t2))

    def test_item_attributes(self):
        t1 = LookupTableRow(item_attributes={"color": "blue"})
        t2 = LookupTableRow(item_attributes={"color": "blue"})
        self.assertEqual(mod.row_key(t1), mod.row_key(t2))

        t3 = LookupTableRow(item_attributes={"color": "red"})
        self.assertNotEqual(mod.row_key(t1), mod.row_key(t3))

    def test_item_attributes_order(self):
        t1 = LookupTableRow(item_attributes={"color": "blue", "shape": "square"})
        t2 = LookupTableRow(item_attributes={"shape": "square", "color": "blue"})
        self.assertEqual(mod.row_key(t1), mod.row_key(t2))

    def test_sort_key(self):
        t1 = LookupTableRow(sort_key=1)
        t2 = LookupTableRow(sort_key=1)
        self.assertEqual(mod.row_key(t1), mod.row_key(t2))

        t3 = LookupTableRow(sort_key=2)
        self.assertNotEqual(mod.row_key(t1), mod.row_key(t3))


class TestOwnerKey(SimpleTestCase):

    def test_owner_type(self):
        t1 = LookupTableRowOwner(owner_type=OwnerType.User)
        t2 = LookupTableRowOwner(owner_type=OwnerType.User)
        self.assertEqual(mod.owner_key(t1), mod.owner_key(t2))

        t3 = LookupTableRowOwner(owner_type=OwnerType.Group)
        self.assertNotEqual(mod.owner_key(t1), mod.owner_key(t3))

    def test_owner_id(self):
        t1 = LookupTableRowOwner(owner_id="abc")
        t2 = LookupTableRowOwner(owner_id="abc")
        self.assertEqual(mod.owner_key(t1), mod.owner_key(t2))

        t3 = LookupTableRowOwner(owner_id="def")
        self.assertNotEqual(mod.owner_key(t1), mod.owner_key(t3))
