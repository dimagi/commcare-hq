import csv
from datetime import datetime
from inspect import cleandoc

from django.core.management.base import BaseCommand, CommandError

from couchdbkit import ResourceNotFound
from openpyxl import Workbook

from corehq.motech.repeaters.dbaccessors import (
    iter_sql_repeat_records_by_domain,
)
from corehq.motech.repeaters.models import SQLRepeatRecord
from corehq.util.log import with_progress_bar


class Command(BaseCommand):
    help = cleandoc("""
    Pass Repeater ID along with domain and State(Optional) or
    Pass a csv file path with a list of repeat records IDs to get a report(xlsx)
    with final state and message for all attempts(if available)
    """)

    def __init__(self, *args, **kwargs):
        super(Command, self).__init__(*args, **kwargs)
        # keep a counter of max attempts a repeat record has in the report
        # to add those many headers later
        self.max_attempts_in_sheet = 0
        self.record_ids = []
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['Record ID', 'Payload ID', 'State', 'Failure Reason'])

    def add_arguments(self, parser):
        parser.add_argument(
            '--domain',
            dest='domain',
            help='Domain for which the repeater is passed'
        )
        parser.add_argument(
            '--repeater_id',
            dest='repeater_id',
            help='ID for repeater for which the report is to be generated',
        )
        parser.add_argument(
            '--state',
            dest='state',
            help='State of Record: CANCELLED, SUCCESS, PENDING, FAIL'
        )
        parser.add_argument(
            '--records_file_path',
            dest='records_file_path',
            help='Path to a csv file with ids under header record_id'
        )

    def _add_row(self, repeat_record):
        row = [
            repeat_record.pk,
            repeat_record.payload_id,
            repeat_record.state,
            repeat_record.failure_reason,
        ]
        if repeat_record.num_attempts > self.max_attempts_in_sheet:
            self.max_attempts_in_sheet = repeat_record.num_attempts
        for attempt in repeat_record.attempts:
            row.append(attempt.message)
        self.ws.append(row)

    def _add_header_for_attempts(self):
        # find first empty header cell:
        # iterate first row till highest column length which is highest column length over the sheet
        # Even if there are no values in header, empty cells are inserted to match highest
        # column length. So iterate over cells and find index for the first blank value
        max_columns = self.ws.max_column
        first_empty_index = max_columns
        for i in range(1, max_columns + 1):
            if not self.ws.cell(row=1, column=i).value:
                first_empty_index = i
                break
        for i in range(0, self.max_attempts_in_sheet):
            column_num = first_empty_index + i
            self.ws.cell(row=1, column=column_num).value = "Attempt {index}".format(index=i + 1)

    def _save_file(self, repeater_id, state):
        if self.max_attempts_in_sheet > 0:
            self._add_header_for_attempts()
        file_name = "repeater_report_{repeater_id}_{state}_{timestamp}.xlsx".format(
            repeater_id=repeater_id if repeater_id else '',
            state=state if state else '',
            timestamp=datetime.now().strftime("%Y-%m-%d-%H-%M-%S"),
        )
        self.wb.save(file_name)
        return file_name

    def _load_record_ids_from_file(self, file_path):
        with open(file_path, 'rU', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                self.record_ids.append(row['record_id'])

    def handle(self, *args, **options):
        domain = options.get('domain')
        repeater_id = options.get('repeater_id')
        state = options.get('state')
        records_file_path = options.get('records_file_path')

        if records_file_path:
            self._load_record_ids_from_file(records_file_path)
            records = self.record_ids
            record_count = len(records)
        elif domain and repeater_id:
            records, record_count = iter_sql_repeat_records_by_domain(
                domain,
                repeater_id,
                [state] if state else None,
            )
        else:
            raise CommandError("Insufficient Arguments")

        for record in with_progress_bar(records, length=record_count):
            if isinstance(record, str):
                record_id = record
                try:
                    record = SQLRepeatRecord.objects.get(pk=record_id)
                except ResourceNotFound:
                    self.ws.append([record_id, '', 'Not Found'])
                    continue
            self._add_row(record)

        file_name = self._save_file(repeater_id, state)
        print("Report saved in file:{filename}".format(filename=file_name))
