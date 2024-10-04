import io
from codecs import BOM_UTF8
import os
import re
import tempfile
import zipfile
import csv
import json
from collections import OrderedDict
import openpyxl
import math

from django.template.loader import render_to_string, get_template
from django.utils.functional import Promise
import xlwt

from couchexport.models import Format
from openpyxl.styles import numbers
from openpyxl.cell import WriteOnlyCell

from couchexport.util import get_excel_format_value, get_legacy_excel_safe_value

MAX_XLS_COLUMNS = 256


class XlsLengthException(Exception):
    pass


class UniqueHeaderGenerator(object):

    def __init__(self, max_column_size=None):
        self.used = set()
        self.max_column_size = max_column_size or 2000

    def next_unique(self, header):
        header = self._next_unique(header)
        self.used.add(header.lower())
        return header

    def _next_unique(self, string):
        counter = 1
        split = (self.max_column_size - 3) / 2
        if len(string) > self.max_column_size:
            # truncate the middle
            string = "{}...{}".format(string[:math.ceil(split)], string[-math.floor(split):])
        orig_string = string
        while string.lower() in self.used:
            string = "%s%s" % (orig_string, counter)
            if len(string) > self.max_column_size:
                string = "{}...{}".format(string[:math.ceil(split)], string[-math.floor(split):])
            counter += 1

        return string

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass


class ExportFileWriter(object):

    def __init__(self):
        self.name = None
        self._isopen = False
        self._file = None
        self._path = None

    def get_path(self):
        assert self._isopen
        return self._path

    def get_file(self):
        assert self._isopen
        return self._file

    def open(self, name):
        assert not self._isopen
        self._isopen = True
        self.name = name
        fd, path = tempfile.mkstemp()
        self._file = os.fdopen(fd, 'wb+')
        self._path = path
        self._open()
        self._begin_file()

    def _open(self):
        pass

    def _begin_file(self):
        pass

    def write_row(self, row):
        raise NotImplementedError

    def _end_file(self):
        pass

    def finish(self):
        self._end_file()
        self._file.seek(0)

    def close(self):
        assert self._isopen
        self._file.close()
        os.remove(self._path)
        self._isopen = False


class CsvFileWriter(ExportFileWriter):

    def _open(self):
        # Excel needs UTF8-encoded CSVs to start with the UTF-8 byte-order mark (FB 163268)
        self._file.write(BOM_UTF8)

    def write_row(self, row):
        buffer = io.StringIO()
        csvwriter = csv.writer(buffer, csv.excel)
        csvwriter.writerow([
            col.decode('utf-8') if isinstance(col, bytes) else col
            for col in row
        ])
        self._file.write(buffer.getvalue().encode('utf-8'))


class PartialHtmlFileWriter(ExportFileWriter):

    def _write_from_template(self, context):
        self._file.write(self.template.render(context).encode('utf-8'))

    def _open(self):
        self.template = get_template("couchexport/html_export.html")
        self._on_first_row = True

    def write_row(self, row):
        section = "row" if not self._on_first_row else "first_row"
        self._on_first_row = False
        self._write_from_template({"row": row, "section": section})

    def _end_file(self):
        if self._on_first_row:
            # There were no rows
            self._write_from_template({"section": "no_rows"})


class HtmlFileWriter(PartialHtmlFileWriter):

    def _begin_file(self):
        self._write_from_template({"section": "doc_begin"})
        self._write_from_template({"section": "table_begin", "name": self.name})

    def _end_file(self):
        super(HtmlFileWriter, self)._end_file()
        self._write_from_template({"section": "table_end"})
        self._write_from_template({"section": "doc_end"})


class ExportWriter(object):
    max_table_name_size = 500
    target_app = 'Excel'  # Where does this writer export to? Export button to say "Export to Excel"

    def open(self, header_table, file, max_column_size=2000, table_titles=None, archive_basepath=''):
        """
        Create any initial files, headings, etc necessary.
        :param header_table: tuple of one of the following formats
            tuple(sheet_name, [['col1header', 'col2header', ....]])
            tuple(sheet_name, [FormattedRow])
        """
        table_titles = table_titles or {}

        self._isopen = True
        self.max_column_size = max_column_size
        self._current_primary_id = 0
        self.file = file
        self.archive_basepath = archive_basepath

        self._init()
        self.table_name_generator = UniqueHeaderGenerator(
            self.max_table_name_size
        )
        for table_index, table in header_table:
            self.add_table(
                table_index,
                list(table)[0],
                table_title=table_titles.get(table_index)
            )

    def add_table(self, table_index, headers, table_title=None):
        def _clean_name(name):
            if isinstance(name, bytes):
                name = name.decode('utf8')
            elif isinstance(name, Promise):
                # noinspection PyCompatibility
                name = str(name)
            return re.sub(r"[\n]", '', re.sub(r"[\[\\?*/:\]]", "-", name))

        table_title_truncated = self.table_name_generator.next_unique(
            _clean_name(table_title or table_index)
        )

        # make sure we trim the headers
        with UniqueHeaderGenerator(self.max_column_size) as g:
            try:
                headers.data = [g.next_unique(header) for header in headers.data]
            except AttributeError:
                headers = [g.next_unique(header) for header in headers]

        self._init_table(table_index, table_title_truncated)
        self.write_row(table_index, headers)

    def write(self, document_table, skip_first=False):
        """
        Given a document that's been parsed into the appropriate
        rows, write those rows to the resulting files.
        """
        assert self._isopen
        for table_index, table in document_table:
            for i, row in enumerate(table):
                if skip_first and i == 0:
                    continue
                # update the primary component of the ID to match
                # how many docs we've seen
                try:
                    row_has_id = row.has_id()
                except AttributeError:
                    row_has_id = False
                if row_has_id:
                    row.id = (self._current_primary_id,) + tuple(row.id[1:])

                self.write_row(table_index, row)

        self._current_primary_id += 1

    def write_row(self, table_index, row):
        """
        Currently just calls the subclass's implementation
        but if we were to add a universal validation step,
        such a thing would happen here.
        """
        return self._write_row(table_index, row)

    def close(self):
        """
        Close any open file references, do any cleanup.
        """
        assert self._isopen
        self._close()
        self._isopen = False

    def _init(self):
        raise NotImplementedError

    def _init_table(self, table_index, table_title):
        raise NotImplementedError

    def _write_row(self, sheet_index, row):
        raise NotImplementedError

    def _close(self):
        raise NotImplementedError

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._close()


class OnDiskExportWriter(ExportWriter):
    """
    Keeps tables in temporary csv files. Subclassed by other export writers.
    """
    writer_class = CsvFileWriter
    _write_row_force_to_bytes = True

    def _init(self):
        self.tables = OrderedDict()
        self.table_names = OrderedDict()

    def _init_table(self, table_index, table_title):
        writer = self.writer_class()
        self.tables[table_index] = writer
        writer.open(table_title)
        self.table_names[table_index] = table_title

    def _write_row(self, sheet_index, row):

        def _transform(val):
            if val is None:
                val = ''
            if self._write_row_force_to_bytes and isinstance(val, str):
                val = val.encode("utf8")
            return val

        row = list(map(_transform, row))
        self.tables[sheet_index].write_row(row)

    def _close(self):
        """
        Close any open file references, do any cleanup.
        """
        for writer in self.tables.values():
            writer.finish()

        self._write_final_result()

        for writer in self.tables.values():
            writer.close()

    def _write_final_result(self):
        """
        Subclasses should call this method then write to a zip file, html files, or whatever.
        """
        raise NotImplementedError


class ZippedExportWriter(OnDiskExportWriter):
    """
    Writer that creates a zip file containing a csv for each table.
    """
    table_file_extension = ".csv"

    def _write_final_result(self):
        archive = zipfile.ZipFile(self.file, 'w', zipfile.ZIP_DEFLATED)
        for index, name in self.table_names.items():
            if isinstance(name, bytes):
                name = name.decode('utf-8')
            path = self.tables[index].get_path()
            archive_filename = self._get_archive_filename(name)
            archive.write(path, archive_filename)
        archive.close()
        self.file.seek(0)

    def _get_archive_filename(self, name):
        return os.path.join(self.archive_basepath, '{}{}'.format(name, self.table_file_extension))


class CsvExportWriter(ZippedExportWriter):
    """
    CSV writer that creates a zip file containing a csv for each table.
    """
    format = Format.CSV


class UnzippedCsvExportWriter(OnDiskExportWriter):
    """
    Serve the first table as a csv
    """
    format = Format.UNZIPPED_CSV

    def _write_final_result(self):
        tablefile = list(self.tables.values())[0].get_file()
        for line in tablefile:
            self.file.write(line)
        self.file.seek(0)


class Excel2007ExportWriter(ExportWriter):
    format = Format.XLS_2007
    max_table_name_size = 31

    def __init__(self, format_as_text=False, use_formatted_cells=False):
        super(Excel2007ExportWriter, self).__init__()
        self.format_as_text = format_as_text
        self.use_formatted_cells = use_formatted_cells

    def _init(self):
        # https://openpyxl.readthedocs.io/en/latest/optimized.html
        self.book = openpyxl.Workbook(write_only=True)

        self.tables = {}
        self.table_indices = {}

    def _init_table(self, table_index, table_title):
        sheet = self.book.create_sheet()
        sheet.title = table_title
        self.tables[table_index] = sheet
        self.table_indices[table_index] = 0

    def _write_row(self, sheet_index, row):
        from couchexport.export import FormattedRow
        sheet = self.tables[sheet_index]

        cells = []
        for col_ind, val in enumerate(row):
            skip_formatting_on_row = (isinstance(row, FormattedRow)
                                      and col_ind in row.skip_excel_formatting)

            if (self.use_formatted_cells
                    and not skip_formatting_on_row
                    and not self.format_as_text):
                excel_format, val_fmt = get_excel_format_value(val)
                cell = WriteOnlyCell(sheet, val_fmt)
                cell.number_format = excel_format
            else:
                cell = WriteOnlyCell(sheet, get_legacy_excel_safe_value(val))
                if self.format_as_text:
                    cell.number_format = numbers.FORMAT_TEXT

            cells.append(cell)

        if isinstance(row, FormattedRow):
            for hyperlink_column_index in row.hyperlink_column_indices:
                cells[hyperlink_column_index].hyperlink = cells[hyperlink_column_index].value
                cells[hyperlink_column_index].style = 'Hyperlink'

        sheet.append(cells)

    def _close(self):
        """
        Close any open file references, do any cleanup.
        """
        self.book.save(self.file)


class Excel2003ExportWriter(ExportWriter):
    format = Format.XLS
    max_table_name_size = 31

    def _init(self):
        self.book = xlwt.Workbook()
        self.tables = {}
        self.table_indices = {}

    def _init_table(self, table_index, table_title):
        sheet = self.book.add_sheet(table_title)
        self.tables[table_index] = sheet
        self.table_indices[table_index] = 0

    def _write_row(self, sheet_index, row):
        row_index = self.table_indices[sheet_index]
        sheet = self.tables[sheet_index]

        # have to deal with primary ids
        for i, val in enumerate(row):
            if i >= MAX_XLS_COLUMNS:
                raise XlsLengthException()
            sheet.write(row_index, i, str(val))
        self.table_indices[sheet_index] = row_index + 1

    def _close(self):
        self.book.save(self.file)


class InMemoryExportWriter(ExportWriter):
    """
    Keeps tables in memory. Subclassed by other export writers.
    """

    def _init(self):
        self.tables = {}
        self.table_names = {}

    def _init_table(self, table_index, table_title):
        self.table_names[table_index] = table_title
        self.tables[table_index] = []

    def _write_row(self, sheet_index, row):
        table = self.tables[sheet_index]
        # have to deal with primary ids
        table.append(list(row))

    def _close(self):
        pass


class JsonExportWriter(InMemoryExportWriter):
    """
    Write tables to JSON
    """
    format = Format.JSON

    class ConstantEncoder(json.JSONEncoder):

        def default(self, obj):
            from dimagi.utils.web import json_handler
            from couchexport.export import Constant
            if isinstance(obj, Constant):
                return obj.message
            else:
                return json_handler(obj)

    def _close(self):
        new_tables = {}
        for tablename, data in self.tables.items():
            new_tables[self.table_names[tablename]] = {"headers": data[0], "rows": data[1:]}

        json_dump = json.dumps(new_tables, cls=self.ConstantEncoder).encode('utf-8')
        self.file.write(json_dump)


class PythonDictWriter(InMemoryExportWriter):
    format = Format.PYTHON_DICT

    class ConstantEncoder(json.JSONEncoder):

        def default(self, obj):
            from dimagi.utils.web import json_handler
            from couchexport.export import Constant
            if isinstance(obj, Constant):
                return obj.message
            else:
                return json_handler(obj)

    def get_preview(self):
        new_tables = []
        for tablename, data in self.tables.items():
            new_tables.append({
                'table_name': self.table_names[tablename],
                'headers': data[0],
                'rows': data[1:],
            })
        dumps = json.dumps(new_tables, cls=self.ConstantEncoder)
        loads = json.loads(dumps)
        return loads


class HtmlExportWriter(OnDiskExportWriter):
    """
    Write tables to a single HTML file.
    """
    format = Format.HTML
    writer_class = PartialHtmlFileWriter

    _write_row_force_to_bytes = False

    def _write_final_result(self):

        def write(context):
            self.file.write(
                render_to_string(
                    "couchexport/html_export.html", context
                ).encode("utf-8")
            )

        write({"section": "doc_begin"})
        for index, name in self.table_names.items():
            table_writer = self.tables[index]
            write({"section": "table_begin", "name": name})
            for line in table_writer.get_file():
                self.file.write(line)
            write({"section": "table_end"})
        write({"section": "doc_end"})

        self.file.seek(0)


class ZippedHtmlExportWriter(ZippedExportWriter):
    """
    Write each table to an HTML file in a zipfile
    """
    writer_class = HtmlFileWriter
    table_file_extension = ".html"
    format = Format.ZIPPED_HTML


class GeoJSONWriter(JsonExportWriter):
    format = Format.GEOJSON
    table_file_extension = ".geojson"

    @staticmethod
    def parse_feature(coordinates, properties):
        return {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": coordinates},
            "properties": properties,
        }

    @staticmethod
    def _find_geo_property_by_path(table):
        # For form exports we store the path to the geo property, so this is an
        # attempt to find the header by the path specification
        column = table.get_column_by_path_str(
            path=table.selected_geo_property,
            doc_type="GeopointItem",
        )
        if column:
            return column.get_headers()[0]

        return table.selected_geo_property

    def get_features(self, table, data):
        geo_property_name = table.selected_geo_property
        table_headers = data[0]
        if geo_property_name not in table_headers:
            # This might happen for some form export metadata columns
            geo_property_name = self._find_geo_property_by_path(table)
            if geo_property_name not in table_headers:
                return []

        geo_data_index = table_headers.index(geo_property_name)
        features = []
        for row in data[1:]:
            try:
                # row[geo_data_index] could look like "<lat> <lng>" or "<lat> <lng> 0 0"
                result = row[geo_data_index].split(" ")
                lat = result[0]
                lng = result[1]
            except IndexError:
                continue
            properties = {header: row[i] for i, header in enumerate(table_headers) if header != geo_property_name}
            features.append(self.parse_feature(
                coordinates=[lng, lat],
                properties=properties,
            ))
        return features

    def _close(self):
        feature_collections = []
        for table, data in self.tables.items():
            feature_collections.append(
                self.get_features(table, data)
            )
        new_tables = {
            "type": "FeatureCollection",
            "features": feature_collections[0],
        }

        json_dump = json.dumps(new_tables, cls=self.ConstantEncoder).encode('utf-8')
        self.file.write(json_dump)
