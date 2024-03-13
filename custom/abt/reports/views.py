import io
import json

from django.http import HttpResponse
from memoized import memoized

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from sqlagg import MaxColumn
from sqlagg.columns import SimpleColumn
from sqlagg.filters import IN, BETWEEN

from corehq.apps.reports.sqlreport import SqlData, DatabaseColumn
from corehq.apps.reports.util import get_INFilter_bindparams
from corehq.apps.userreports.reports.util import ReportExport
from corehq.apps.userreports.reports.view import CustomConfigurableReport
from corehq.apps.userreports.util import get_table_name
from custom.utils.utils import clean_IN_filter_value

# Copied from custom/abt/reports/data_sources/supervisory.json
MAX_LOCATION_COLUMNS = 350


def _invert_table(table):
    return [
        [row[column_index] for row in table]
        for column_index in range(len(table[0]))
    ]


class FormattedSupervisoryReport(CustomConfigurableReport):

    @property
    def export_table(self):
        data = super(FormattedSupervisoryReport, self).export_table

        # remove zeroes
        table = list(data[0][1])
        for row in range(1, len(table) - 1):
            for column in range(2, len(table[row])):
                if table[row][column] == 0:
                    table[row][column] = ''

        # remove hyphen prefixes from location column headers
        for column_index in range(2, len(table[0])):
            table[0][column_index] = table[0][column_index][1:]

        # sort columns by location
        inverted_table = _invert_table(table)
        inverted_incident_and_total_columns = inverted_table[:2]
        inverted_location_columns = inverted_table[2:]
        sorted_inverted_location_columns = sorted(
            inverted_location_columns,
            key=lambda inverted_location_column: inverted_location_column[0].lower()
        )
        data[0][1] = _invert_table(
            inverted_incident_and_total_columns + sorted_inverted_location_columns
        )
        data[0][1] = table
        return data

    @property
    def excel_response(self):
        unformatted_excel_file = super(FormattedSupervisoryReport, self).excel_response

        workbook = openpyxl.load_workbook(unformatted_excel_file)
        worksheet = workbook.get_active_sheet()

        red = PatternFill(
            start_color='FFEE1111',
            end_color='FFEE1111',
            fill_type='solid',
        )

        max_row = worksheet.max_row
        max_column = get_column_letter(worksheet.max_column)

        def percentile_fill(start_column, start_row, end_column, end_row,
                            percentile, fill):
            format_range = {
                'start_column': start_column,
                'start_row': start_row,
                'end_column': end_column,
                'end_row': end_row,
            }
            worksheet.conditional_formatting.add(
                "%(start_column)s%(start_row)d:%(end_column)s%(end_row)d" % format_range,
                CellIsRule(
                    operator='greaterThan',
                    formula=[(
                        'PERCENTILE($%(start_column)s$%(start_row)d:'
                        '$%(end_column)s$%(end_row)d,{})'
                    ).format(percentile) % format_range],
                    fill=fill
                )
            )

        # total column
        percentile_fill('B', 2, 'B', max_row - 1, 0.95, red)

        # total row
        percentile_fill('C', max_row, max_column, max_row, 0.90, red)

        # body
        percentile_fill('C', 2, max_column, max_row - 1, 0.90, red)

        f = io.BytesIO()
        workbook.save(f)
        return f

    @property
    def email_response(self):
        return HttpResponse(json.dumps({
            'report': '',
        }))


class UniqueSOPSumDataSource(SqlData):

    def __init__(self, domain, filter_values):
        config = {
            'domain': domain,
        }
        for (key, value) in filter_values.items():
            if key == 'date_of_data_collection' and value:
                config['start_date'] = value.startdate
                config['end_date'] = value.enddate
            elif key and value:
                self.__setattr__(key, [x.value for x in value])
                config[key] = self.__getattribute__(key)
                clean_IN_filter_value(config, key)
        self.config = config

    @property
    def table_name(self):
        return get_table_name(self.config['domain'], "static-sms-indicators-001")

    @property
    def filters(self):
        filters = []
        if 'start_date' in self.config and 'end_date' in self.config:
            filters.append(BETWEEN("date_of_data_collection", "start_date", "end_date"))
        if 'country' in self.config:
            filters.append(IN('country', get_INFilter_bindparams('country', self.__getattribute__("country"))))
        if 'level_1' in self.config:
            filters.append(IN('level_1', get_INFilter_bindparams('level_1', self.__getattribute__("level_1"))))
        if 'level_2' in self.config:
            filters.append(IN('level_2', get_INFilter_bindparams('level_2', self.__getattribute__("level_2"))))
        if 'level_3' in self.config:
            filters.append(IN('level_3', get_INFilter_bindparams('level_3', self.__getattribute__("level_3"))))
        if 'level_4' in self.config:
            filters.append(IN('level_4', get_INFilter_bindparams('level_4', self.__getattribute__("level_4"))))
        return filters

    @property
    def group_by(self):
        return ['country', 'level_1', 'level_2', 'level_3', 'level_4']

    @property
    def columns(self):
        return [
            DatabaseColumn('country', SimpleColumn('country')),
            DatabaseColumn('level_1', SimpleColumn('level_1')),
            DatabaseColumn('level_2', SimpleColumn('level_2')),
            DatabaseColumn('level_3', SimpleColumn('level_3')),
            DatabaseColumn('level_4', SimpleColumn('level_4')),
            DatabaseColumn('total_sprayers', MaxColumn('total_sprayers'))
        ]


class CombinedDataSource(object):

    def __init__(self, original_data_source, filter_values):
        self.original_data_source = original_data_source
        self.filters_data = filter_values

    def get_data(self, start=None, limit=None):
        original_data = self.original_data_source.get_data(start, limit)
        unique_sops = UniqueSOPSumDataSource(self.original_data_source.domain, self.filters_data).get_data()
        group_columns = tuple(self.original_data_source.group_by[:-1])
        unique = {}

        for row in unique_sops:
            loc_names = tuple(row[x] for x in group_columns)
            if loc_names not in unique:
                unique[loc_names] = row['total_sprayers']['sort_key']
            else:
                unique[loc_names] += row['total_sprayers']['sort_key']

        for row in original_data:
            loc_names = tuple(row[x] for x in group_columns)
            row['total_sprayers'] = unique[loc_names]
        return original_data

    def set_filter_values(self, filter_values):
        self.original_data_source.set_filter_values(filter_values)

    def set_defer_fields(self, defer_fields):
        self.original_data_source.set_defer_fields(defer_fields)

    def set_order_by(self, columns):
        self.original_data_source.set_order_by(columns)

    @property
    def column_warnings(self):
        return self.original_data_source.column_warnings

    @property
    def group_by(self):
        return self.original_data_source.group_by

    @property
    def columns(self):
        return self.original_data_source.columns

    @property
    def inner_columns(self):
        return self.original_data_source.inner_columns

    @property
    def has_total_row(self):
        return self.original_data_source.has_total_row

    def get_total_records(self):
        return self.original_data_source.get_total_records()

    def get_total_row(self):
        return self.original_data_source.get_total_row()

    @property
    def top_level_columns(self):
        return self.original_data_source.top_level_columns

    @property
    def config(self):
        return self.original_data_source.config


class CustomReportExport(ReportExport):

    @property
    @memoized
    def data_source(self):
        original_data_source = super(CustomReportExport, self).data_source
        return CombinedDataSource(original_data_source, self.filter_values)


class FormattedSprayProgressReport(CustomConfigurableReport):

    @property
    @memoized
    def data_source(self):
        original_data_source = super(FormattedSprayProgressReport, self).data_source
        return CombinedDataSource(original_data_source, self.filter_values)

    @property
    @memoized
    def report_export(self):
        return CustomReportExport(self.domain, self.title, self.spec, self.lang, self.filter_values)
