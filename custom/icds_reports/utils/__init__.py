import copy
import csv
import json
import os
import string
import time
import zipfile

from collections import defaultdict
from datetime import datetime, timedelta, date
from dateutil.parser import parse
from functools import wraps
from memoized import memoized
from tempfile import mkstemp

import attr
import operator

import pytz
import qrcode
from base64 import b64encode
from io import BytesIO
from dateutil.relativedelta import relativedelta
from django.template.loader import render_to_string, get_template
from django.conf import settings
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS

from corehq.apps.app_manager.dbaccessors import get_latest_released_build_id
from corehq.apps.locations.models import SQLLocation
from corehq.apps.reports.datatables import DataTablesColumn
from corehq.apps.reports_core.filters import Choice
from corehq.apps.userreports.models import StaticReportConfiguration
from corehq.apps.userreports.reports.data_source import ConfigurableReportDataSource
from corehq.blobs.mixin import safe_id
from corehq.const import ONE_DAY
from corehq.util.files import TransientTempfile
from corehq.util.quickcache import quickcache
from corehq.util.timer import TimingContext
from custom.icds_reports import const
from custom.icds_reports.const import (
    ISSUE_TRACKER_APP_ID,
    LOCATION_TYPES,
    AggregationLevels,
    THR_REPORT_CONSOLIDATED,
    THR_REPORT_BENEFICIARY_TYPE,
    THR_REPORT_DAY_BENEFICIARY_TYPE,
    THR_21_DAYS_THRESHOLD_DATE
)

from custom.icds_reports.models.helper import IcdsFile
from custom.icds_reports.queries import get_test_state_locations_id, get_test_district_locations_id
from couchexport.export import export_from_tables
from dimagi.utils.dates import DateSpan
from dimagi.utils.parsing import ISO_DATE_FORMAT
from django.db.models import Case, When, Q, F, IntegerField, Min
import uuid
from sqlagg.filters import EQ, NOT
from pillowtop.models import KafkaCheckpoint
from custom.icds_reports.cache import icds_quickcache

OPERATORS = {
    "==": operator.eq,
    "!=": operator.ne,
    ">": operator.gt,
    ">=": operator.ge,
    "<": operator.lt,
    "<=": operator.le,
    "in": operator.contains,
}

RED = '#de2d26'
ORANGE = '#fc9272'
BLUE = '#006fdf'
PINK = '#fee0d2'
GREY = '#9D9D9D'


DATA_NOT_ENTERED = "Data Not Entered"
DEFAULT_VALUE = DATA_NOT_ENTERED
DATA_NOT_VALID = "Data Not Valid"

india_timezone = pytz.timezone('Asia/Kolkata')


class MPRData(object):
    resource_file = ('custom', 'icds_reports', 'resources', 'block_mpr.json')


class ASRData(object):
    resource_file = ('custom', 'icds_reports', 'resources', 'block_asr.json')


class ICDSData(object):

    def __init__(self, domain, filters, report_id, override_agg_column=None):
        report_config = ConfigurableReportDataSource.from_spec(
            self._get_static_report_configuration_without_owner_transform(report_id, domain, override_agg_column)
        )
        report_config.set_filter_values(filters)
        self.report_config = report_config

    def _get_static_report_configuration_without_owner_transform(self, report_id, domain, override_agg_column):
        report_id = report_id.format(domain=domain)
        static_report_configuration = StaticReportConfiguration.by_id(report_id, domain)

        if override_agg_column and override_agg_column != 'awc_id':
            static_report_configuration = self._override_agg(static_report_configuration, override_agg_column)

        # this is explicitly after override, otherwise 'report_columns' attrib gets memoized too early
        for report_column in static_report_configuration.report_columns:
            transform = report_column.transform
            if transform.get('type') == 'custom' and transform.get('custom_type') == 'owner_display':
                report_column.transform = {}
        return static_report_configuration

    def _override_agg(self, static_report_configuration, override_agg_column):
        level_order = ['owner_id', 'awc_id', 'supervisor_id', 'block_id', 'district_id', 'state_id']
        # override aggregation level
        static_report_configuration.aggregation_columns = [override_agg_column]
        # remove columns below agg level
        columns_to_remove = level_order[0:level_order.index(override_agg_column)]
        for column in static_report_configuration.columns:
            if column.get('column_id') in columns_to_remove:
                static_report_configuration.columns.remove(column)
        return static_report_configuration

    def data(self):
        return self.report_config.get_data()


class ICDSMixin(object):
    has_sections = False
    posttitle = None

    def __init__(self, config, allow_conditional_agg=False):
        self.config = config
        self.allow_conditional_agg = allow_conditional_agg

    @property
    def subtitle(self):
        return []

    @property
    def headers(self):
        return []

    @property
    def rows(self):
        return [[]]

    @property
    def sources(self):
        with open(os.path.join(*self.resource_file), encoding='utf-8') as f:
            return json.loads(f.read())[self.slug]

    @property
    @memoized
    def selected_location(self):
        if self.config['location_id']:
            return SQLLocation.objects.get(
                location_id=self.config['location_id']
            )

    @property
    @memoized
    def awc(self):
        if self.config['location_id']:
            return self.selected_location.get_descendants(include_self=True).filter(
                location_type__name='awc'
            )

    @property
    def awc_number(self):
        if self.awc:
            return len(
                [
                    loc for loc in self.awc
                    if 'test' not in loc.metadata and loc.metadata.get('test', '').lower() != 'yes'
                ]
            )

    @memoized
    def custom_data(self, selected_location, domain):
        timer = TimingContext()
        with timer:
            to_ret = self._custom_data(selected_location, domain)
        if selected_location:
            loc_type = selected_location.location_type.name
        else:
            loc_type = None
        tags = ["location_type:{}".format(loc_type), "report_slug:{}".format(self.slug)]
        if self.allow_conditional_agg:
            tags.append("allow_conditional_agg:yes")
        return to_ret

    def _custom_data(self, selected_location, domain):
        data = {}

        for config in self.sources['data_source']:
            filters = {}
            location_type_column = None
            if selected_location:
                location_type_column = selected_location.location_type.name.lower() + '_id'
                filters = {
                    location_type_column: [Choice(value=selected_location.location_id, display=selected_location.name)]
                }
            if 'date_filter_field' in config:
                filters.update({config['date_filter_field']: self.config['date_span']})
            if 'filter' in config:
                for fil in config['filter']:
                    if 'type' in fil:
                        now = datetime.now()
                        start_date = now if 'start' not in fil else now - timedelta(days=fil['start'])
                        end_date = now if 'end' not in fil else now - timedelta(days=fil['end'])
                        datespan = DateSpan(start_date, end_date)
                        filters.update({fil['column']: datespan})
                    else:
                        filters.update({
                            fil['column']: {
                                'operator': fil['operator'],
                                'operand': fil['value']
                            }
                        })

            timer = TimingContext()
            with timer:
                allow_conditional_agg = self.allow_conditional_agg and not config.get('disallow_conditional_agg', False)
                override_agg_column = location_type_column if allow_conditional_agg else None
                report_data = ICDSData(domain, filters, config['id'], override_agg_column).data()
            if selected_location:
                loc_type = selected_location.location_type.name
            else:
                loc_type = None
            tags = ["location_type:{}".format(loc_type), "report_slug:{}".format(self.slug), "config:{}".format(config['id'])]
            if allow_conditional_agg:
                tags.append("allow_conditional_agg:yes")

            for column in config['columns']:
                column_agg_func = column['agg_fun']
                column_name = column['column_name']
                column_data = 0
                if column_agg_func == 'sum':
                    column_data = sum([x.get(column_name, 0) or 0 for x in report_data])
                elif column_agg_func == 'count':
                    column_data = len(report_data)
                elif column_agg_func == 'count_if':
                    value = column['condition']['value']
                    op = column['condition']['operator']

                    def check_condition(v):
                        if isinstance(v, str):
                            fil_v = str(value)
                        elif isinstance(v, int):
                            fil_v = int(value)
                        else:
                            fil_v = value

                        if op == "in":
                            return OPERATORS[op](fil_v, v)
                        else:
                            return OPERATORS[op](v, fil_v)

                    column_data = len([val for val in report_data if check_condition(val[column_name])])
                elif column_agg_func == 'avg':
                    values = [x.get(column_name, 0) for x in report_data]
                    column_data = sum(values) / (len(values) or 1)
                elif column_agg_func == 'last_value':
                    group_by = column['group_by']
                    awc_mapping = {x.get(group_by): x.get(column_name, 0) for x in report_data}
                    column_data = sum(awc_mapping.values())
                column_display = column_name if 'column_in_report' not in column else column['column_in_report']
                data.update({
                    column_display: data.get(column_display, 0) + column_data
                })
        return data


class ICDSDataTableColumn(DataTablesColumn):

    @property
    def render_html(self):
        column_params = dict(
            title=self.html,
            sort=self.sortable,
            rotate=self.rotate,
            css="span%d" % self.css_span if self.css_span > 0 else '',
            rowspan=self.rowspan,
            help_text=self.help_text,
            expected=self.expected
        )
        return render_to_string("icds_reports/partials/column.html", dict(
            col=column_params
        ))


PREVIOUS_PERIOD_ZERO_DATA = "Data in the previous reporting period was 0"


def percent_increase(prop, data, prev_data):
    current = 0
    previous = 0
    if data:
        current = data[0][prop]
    if prev_data:
        previous = prev_data[0][prop]

    if previous:
        tenths_of_promils = (((current or 0) - (previous or 0)) * 10000) / float(previous or 1)
        return tenths_of_promils / 100 if (tenths_of_promils < -1 or 1 < tenths_of_promils) else 0
    else:
        return PREVIOUS_PERIOD_ZERO_DATA


def percent_diff(property, current_data, prev_data, all):
    current = 0
    curr_all = 1
    prev = 0
    prev_all = 1
    if current_data:
        current = (current_data[0][property] or 0)
        curr_all = (current_data[0][all] or 1)

    if prev_data:
        prev = (prev_data[0][property] or 0)
        prev_all = (prev_data[0][all] or 1)

    current_percent = current / float(curr_all) * 100
    prev_percent = prev / float(prev_all) * 100

    if prev_percent:
        tenths_of_promils = ((current_percent - prev_percent) * 10000) / (prev_percent or 1.0)
        return tenths_of_promils / 100 if (tenths_of_promils < -1 or 1 < tenths_of_promils) else 0
    else:
        return PREVIOUS_PERIOD_ZERO_DATA


def get_color_with_green_positive(val):
    if isinstance(val, (int, float)):
        if val > 0:
            return 'green'
        else:
            return 'red'
    else:
        assert val == PREVIOUS_PERIOD_ZERO_DATA, val
        return 'green'


def get_color_with_red_positive(val):
    if isinstance(val, (int, float)):
        if val > 0:
            return 'red'
        else:
            return 'green'
    else:
        assert val == PREVIOUS_PERIOD_ZERO_DATA, val
        return 'red'


def get_value(data, prop):
    return (data[0][prop] or 0) if data else 0


def apply_exclude(domain, queryset):
    return queryset.exclude(
        Q(state_id__in=get_test_state_locations_id(domain)) |
        Q(district_id__in=get_test_district_locations_id(domain))
    )


def get_age_filter(age_value):
    """
        When age_value = 6 it means first range is chosen 0-6 months.
        For that range we want to include 0 and 6 in results.
    """
    if age_value == '6':
        return {'age_tranche__in': ['0', '6']}
    else:
        return {'age_tranche': age_value}


def get_age_filter_in_months(age_value):
    """
        When age_value = 6 it means first range is chosen 0-6 months.
        For that range we want to include 0 and 6 in results.
    """
    if age_value == '6':
        return {'age_in_months__range': ['0', '6']}
    elif age_value == '12':
        return {'age_in_months__range': ['7', '12']}
    elif age_value == '24':
        return {'age_in_months__range': ['13', '24']}
    elif age_value == '36':
        return {'age_in_months__range': ['25', '36']}
    elif age_value == '48':
        return {'age_in_months__range': ['37', '48']}
    elif age_value == '60':
        return {'age_in_months__range': ['49', '60']}
    elif age_value == '72':
        return {'age_in_months__range': ['61', '72']}


def match_age(age):
    if 0 <= age <= 1:
        return '0-1 month'
    elif 1 < age <= 6:
        return '1-6 months'
    elif 6 < age <= 12:
        return '6-12 months'
    elif 12 < age <= 36:
        return '1-3 years'
    elif 36 < age <= 72:
        return '3-6 years'


def get_location_filter(location_id, domain, include_object=False):
    """
    Args:
        location_id (str)
        domain (str)
    Returns:
        dict
    """
    if not location_id:
        return {}

    config = {}
    try:
        sql_location = SQLLocation.objects.get(location_id=location_id, domain=domain)
    except SQLLocation.DoesNotExist:
        return {'aggregation_level': 1}
    config.update(
        {
            ('%s_id' % ancestor.location_type.code): ancestor.location_id
            for ancestor in sql_location.get_ancestors(include_self=True)
        }
    )
    config['aggregation_level'] = len(config) + 1

    if include_object:
        config['sql_location'] = sql_location

    return config


def get_location_level(aggregation_level):
    if not aggregation_level:
        return LOCATION_TYPES[0]
    elif aggregation_level >= len(LOCATION_TYPES):
        return LOCATION_TYPES[-1]
    return LOCATION_TYPES[aggregation_level - 1]


@quickcache([])
def get_latest_issue_tracker_build_id():
    return get_latest_released_build_id('icds-cas', ISSUE_TRACKER_APP_ID)


def get_status(value, second_part='', normal_value='', exportable=False, data_entered=False):
    status = {'value': DATA_NOT_VALID if data_entered else DATA_NOT_ENTERED, 'color': 'black'}
    if not value or value in ['unweighed', 'unmeasured', 'unknown']:
        status = {'value': DATA_NOT_VALID if data_entered else DATA_NOT_ENTERED, 'color': 'black'}
    elif value in ['severely_underweight', 'severe']:
        status = {'value': 'Severely ' + second_part, 'color': 'red'}
    elif value in ['moderately_underweight', 'moderate']:
        status = {'value': 'Moderately ' + second_part, 'color': 'black'}
    elif value in ['normal']:
        status = {'value': normal_value, 'color': 'black'}
    elif value in ['N/A']:
        return 'N/A'
    return status if not exportable else status['value']


def is_anemic(value):
    if value['anemic_severe']:
        return 'Y'
    elif value['anemic_moderate']:
        return 'Y'
    elif value['anemic_normal']:
        return 'N'
    else:
        return DATA_NOT_ENTERED


def get_anemic_status(value):
    if value['anemic_severe']:
        return 'Severe'
    elif value['anemic_moderate']:
        return 'Moderate'
    elif value['anemic_normal']:
        return 'Normal'
    else:
        return DATA_NOT_ENTERED


def get_symptoms(value):
    if value['bleeding']:
        return 'Bleeding'
    elif value['swelling']:
        return 'Face, hand or genital swelling'
    elif value['blurred_vision']:
        return 'Blurred vision / headache'
    elif value['convulsions']:
        return 'Convulsions / unconsciousness'
    elif value['rupture']:
        return 'Water ruptured without labor pains'
    else:
        return 'None'


def get_counseling(value):
    counseling = []
    if value['eating_extra']:
        counseling.append('Eating Extra')
    if value['resting']:
        counseling.append('Taking Rest')
    if value['immediate_breastfeeding']:
        counseling.append('Counsel on Immediate Breastfeeding')
    if counseling:
        return ', '.join(counseling)
    else:
        return 'None'


def get_tt_dates(value):
    tt_dates = []
    # ignore 1970-01-01 as that is default date for ledger dates
    default = date(1970, 1, 1)
    if value['tt_1'] and value['tt_1'] != default:
        tt_dates.append(str(value['tt_1']))
    if value['tt_2'] and value['tt_2'] != default:
        tt_dates.append(str(value['tt_2']))
    if tt_dates:
        return '; '.join(tt_dates)
    else:
        return 'None'


def get_delivery_nature(value):
    delivery_natures = {
        1: 'Vaginal',
        2: 'Caesarean',
        3: 'Instrumental',
        0: DATA_NOT_ENTERED,
    }
    return delivery_natures.get(value['delivery_nature'], DATA_NOT_ENTERED)


def current_age(dob, selected_date):
    age = relativedelta(selected_date, dob)
    age_format = ""
    if age.years:
        age_format += "%s year%s " % (age.years, '' if age.years == 1 else 's')
    if age.months:
        age_format += "%s month%s " % (age.months, '' if age.months == 1 else 's')
    if not age.years and not age.months:
        if age.days > 0:
            age_format = "%s day%s" % (age.days, '' if age.days == 1 else 's')
        else:
            age_format = "0 days"
    return age_format


def exclude_records_by_age_for_column(exclude_config, column):
    return Case(
        When(~Q(**exclude_config), then=F(column)),
        default=0,
        output_field=IntegerField()
    )


def include_records_by_age_for_column(include_config, column):
    return Case(
        When(Q(**include_config), then=F(column)),
        default=0,
        output_field=IntegerField()
    )


def generate_data_for_map(data, loc_level, num_prop, denom_prop, fill_key_lower, fill_key_bigger,
                          all_property=None, location_launched_status=None):
    data_for_map = defaultdict(lambda: {
        num_prop: 0,
        denom_prop: 0,
        'original_name': []
    })

    if all_property:
        data_for_map = defaultdict(lambda: {
            num_prop: 0,
            denom_prop: 0,
            'original_name': [],
            all_property: 0
        })

    valid_total = 0
    in_month_total = 0
    total = 0
    values_to_calculate_average = {'numerator': 0, 'denominator': 0}


    for row in data:
        if location_launched_status is not None:
            launched_status = location_launched_status.get(row['%s_name' % loc_level])
            if launched_status is None or launched_status <= 0:
                continue
        valid = row[denom_prop] or 0
        name = row['%s_name' % loc_level]
        on_map_name = row['%s_map_location_name' % loc_level] or name
        in_month = row[num_prop] or 0

        values_to_calculate_average['numerator'] += in_month if in_month else 0
        values_to_calculate_average['denominator'] += row[denom_prop] if row[denom_prop] else 0

        valid_total += valid
        in_month_total += in_month
        if all_property:
            all_data = row[all_property] or 0
            data_for_map[on_map_name][all_property] += all_data
            total += all_data
        data_for_map[on_map_name][num_prop] += in_month
        data_for_map[on_map_name][denom_prop] += valid
        data_for_map[on_map_name]['original_name'].append(name)

    for data_for_location in data_for_map.values():
        value = data_for_location[num_prop] * 100 / (data_for_location[denom_prop] or 1)
        fill_format = '%s%%-%s%%'
        if value < fill_key_lower:
            data_for_location.update({'fillKey': (fill_format % (0, fill_key_lower))})
        elif fill_key_lower <= value < fill_key_bigger:
            data_for_location.update({'fillKey': (fill_format % (fill_key_lower, fill_key_bigger))})
        elif value >= fill_key_bigger:
            data_for_location.update({'fillKey': (fill_format % (fill_key_bigger, 100))})

    average = (
        (values_to_calculate_average['numerator'] * 100) /
        float(values_to_calculate_average['denominator'] or 1)
    )
    return data_for_map, valid_total, in_month_total, average, total


def calculate_date_for_age(dob, date):
    now = datetime.utcnow().date()
    if now.month == date.month and now.year == date.year:
        date_for_age = now
    else:
        date_for_age = (date + relativedelta(months=1)) - relativedelta(days=1)
    return current_age(dob, date_for_age)


def chosen_filters_to_labels(config, default_interval=''):
    gender_types = {
        'M': 'Male',
        'F': 'Female'
    }

    age_intervals = {
        '6': '0-6 months (0-180 days)',
        '12': '6-12 months (181-365 days)',
        '24': '12-24 months (366-730 days)',
        '36': '24-36 months (731-1095 days)',
        '48': '36-48 months (1096-1460 days)',
        '60': '48-60 months (1461-1825 days)',
        '72': '60-72 months (1826-2190 days)'
    }

    gender = config.get('gender')
    gender_label = ' ({gender})'.format(gender=gender_types.get(gender)) if gender else ''
    chosen_gender = '{gender}'.format(gender=gender_types.get(gender)) if gender else ''

    age = config.get('age_tranche')
    age_in = config.get('age_tranche__in')
    if age:
        chosen_age = '{age}'.format(age=age_intervals.get(age))
    elif age_in:
        chosen_age = '{age}'.format(age=age_intervals.get(age_in[-1]))
    else:
        chosen_age = default_interval

    delimiter = ', ' if gender and chosen_age else ''
    chosen_filters = ' ({gender}{delimiter}{age})'\
        .format(gender=chosen_gender, delimiter=delimiter, age=chosen_age) if gender or chosen_age else ''

    return gender_label, chosen_age, chosen_filters


def zip_folder(pdf_files):
    zip_hash = uuid.uuid4().hex
    icds_file = IcdsFile(blob_id=zip_hash, data_type='issnip_monthly')
    in_memory = BytesIO()
    zip_file = zipfile.ZipFile(in_memory, 'w', zipfile.ZIP_DEFLATED)
    files_to_zip = IcdsFile.objects.filter(blob_id__in=list(pdf_files.keys()), data_type='issnip_monthly')

    for pdf_file in files_to_zip:
        zip_file.writestr(
            'ICDS_CAS_monthly_register_{}.pdf'.format(pdf_files[pdf_file.blob_id]),
            pdf_file.get_file_from_blobdb().read()
        )
    zip_file.close()

    # we need to reset buffer position to the beginning after creating zip, if not read() will return empty string
    # we read this to save file in blobdb
    in_memory.seek(0)
    icds_file.store_file_in_blobdb(in_memory, expired=ONE_DAY)
    icds_file.save()
    return zip_hash


def create_excel_file(excel_data, data_type, file_format, blob_key=None, timeout=ONE_DAY):
    key = blob_key or uuid.uuid4().hex
    export_file = BytesIO()
    icds_file, _ = IcdsFile.objects.get_or_create(blob_id=key, data_type=data_type)
    export_from_tables(excel_data, export_file, file_format)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=timeout)
    icds_file.save()
    return key


def create_pdf_file(pdf_context):
    pdf_hash = uuid.uuid4().hex
    template = get_template("icds_reports/icds_app/pdf/issnip_monthly_register.html")
    resultFile = BytesIO()
    icds_file = IcdsFile(blob_id=pdf_hash, data_type='issnip_monthly')
    try:
        pdf_page = template.render(pdf_context)
    except Exception as ex:
        pdf_page = str(ex)
    base_url = os.path.join(settings.FILEPATH, 'custom', 'icds_reports', 'static')
    resultFile.write(HTML(string=pdf_page, base_url=base_url).write_pdf(
        stylesheets=[CSS(os.path.join(base_url, 'css', 'issnip_monthly_print_style.css')), ])
    )
    # we need to reset buffer position to the beginning after creating pdf, if not read() will return empty string
    # we read this to save file in blobdb
    resultFile.seek(0)

    icds_file.store_file_in_blobdb(resultFile, expired=ONE_DAY)
    icds_file.save()
    return pdf_hash


def generate_qrcode(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4
    )
    qr.add_data(data)
    qr.make(fit=True)
    image = qr.make_image()
    output = BytesIO()
    image.save(output, "PNG")
    qr_content = b64encode(output.getvalue())
    return qr_content


def indian_formatted_number(number):
    s = str(number)
    if s.isdigit():
        r = ",".join([s[x - 2:x] for x in range(-3, -len(s), -2)][::-1] + [s[-3:]])
        return "".join(r)
    else:
        return 0


@quickcache(['domain', 'location_id', 'show_test'], timeout=5 * 60)
def get_child_locations(domain, location_id, show_test):
    if location_id:
        locations = SQLLocation.objects.get(domain=domain, location_id=location_id).get_children()
    else:
        locations = SQLLocation.objects.filter(domain=domain, location_type__code=const.LocationTypes.STATE)

    if not show_test:
        return [
            sql_location for sql_location in locations
            if sql_location.metadata.get('is_test_location', 'real') != 'test'
        ]
    else:
        return list(locations)


def person_has_aadhaar_column(beta):
    return 'cases_person_has_aadhaar_v2'


def person_is_beneficiary_column(beta):
    return 'cases_person_beneficiary_v2'


def wasting_moderate_column(beta):
    return 'wasting_moderate'


def wasting_severe_column(beta):
    return 'wasting_severe'


def wasting_normal_column(beta):
    return 'wasting_normal'


def stunting_moderate_column(beta):
    return 'stunting_moderate'


def stunting_severe_column(beta):
    return 'stunting_severe'


def stunting_normal_column(beta):
    return 'stunting_normal'


def current_month_stunting_column(beta):
    return 'current_month_stunting'


def current_month_wasting_column(beta):
    return 'current_month_wasting'


def hfa_recorded_in_month_column(beta):
    return 'height_measured_in_month'


def wfh_recorded_in_month_column(beta):
    return 'weighed_and_height_measured_in_month'


def default_age_interval(beta):
    return '0 - 5 years'


def get_age_filters(beta):
    return [
        NOT(EQ('age_tranche', 'age_72'))
    ]


def get_age_condition(beta):
    return "age_tranche != :age_72"


def track_time(func):
    """A decorator to track the duration an aggregation script takes to execute"""
    from custom.icds_reports.models import AggregateSQLProfile

    def get_sync_datasource_time():
        return KafkaCheckpoint.objects.filter(checkpoint_id__in=const.UCR_PILLOWS) \
            .exclude(doc_modification_time__isnull=True)\
            .aggregate(Min('doc_modification_time'))['doc_modification_time__min']


    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()

        sync_latest_ds_update = get_sync_datasource_time()

        AggregateSQLProfile.objects.create(
            name=func.__name__,
            duration=int(end - start),
            last_included_doc_time=sync_latest_ds_update
        )
        return result

    return wrapper


def percent_num(x, y):
    return (x or 0) * 100 / float(y or 1)


def percent(x, y):
    return "%.2f %%" % (percent_num(x, y))


def format_decimal(num):
    return "%.2f" % num


def percent_or_not_entered(x, y):
    return percent(x, y) if y and x is not None else DATA_NOT_ENTERED


def format_data_not_entered_to_zero(value):
    if value == DATA_NOT_ENTERED:
        return 0
    return value


def india_now():
    utc_now = datetime.now(pytz.utc)
    india_now = utc_now.astimezone(india_timezone)
    return india_now.strftime("%H:%M:%S %d %B %Y")


def day_suffix(day):
    return 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')


def custom_strftime(format_to_use, date_to_format):
    # adds {S} option to strftime that formats day as 1st, 3rd, 11th etc.
    return date_to_format.strftime(format_to_use).replace(
        '{S}', str(date_to_format.day) + day_suffix(date_to_format.day)
    )


def create_aww_performance_excel_file(excel_data, data_type, month, state, district=None, block=None):
    aggregation_level = 3 if block else (2 if district else 1)
    export_info = excel_data[1][1]
    excel_data = [line[aggregation_level:] for line in excel_data[0][1]]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    warp_text_alignment = Alignment(wrap_text=True)
    bold_font = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="B3C5E5")
    grey_fill = PatternFill("solid", fgColor="BFBFBF")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AWW Performance Report"
    worksheet.sheet_view.showGridLines = False
    # sheet title
    worksheet.merge_cells('B2:{0}2'.format(
        "L" if aggregation_level == 3 else ("M" if aggregation_level == 2 else "N")
    ))
    title_cell = worksheet['B2']
    title_cell.fill = PatternFill("solid", fgColor="4472C4")
    title_cell.value = "AWW Performance Report for the month of {}".format(month)
    title_cell.font = Font(size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")

    # sheet header
    header_cells = ["B3", "C3", "D3", "E3", "F3", "G3", "H3", "I3", "J3", "K3", "L3"]
    if aggregation_level < 3:
        header_cells.append("M3")
    if aggregation_level < 2:
        header_cells.append("N3")

    for cell in header_cells:
        worksheet[cell].fill = blue_fill
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
    worksheet.merge_cells('B3:C3')
    worksheet['B3'].value = "State: {}".format(state)
    if district:
        worksheet['D3'].value = "District: {}".format(district)
    worksheet.merge_cells('E3:F3')
    if block:
        worksheet['E3'].value = "Block: {}".format(block)

    # table header
    table_header_position_row = 5
    headers = ["S.No"]
    if aggregation_level < 2:
        headers.append("District")
    if aggregation_level < 3:
        headers.append("Block")

    headers.extend([
        'Supervisor', 'AWC', 'AWC Site Code', 'AWW Name', 'AWW Contact Number',
        'Home Visits Conducted', 'Weighing Efficiency', 'AWW Eligible for Incentive',
        'Number of Days AWC was Open', 'AWH Eligible for Incentive'
    ])
    columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    if aggregation_level < 3:
        columns.append('M')
    if aggregation_level < 2:
        columns.append('N')

    table_header = {}
    for col, header in zip(columns, headers):
        table_header[col] = header
    for column, value in table_header.items():
        cell = "{}{}".format(column, table_header_position_row)
        worksheet[cell].fill = grey_fill
        worksheet[cell].border = thin_border
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
        worksheet[cell].value = value

    # table contents
    row_position = table_header_position_row + 1

    for enum, row in enumerate(excel_data[1:], start=1):
        for column_index in range(len(columns)):
            column = columns[column_index]
            cell = "{}{}".format(column, row_position)
            worksheet[cell].border = thin_border
            if column_index == 0:
                worksheet[cell].value = enum
            else:
                worksheet[cell].value = row[column_index - 1]
        row_position += 1

    # sheet dimensions
    title_row = worksheet.row_dimensions[2]
    title_row.height = 23
    worksheet.row_dimensions[table_header_position_row].height = 46
    worksheet.row_dimensions[3].height = 16
    widths = {}
    widths_columns = ['A']
    widths_columns.extend(columns)
    standard_widths = [4, 7, 15]
    standard_widths.extend([15] * (3 - aggregation_level))
    standard_widths.extend([13, 12, 12, 13, 15, 11, 14, 14])
    standard_widths.append(14)

    for col, width in zip(widths_columns, standard_widths):
        widths[col] = width
    widths['C'] = max(widths['C'], len(state) * 4 // 3 if state else 0)
    widths['D'] = 13 + (len(district) * 4 // 3 if district else 0)
    widths['F'] = max(widths['F'], len(block) * 4 // 3 if block else 0)
    for column in ["C", "E", "G"]:
        if widths[column] > 25:
            worksheet.row_dimensions[3].height = max(
                16 * ((widths[column] // 25) + 1),
                worksheet.row_dimensions[3].height
            )
            widths[column] = 25
    columns = columns[1:]
    # column widths based on table contents
    for column_index in range(len(columns)):
        widths[columns[column_index]] = max(
            widths[columns[column_index]],
            max(
                len(row[column_index].decode('utf-8') if isinstance(row[column_index], bytes)
                    else str(row[column_index])
                    )
                for row in excel_data[1:]) * 4 // 3 if len(excel_data) >= 2 else 0
        )

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    # saving file
    key = get_performance_report_blob_key(state, district, block, month, 'xlsx')
    export_file = BytesIO()
    icds_file, _ = IcdsFile.objects.get_or_create(blob_id=key, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=None)
    icds_file.save()
    return key


def get_performance_report_blob_key(state, district, block, month, file_format):
    key_safe_date = datetime.strptime(month, '%B %Y').strftime('%Y_%m')
    key = 'performance_report-{}-{}-{}-{}-{}'.format(state, district, block, key_safe_date, file_format)
    safe_key = key.replace(' ', '_')
    return safe_id(safe_key)


def create_excel_file_in_openpyxl(excel_data, data_type):
    workbook = Workbook()
    first_worksheet = True
    for worksheet_data in excel_data:
        if first_worksheet:
            worksheet = workbook.active
            worksheet.title = worksheet_data[0]
            first_worksheet = False
        else:
            worksheet = workbook.create_sheet(worksheet_data[0])
        for row_number, row_data in enumerate(worksheet_data[1], start=1):
            for column_number, cell_data in enumerate(row_data, start=1):
                worksheet.cell(row=row_number, column=column_number).value = cell_data

    # saving file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()
    return file_hash


def create_thr_report_excel_file(excel_data, data_type, month, aggregation_level, report_type='consolidated',
                                 beta=False):
    export_info = excel_data[1][1]
    national = 'National Level' if aggregation_level == 0 else ''
    state = export_info[1][1] if aggregation_level > 0 else ''
    district = export_info[2][1] if aggregation_level > 1 else ''
    block = export_info[3][1] if aggregation_level > 2 else ''
    supervisor = export_info[4][1] if aggregation_level > 3 else ''

    excel_data = [line[aggregation_level:] for line in excel_data[0][1]]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    warp_text_alignment = Alignment(wrap_text=True)
    bold_font = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="B3C5E5")
    grey_fill = PatternFill("solid", fgColor="BFBFBF")
    purple_fill = PatternFill("solid", fgColor="CFC6F5")

    workbook = Workbook()
    worksheet = workbook.active
    # sheet title
    worksheet.title = "THR Report"
    worksheet.sheet_view.showGridLines = False

    thr_days_info = ""
    if report_type == THR_REPORT_DAY_BENEFICIARY_TYPE:
        total_column_count = 30
        data_start_row_diff = 3
        secondary_headers = ['Not provided',
                             'Provided for 1-7 days',
                             'Provided for 8-14 days',
                             'Provided for 15-20 days',
                             'Provided for 21-24 days',
                             'Provided for at least 25 days (>=25 days)']

    else:
        if report_type == THR_REPORT_BENEFICIARY_TYPE:
            total_column_count = 15
            data_start_row_diff = 2

        else:
            total_column_count = 11
            data_start_row_diff = 1

        if parse(month).date() <= THR_21_DAYS_THRESHOLD_DATE:
            thr_days_info = "for at least 21 days"
        else:
            thr_days_info = "for at least 25 days"

    if report_type != THR_REPORT_CONSOLIDATED:
        beneficiary_type_columns = [
            'Pregnant women',
            'Lactating women',
            'Children (6-36 months)'
        ]

    amount_of_columns = total_column_count - aggregation_level
    last_column = get_column_letter(amount_of_columns)
    worksheet.merge_cells('A2:{0}2'.format(last_column))
    title_cell = worksheet['A2']
    title_cell.fill = PatternFill("solid", fgColor="4472C4")
    title_cell.value = "Take Home Ration(THR) Report for the {}".format(month)
    title_cell.font = Font(size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")

    columns = [get_column_letter(i) for i in range(1, amount_of_columns + 1)]

    # sheet header
    header_cells = ['{0}3'.format(column) for column in columns]
    for cell in header_cells:
        worksheet[cell].fill = blue_fill
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment

    if national:
        worksheet['A3'].value = national
        worksheet.merge_cells('A3:B3')
    else:
        if state:
            worksheet['A3'].value = "State: {}".format(state)
            worksheet.merge_cells('A3:B3')
        if district:
            worksheet['C3'].value = "District: {}".format(district)
        if block:
            worksheet['D3'].value = "Block: {}".format(block)
        if supervisor:
            worksheet['E3'].value = "Sector: {}".format(supervisor)

    date_cell = '{0}3'.format(last_column)
    date_description_cell = '{0}3'.format(get_column_letter(amount_of_columns - 1))
    worksheet[date_description_cell].value = "Date when downloaded:"
    worksheet[date_description_cell].alignment = Alignment(horizontal="right")
    utc_now = datetime.now(pytz.utc)
    now_in_india = utc_now.astimezone(india_timezone)
    worksheet[date_cell].value = custom_strftime('{S} %b %Y', now_in_india)
    worksheet[date_cell].alignment = Alignment(horizontal="right")

    # table header
    table_header_position_row = 5
    headers = ["S.No"]
    main_headers = ['State', 'District', 'Block', 'Sector', 'Awc Name', 'AWW Name', 'AWW Phone No.',
                   'Total No. of Beneficiaries eligible for THR',
                   f'Total No. of beneficiaries received THR {thr_days_info} in given month',
                   'Total No of Pictures taken by AWW']
    headers.extend(main_headers[aggregation_level:])

    def set_beneficiary_columns(start_column_index, end_column_index, row):
        for i in range(end_column_index - start_column_index + 1):
            cell = "{}{}".format(columns[start_column_index + i], row)
            worksheet[cell].fill = purple_fill
            worksheet[cell].border = thin_border
            worksheet[cell].font = bold_font
            worksheet[cell].alignment = warp_text_alignment
            worksheet[cell].value = beneficiary_type_columns[i % len(beneficiary_type_columns)]

    def set_service_delivery_columns(start_column_index, row):
        for i in range(6):
            column_index = start_column_index + i * 3
            cell = "{}{}".format(columns[column_index], row)
            worksheet[cell].fill = blue_fill
            worksheet[cell].border = thin_border
            worksheet[cell].font = bold_font
            worksheet[cell].alignment = warp_text_alignment
            worksheet[cell].value = secondary_headers[i]
            next_cell = "{}{}".format(columns[column_index + 2], row)
            worksheet.merge_cells(f"{cell}:{next_cell}")

    next_deviated_column = 0
    column_deviation_2 = 2
    column_deviation_17 = 17
    for index, value in enumerate(headers):
        column_index = index + next_deviated_column
        cell = "{}{}".format(columns[column_index], table_header_position_row)

        worksheet[cell].fill = grey_fill
        worksheet[cell].border = thin_border
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
        worksheet[cell].value = value

        if report_type == THR_REPORT_BENEFICIARY_TYPE:
            if value in ('Total No. of Beneficiaries eligible for THR',
                         f'Total No. of beneficiaries received THR {thr_days_info} in given month'):
                next_deviated_column += column_deviation_2
                next_cell = "{}{}".format(columns[column_index + column_deviation_2],
                                          table_header_position_row + data_start_row_diff - 2)
                worksheet.merge_cells(f'{cell}:{next_cell}')
                set_beneficiary_columns(column_index, column_index + column_deviation_2,
                                        table_header_position_row + data_start_row_diff - 1)
            else:
                next_cell = "{}{}".format(columns[column_index], table_header_position_row+ data_start_row_diff - 1)
                worksheet.merge_cells(f'{cell}:{next_cell}')

        elif report_type == THR_REPORT_DAY_BENEFICIARY_TYPE:
            if value == 'Total No. of Beneficiaries eligible for THR':
                next_deviated_column += column_deviation_2
                next_cell = "{}{}".format(columns[column_index + column_deviation_2],
                                          table_header_position_row + data_start_row_diff - 2)
                worksheet.merge_cells(f'{cell}:{next_cell}')
                set_beneficiary_columns(column_index, column_index + column_deviation_2,
                                        table_header_position_row + data_start_row_diff - 1)
            elif value == f'Total No. of beneficiaries received THR {thr_days_info} in given month':
                next_deviated_column += column_deviation_17
                next_cell = "{}{}".format(columns[column_index + column_deviation_17], table_header_position_row)
                worksheet.merge_cells(f'{cell}:{next_cell}')
                set_service_delivery_columns(column_index,
                                             table_header_position_row + data_start_row_diff - 2)
                set_beneficiary_columns(column_index, column_index + column_deviation_17,
                                        table_header_position_row + data_start_row_diff - 1)
            else:
                next_cell = "{}{}".format(columns[column_index],
                                          table_header_position_row + data_start_row_diff - 1)
                worksheet.merge_cells(f'{cell}:{next_cell}')


    # table contents
    row_position = table_header_position_row + data_start_row_diff

    for enum, row in enumerate(excel_data[1:], start=1):
        for column_index in range(len(columns)):
            column = columns[column_index]
            cell = "{}{}".format(column, row_position)
            worksheet[cell].border = thin_border
            if column_index == 0:
                worksheet[cell].value = enum
            else:
                worksheet[cell].value = row[column_index - 1]
        row_position += 1

    # sheet dimensions
    title_row = worksheet.row_dimensions[2]
    title_row.height = 23
    worksheet.row_dimensions[table_header_position_row].height = 46
    widths = {}
    widths_columns = ['A']
    widths_columns.extend(columns)
    standard_widths = [4, 7]
    standard_widths.extend([15] * (4 - aggregation_level))
    standard_widths.extend([25, 15, 25])
    standard_widths += [15] * (len(widths_columns) - len(standard_widths))

    for col, width in zip(widths_columns, standard_widths):
        widths[col] = width

    widths['C'] = max(widths['C'], len(state) * 4 // 3 if state else 0)
    widths['D'] = 9 + (len(district) * 4 // 3 if district else 0)
    widths['E'] = 8 + (len(block) * 4 // 3 if district else 0)
    widths['F'] = 8 + (len(supervisor) * 4 // 3 if district else 0)

    columns = columns[1:]
    # column widths based on table contents
    for column_index in range(len(columns)):
        widths[columns[column_index]] = max(
            widths[columns[column_index]],
            max(
                len(row[column_index].decode('utf-8') if isinstance(row[column_index], bytes)
                    else str(row[column_index])
                    )
                for row in excel_data[1:]) * 4 // 3 if len(excel_data) >= 2 else 0
        )

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    # saving file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()
    return file_hash


def create_child_report_excel_file(excel_data, data_type, month, aggregation_level):
    export_info = excel_data[1][1]
    max_merged_column_no = aggregation_level
    if aggregation_level == 5:
        max_merged_column_no = aggregation_level + 2

    primary_headers = ['Children weighed','Height measured for Children', '', 'Severely Underweight Children',
                       'Moderately Underweight Children','Children with normal weight for age (WfA)',
                       'Severely wasted Children(SAM)', 'Moderately wasted children(MAM)',
                       'Children with normal weight for height',
                       'Severely stunted children', 'Moderately Stunted Children',
                       'Children with normal height for age',
                       'Newborn with Low birth weight', 'Children completed immunization prescribed for 1 year',
                       'Children breastfed at the time of birth','Children exclusively breastfed',
                       'Children initiated with complementary feeding',
                       'Children initiated with complementary feeding appropriately',
                       'Children initiated with complementary feeding with adequate diet diversity',
                       'Children initiated with complementary feeding with adequate diet quantity',
                       'Children initiated with complementary feeding with appropriate handwashing before feeding'
                       ]

    location_padding_columns = ([''] * max_merged_column_no)
    primary_headers = location_padding_columns + primary_headers

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Children"

    # Styling initialisation
    bold_font = Font(size=14, color="FFFFFF")
    bold_font_black = Font(size=14, color="000000")
    cell_pattern = PatternFill("solid", fgColor="B3C5E5")
    cell_pattern_blue = PatternFill("solid", fgColor="4472C4")
    cell_pattern_red = PatternFill("solid", fgColor="ff0000")
    cell_pattern_yellow = PatternFill("solid", fgColor="ffff00")
    text_alignment = Alignment(horizontal="center", vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Primary Header
    main_header = worksheet.row_dimensions[1]
    main_header.height = 40
    current_column_location = 1
    for index, primary_header in enumerate(primary_headers):
        cell_name = get_column_letter(current_column_location)
        cell = worksheet['{}1'.format(cell_name)]
        cell.alignment = text_alignment

        cell.value = primary_header

        if primary_header == 'Severely wasted Children(SAM)':
            cell.fill = cell_pattern_red
            cell.font = bold_font
        elif primary_header == 'Moderately wasted children(MAM)':
            cell.fill = cell_pattern_yellow
            cell.font = bold_font_black
        else:
            cell.fill = cell_pattern_blue
            cell.font = bold_font

        if current_column_location <= max_merged_column_no or current_column_location == max_merged_column_no + 7:
            worksheet.merge_cells('{}1:{}2'.format(get_column_letter(current_column_location),
                                                   get_column_letter(current_column_location)))
            current_column_location += 1
        else:
            worksheet.merge_cells('{}1:{}1'.format(get_column_letter(current_column_location),
                                                   get_column_letter(current_column_location + 2)))
            current_column_location += 3

    # Secondary Header
    secondary_header = worksheet.row_dimensions[2]
    secondary_header.height = 80
    headers = excel_data[0][1][0]
    bold_font_black = Font(size=14)
    for index, header in enumerate(headers):
        location_column = get_column_letter(index + 1)
        cell = worksheet['{}{}'.format(location_column, 1 if index+1 <= max_merged_column_no or
                                                             index == max_merged_column_no + 6 else 2)]
        cell.alignment = text_alignment
        worksheet.column_dimensions[location_column].width = 30
        cell.value = header
        if index != max_merged_column_no + 6 and index + 1 > max_merged_column_no:
            cell.fill = cell_pattern
            cell.font = bold_font_black
            cell.border = thin_border

    # Fill data
    for row_index,row in enumerate(excel_data[0][1][1:]):
        for col_index, col_value in enumerate(row):
            row_num = row_index + 3
            column_name = get_column_letter(col_index + 1)
            cell = worksheet['{}{}'.format(column_name, row_num)]
            cell.value = col_value
            cell.border = thin_border

    # Export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    #Export to icds file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()

    return file_hash


def create_service_delivery_report(excel_data, data_type, config):

    export_info = excel_data[1][1]
    location_padding_columns = ([''] * config['aggregation_level'])

    if config['beneficiary_category'] == 'pw_lw_children':
        primary_headers = location_padding_columns + ['Home Visits',
                                                      'Growth Monitoring',
                                                      'Community Based Events',
                                                      'VHSND',
                                                      'Take Home Ration  (Pregnant women, lactating women and children 0-3 years)',
                                                      ]
        if config['aggregation_level'] < 5:
            primary_headers = [''] + primary_headers
    else:
        primary_headers = location_padding_columns + ['Supplementary Nutrition (Children 3-6 years)',
                                                      'Pre-school Education (Children 3-6 years)',
                                                      'Growth Monitoring (Children 3-5 years)',
                                                      ]

    secondary_headers = ['Not provided',
                         'Provided for 1-7 days',
                         'Provided for 8-14 days',
                         'Provided for 15-20 days',
                         'Provided for 21-24 days',
                         'Provided for at least 25 days (>=25 days)']

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SDR"

    # Styling initialisation
    bold_font = Font(size=14, color="FFFFFF")
    bold_font_black = Font(size=14, color="000000")
    bold_font_black_normal = Font(color="000000", bold=True)
    cell_pattern = PatternFill("solid", fgColor="B3C5E5")
    cell_pattern_blue = PatternFill("solid", fgColor="4472C4")
    cell_pattern_grey = PatternFill("solid", fgColor="C3C3C3")
    text_alignment = Alignment(horizontal="center", vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_column_location = 1
    location_header_count = 0
    for index, primary_header in enumerate(primary_headers):
        cell_name = get_column_letter(current_column_location)
        cell = worksheet['{}1'.format(cell_name)]
        cell.alignment = text_alignment

        cell.value = primary_header
        cell.fill = cell_pattern_blue
        cell.font = bold_font
        cell.border = thin_border
        if primary_header == '':
            worksheet.merge_cells('{}1:{}3'.format(get_column_letter(current_column_location),
                                                   get_column_letter(current_column_location)))
            location_header_count += 1
            current_column_location += 1
        elif primary_header not in [
            'Take Home Ration  (Pregnant women, lactating women and children 0-3 years)',
            'Supplementary Nutrition (Children 3-6 years)',
            'Pre-school Education (Children 3-6 years)'
        ]:
            merging_width = 2
            if primary_header == 'Community Based Events' and config['aggregation_level'] == 5:
                merging_width = 1

            if primary_header == 'VHSND':
                if config['aggregation_level'] == 5:
                    merging_width = 1
                else:
                    merging_width = 0

            worksheet.merge_cells('{}1:{}2'.format(get_column_letter(current_column_location),
                                                   get_column_letter(current_column_location + merging_width)))
            current_column_location += merging_width+1
        else:
            worksheet.merge_cells('{}1:{}1'.format(get_column_letter(current_column_location),
                                                   get_column_letter(current_column_location + 17)))

            current_column_location_sec_header = current_column_location
            for sec_header in secondary_headers:
                cell_name = get_column_letter(current_column_location_sec_header)
                cell = worksheet['{}2'.format(cell_name)]
                cell.alignment = text_alignment
                cell.value = sec_header
                cell.fill = cell_pattern_grey
                cell.font = bold_font_black
                cell.border = thin_border
                worksheet.merge_cells('{}2:{}2'.format(get_column_letter(current_column_location_sec_header),
                                                       get_column_letter(current_column_location_sec_header + 2)))
                current_column_location_sec_header += 3

            current_column_location += 18

    # Secondary Header
    headers = excel_data[0][1][0]
    for index, header in enumerate(headers):
        location_column = get_column_letter(index + 1)
        cell = worksheet['{}{}'.format(location_column, 1 if index+1 <= location_header_count else 3)]
        cell.alignment = text_alignment
        worksheet.column_dimensions[location_column].width = 30
        cell.value = header
        cell.border = thin_border
        if index + 1 > location_header_count:
            cell.fill = cell_pattern

    # Fill data
    for row_index, row in enumerate(excel_data[0][1][1:]):
        for col_index, col_value in enumerate(row):
            row_num = row_index + 4
            column_name = get_column_letter(col_index + 1)
            cell = worksheet['{}{}'.format(column_name, row_num)]
            cell.value = col_value
            cell.border = thin_border
            if col_value == 'Grand Total':
                cell.font = bold_font_black_normal
                cell.fill = cell_pattern

    # Export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()

    return file_hash


def create_lady_supervisor_excel_file(excel_data, data_type, month, aggregation_level):
    export_info = excel_data[1][1]
    state = export_info[1][1] if aggregation_level > 0 else ''
    district = export_info[2][1] if aggregation_level > 1 else ''
    block = export_info[3][1] if aggregation_level > 2 else ''
    excel_data = [line[aggregation_level:] for line in excel_data[0][1]]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    warp_text_alignment = Alignment(wrap_text=True)
    bold_font = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="B3C5E5")
    grey_fill = PatternFill("solid", fgColor="BFBFBF")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "LS Performance Report"
    worksheet.sheet_view.showGridLines = False
    # sheet title
    amount_of_columns = 9 - aggregation_level
    last_column = string.ascii_uppercase[amount_of_columns]
    worksheet.merge_cells('B2:{0}2'.format(last_column))
    title_cell = worksheet['B2']
    title_cell.fill = PatternFill("solid", fgColor="4472C4")
    title_cell.value = "Lady Supervisor Performance Report for the {}".format(month)
    title_cell.font = Font(size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")

    columns = [string.ascii_uppercase[i] for i in range(1, amount_of_columns + 1)]

    # sheet header
    header_cells = ['{0}3'.format(column) for column in columns]
    for cell in header_cells:
        worksheet[cell].fill = blue_fill
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
    if state:
        worksheet['B3'].value = "State: {}".format(state)
        worksheet.merge_cells('B3:C3')
    if district:
        worksheet['D3'].value = "District: {}".format(district)
    if block:
        worksheet['E3'].value = "Block: {}".format(block)
    date_cell = '{0}3'.format(last_column)
    date_description_cell = '{0}3'.format(string.ascii_uppercase[amount_of_columns - 1])
    worksheet[date_description_cell].value = "Date when downloaded:"
    worksheet[date_description_cell].alignment = Alignment(horizontal="right")
    utc_now = datetime.now(pytz.utc)
    now_in_india = utc_now.astimezone(india_timezone)
    worksheet[date_cell].value = custom_strftime('{S} %b %Y', now_in_india)
    worksheet[date_cell].alignment = Alignment(horizontal="right")

    # table header
    table_header_position_row = 5
    header_data = excel_data[0]
    headers = ["S.No"]
    headers.extend(header_data)

    table_header = {}
    for col, header in zip(columns, headers):
        table_header[col] = header
    for column, value in table_header.items():
        cell = "{}{}".format(column, table_header_position_row)
        worksheet[cell].fill = grey_fill
        worksheet[cell].border = thin_border
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
        worksheet[cell].value = value

    # table contents
    row_position = table_header_position_row + 1

    for enum, row in enumerate(excel_data[1:], start=1):
        for column_index in range(len(columns)):
            column = columns[column_index]
            cell = "{}{}".format(column, row_position)
            worksheet[cell].border = thin_border
            if column_index == 0:
                worksheet[cell].value = enum
            else:
                worksheet[cell].value = row[column_index - 1]
        row_position += 1

    # sheet dimensions
    title_row = worksheet.row_dimensions[2]
    title_row.height = 23
    worksheet.row_dimensions[table_header_position_row].height = 46
    widths = {}
    widths_columns = ['A']
    widths_columns.extend(columns)
    standard_widths = [4, 7]
    standard_widths.extend([15] * (4 - aggregation_level))
    standard_widths.extend([25, 15, 25, 15])
    for col, width in zip(widths_columns, standard_widths):
        widths[col] = width
    widths['C'] = max(widths['C'], len(state) * 4 // 3 if state else 0)
    widths['D'] = 9 + (len(district) * 4 // 3 if district else 0)
    widths['E'] = 8 + (len(block) * 4 // 3 if district else 0)

    columns = columns[1:]
    # column widths based on table contents
    for column_index in range(len(columns)):
        widths[columns[column_index]] = max(
            widths[columns[column_index]],
            max(
                len(row[column_index].decode('utf-8') if isinstance(row[column_index], bytes)
                    else str(row[column_index])
                    )
                for row in excel_data[1:]) * 4 // 3 if len(excel_data) >= 2 else 0
        )

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    # saving file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()
    return file_hash


def get_dashboard_usage_excel_file(excel_data, data_type):
    export_info = excel_data[1][1]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Dashboard Activity"
    bold_font = Font(size=11, color="FFFFFF", bold=True)
    cell_pattern_blue = PatternFill("solid", fgColor="3387E3")
    text_alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
    thin_border_no_top = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style=None),
        bottom=Side(style='thin')
    )

    thin_border_no_bottom = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style=None)

    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')

    )

    # # Primary Header
    # main_header = worksheet.row_dimensions[2]
    # main_header.height = 30
    # for index, primary_header in enumerate(primary_headers):
    #     cell_name = get_column_letter(current_column_location)
    #     cell = worksheet['{}1'.format(cell_name)]
    #     cell.alignment = text_alignment
    #
    #     cell.value = primary_header
    #     cell.fill = cell_pattern_blue
    #     cell.font = bold_font
    #     cell.border = thin_border_no_bottom

    # Secondary Header
    secondary_header = worksheet.row_dimensions[1]
    secondary_header.height = 40
    headers = excel_data[0][1][0]
    bold_font_black = Font(size=11)
    for index, header in enumerate(headers):
        location_column = get_column_letter(index + 1)
        cell = worksheet['{}1'.format(location_column)]
        cell.alignment = text_alignment
        worksheet.column_dimensions[location_column].width = 30
        cell.value = header
        cell.fill = cell_pattern_blue
        cell.font = bold_font
        cell.border = thin_border_no_top

    # Fill data
    for row_index, row in enumerate(excel_data[0][1][1:]):
        worksheet.row_dimensions[row_index + 2].height = 20
        for col_index, col_value in enumerate(row):
            row_num = row_index + 2
            column_name = get_column_letter(col_index + 1)
            cell = worksheet['{}{}'.format(column_name, row_num)]
            cell.value = col_value
            cell.border = thin_border
            cell.font = bold_font_black

    # Export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    #Export to icds file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()

    return file_hash


def create_child_growth_tracker_report(excel_data, data_type, config, aggregation_level):
    month = config['month']
    export_info = excel_data[1][1]
    national = 'National Level' if aggregation_level == 0 else ''
    state = export_info[1][1] if aggregation_level > 0 else ''
    district = export_info[2][1] if aggregation_level > 1 else ''
    block = export_info[3][1] if aggregation_level > 2 else ''
    supervisor = export_info[3][1] if aggregation_level > 3 else ''

    excel_data = excel_data[0][1]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    warp_text_alignment = Alignment(wrap_text=True)
    bold_font = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="B3C5E5")
    grey_fill = PatternFill("solid", fgColor="BFBFBF")

    workbook = Workbook()
    worksheet = workbook.active
    # sheet title
    worksheet.title = "Child Growth Tracker Report"
    worksheet.sheet_view.showGridLines = False
    amount_of_columns = 25
    last_column = string.ascii_uppercase[amount_of_columns]
    worksheet.merge_cells('B2:{0}2'.format(last_column))
    title_cell = worksheet['B2']
    title_cell.fill = PatternFill("solid", fgColor="4472C4")
    title_cell.value = "Child Growth Tracker Report for the {}".format(month)
    title_cell.font = Font(size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")

    columns = [string.ascii_uppercase[i] for i in range(1, amount_of_columns + 1)]

    # sheet header
    header_cells = ['{0}3'.format(column) for column in columns]
    for cell in header_cells:
        worksheet[cell].fill = blue_fill
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment

    if national:
        worksheet['B3'].value = national
        worksheet.merge_cells('B3:C3')
    else:
        if state:
            worksheet['B3'].value = "State: {}".format(state)
            worksheet.merge_cells('B3:C3')
        if district:
            worksheet['D3'].value = "District: {}".format(district)
        if block:
            worksheet['E3'].value = "Block: {}".format(block)
        if supervisor:
            worksheet['F3'].value = "Sector: {}".format(supervisor)

    date_cell = '{0}3'.format(last_column)
    date_description_cell = '{0}3'.format(string.ascii_uppercase[amount_of_columns - 1])
    worksheet[date_description_cell].value = "Date when downloaded:"
    worksheet[date_description_cell].alignment = Alignment(horizontal="right")
    utc_now = datetime.now(pytz.utc)
    now_in_india = utc_now.astimezone(india_timezone)
    worksheet[date_cell].value = custom_strftime('{S} %b %Y', now_in_india)
    worksheet[date_cell].alignment = Alignment(horizontal="right")

    # table header
    table_header_position_row = 6
    header_data = excel_data[0]
    headers = ["S.No"]
    headers.extend(header_data)

    table_header = {}
    for col, header in zip(columns, headers):
        table_header[col] = header.split('_')[1] if len(header.split('_')) > 1 else header
    for column, value in table_header.items():
        cell = "{}{}".format(column, table_header_position_row)
        worksheet[cell].fill = blue_fill
        worksheet[cell].border = thin_border
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
        worksheet[cell].value = value
    cols = ["Pre-school Education Attendance", "Supplementary Nutrition", "Stunting (height-for-age)",
            "Wasting (weight-for-height)", "Underweight (weight-for-age)"]
    start_col = 11
    for col in cols:
        cell = f"{string.ascii_uppercase[start_col]}{table_header_position_row-1}"
        worksheet[cell].fill = grey_fill
        worksheet[cell].border = thin_border
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = Alignment(horizontal="center")
        worksheet[cell].value = col
        start_col = start_col + 3
        merge_cell = f"{string.ascii_uppercase[start_col - 1]}{table_header_position_row - 1}"
        worksheet.merge_cells(f"{cell}:{merge_cell}")
    # table contents
    row_position = table_header_position_row + 1
    for enum, row in enumerate(excel_data[1:], start=1):
        for column_index in range(len(columns)):
            column = columns[column_index]
            cell = "{}{}".format(column, row_position)
            worksheet[cell].border = thin_border
            if column_index == 0:
                worksheet[cell].value = enum
            else:
                worksheet[cell].value = row[column_index - 1]
        row_position += 1

    # sheet dimensions
    title_row = worksheet.row_dimensions[2]
    title_row.height = 23
    worksheet.row_dimensions[table_header_position_row].height = 46
    widths = {}
    widths_columns = ['A']
    widths_columns.extend(columns)
    standard_widths = [4, 7]
    standard_widths.extend([15] * 24)
    for col, width in zip(widths_columns, standard_widths):
        widths[col] = width

    widths['C'] = max(widths['C'], len(state) * 4 // 3)
    widths['D'] = 9 + (len(district) * 4 // 3)
    widths['E'] = 8 + (len(block) * 4 // 3)
    widths['F'] = 8 + (len(supervisor) * 4 // 3)

    columns = columns[1:]
    # column widths based on table contents
    for column_index in range(len(columns)):
        widths[columns[column_index]] = max(
            widths[columns[column_index]],
            max(
                len(row[column_index].decode('utf-8') if isinstance(row[column_index], bytes)
                    else str(row[column_index])
                    )
                for row in excel_data[1:]) * 4 // 3 if len(excel_data) >= 2 else 0
        )

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    # saving file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()
    return file_hash

def create_poshan_progress_report(excel_data, data_type, config, aggregation_level):
    export_info = excel_data[1][1]
    layout = config['report_layout']
    national = 'National Level' if len(export_info) == 5 else ''
    state = export_info[1][1] if len(export_info) > 5 else ''
    district = export_info[2][1] if len(export_info) > 6 else ''
    block = export_info[3][1] if len(export_info) > 7 else ''
    supervisor = export_info[3][1] if len(export_info) > 8 else ''

    excel_data = excel_data[0][1]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    warp_text_alignment = Alignment(wrap_text=True)
    bold_font = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="B3C5E5")
    grey_fill = PatternFill("solid", fgColor="BFBFBF")

    workbook = Workbook()
    worksheet = workbook.active
    # sheet title
    worksheet.title = "PPR {}".format(layout)
    worksheet.sheet_view.showGridLines = False
    amount_of_columns = 1 + len(excel_data[0])
    last_column = get_column_letter(amount_of_columns+1)
    worksheet.merge_cells('B2:{0}2'.format(last_column))
    title_cell = worksheet['B2']
    title_cell.fill = PatternFill("solid", fgColor="4472C4")
    title_cell.value = "Poshan Progress Report {}".format(layout.title())
    title_cell.font = Font(size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")

    columns = [get_column_letter(i) for i in range(2, amount_of_columns + 2)]

    # sheet header
    header_cells = ['{0}3'.format(column) for column in columns]
    for cell in header_cells:
        worksheet[cell].fill = blue_fill
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment

    if national:
        worksheet['B3'].value = national
        worksheet.merge_cells('B3:C3')
    else:
        if state:
            worksheet['B3'].value = "State: {}".format(state)
            worksheet.merge_cells('B3:C3')
        if district:
            worksheet['D3'].value = "District: {}".format(district)
        if block:
            worksheet['E3'].value = "Block: {}".format(block)
        if supervisor:
            worksheet['F3'].value = "Sector: {}".format(supervisor)

    date_cell = '{0}3'.format(last_column)
    date_description_cell = '{0}3'.format(get_column_letter(amount_of_columns))
    worksheet[date_description_cell].value = "Date when downloaded:"
    worksheet[date_description_cell].alignment = Alignment(horizontal="right")
    utc_now = datetime.now(pytz.utc)
    now_in_india = utc_now.astimezone(india_timezone)
    worksheet[date_cell].value = custom_strftime('{S} %b %Y', now_in_india)
    worksheet[date_cell].alignment = Alignment(horizontal="right")

    # table header
    table_header_position_row = 5
    header_data = excel_data[0]
    headers = ["S.No"]
    headers.extend(header_data)

    table_header = {}
    for col, header in zip(columns, headers):
        table_header[col] = header
    for column, value in table_header.items():
        cell = "{}{}".format(column, table_header_position_row)
        worksheet[cell].fill = grey_fill
        worksheet[cell].border = thin_border
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
        worksheet[cell].value = value

    # table contents
    row_position = table_header_position_row + 1

    for enum, row in enumerate(excel_data[1:], start=1):
        for column_index in range(len(columns)):
            column = columns[column_index]
            cell = "{}{}".format(column, row_position)
            worksheet[cell].border = thin_border
            if column_index == 0:
                worksheet[cell].value = enum
            else:
                worksheet[cell].value = row[column_index - 1]
        row_position += 1

    # sheet dimensions
    title_row = worksheet.row_dimensions[2]
    title_row.height = 23
    worksheet.row_dimensions[table_header_position_row].height = 46
    widths = {}
    widths_columns = ['A']
    widths_columns.extend(columns)
    standard_widths = [4, 7]
    standard_widths.extend([15] * (len(columns) - 1))
    for col, width in zip(widths_columns, standard_widths):
        widths[col] = width

    widths['C'] = max(widths['C'], len(state) * 4 // 3 if state else 0)
    widths['D'] = 9 + (len(district) * 4 // 3 if district else 0)
    columns = columns[1:]
    # column widths based on table contents
    for column_index in range(len(columns)):
        widths[columns[column_index]] = max(
            widths[columns[column_index]],
            max(
                len(row[column_index].decode('utf-8') if isinstance(row[column_index], bytes)
                    else str(row[column_index])
                    )
                for row in excel_data[1:]) * 4 // 3 if len(excel_data) >= 2 else 0
        )

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    # saving file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()
    return file_hash


def create_aww_activity_report(excel_data, data_type, config, aggregation_level):
    export_info = excel_data[1][1]
    national = 'National Level' if aggregation_level == 0 else ''
    state = export_info[1][1] if aggregation_level > 0 else ''
    district = export_info[2][1] if aggregation_level > 1 else ''
    block = export_info[3][1] if aggregation_level > 2 else ''
    supervisor = export_info[3][1] if aggregation_level > 3 else ''

    excel_data = excel_data[0][1]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    warp_text_alignment = Alignment(wrap_text=True)
    bold_font = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="B3C5E5")
    grey_fill = PatternFill("solid", fgColor="BFBFBF")

    workbook = Workbook()
    worksheet = workbook.active
    # sheet title
    worksheet.title = "AWW Activity Report"
    worksheet.sheet_view.showGridLines = False
    amount_of_columns = 11
    last_column = string.ascii_uppercase[amount_of_columns]
    worksheet.merge_cells('B2:{0}2'.format(last_column))
    title_cell = worksheet['B2']
    title_cell.fill = PatternFill("solid", fgColor="4472C4")
    title_cell.value = "AWW Activity Report"
    title_cell.font = Font(size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")

    columns = [string.ascii_uppercase[i] for i in range(1, amount_of_columns + 1)]

    # sheet header
    header_cells = ['{0}3'.format(column) for column in columns]
    for cell in header_cells:
        worksheet[cell].fill = blue_fill
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment

    if national:
        worksheet['B3'].value = national
        worksheet.merge_cells('B3:C3')
    else:
        if state:
            worksheet['B3'].value = "State: {}".format(state)
            worksheet.merge_cells('B3:C3')
        if district:
            worksheet['D3'].value = "District: {}".format(district)
        if block:
            worksheet['E3'].value = "Block: {}".format(block)
        if supervisor:
            worksheet['F3'].value = "Sector: {}".format(supervisor)

    date_cell = '{0}3'.format(last_column)
    date_description_cell = '{0}3'.format(string.ascii_uppercase[amount_of_columns - 1])
    worksheet[date_description_cell].value = "Date when downloaded:"
    worksheet[date_description_cell].alignment = Alignment(horizontal="right")
    utc_now = datetime.now(pytz.utc)
    now_in_india = utc_now.astimezone(india_timezone)
    worksheet[date_cell].value = custom_strftime('{S} %b %Y', now_in_india)
    worksheet[date_cell].alignment = Alignment(horizontal="right")

    # table header
    table_header_position_row = 5
    header_data = excel_data[0]
    headers = ["S.No"]
    headers.extend(header_data)

    table_header = {}
    for col, header in zip(columns, headers):
        table_header[col] = header
    for column, value in table_header.items():
        cell = "{}{}".format(column, table_header_position_row)
        worksheet[cell].fill = grey_fill
        worksheet[cell].border = thin_border
        worksheet[cell].font = bold_font
        worksheet[cell].alignment = warp_text_alignment
        worksheet[cell].value = value

    # table contents
    row_position = table_header_position_row + 1

    for enum, row in enumerate(excel_data[1:], start=1):
        for column_index in range(len(columns)):
            column = columns[column_index]
            cell = "{}{}".format(column, row_position)
            worksheet[cell].border = thin_border
            if column_index == 0:
                worksheet[cell].value = enum
            else:
                worksheet[cell].value = row[column_index - 1]
        row_position += 1

    # sheet dimensions
    title_row = worksheet.row_dimensions[2]
    title_row.height = 23
    worksheet.row_dimensions[table_header_position_row].height = 46
    widths = {}
    widths_columns = ['A']
    widths_columns.extend(columns)
    standard_widths = [4, 7]
    standard_widths.extend([15] * 10)
    for col, width in zip(widths_columns, standard_widths):
        widths[col] = width

    widths['C'] = max(widths['C'], len(state) * 4 // 3 if state else 0)
    widths['D'] = 9 + (len(district) * 4 // 3 if district else 0)
    widths['E'] = 8 + (len(block) * 4 // 3 if district else 0)
    widths['F'] = 8 + (len(supervisor) * 4 // 3 if district else 0)
    columns = columns[1:]
    # column widths based on table contents
    for column_index in range(len(columns)):
        widths[columns[column_index]] = max(
            widths[columns[column_index]],
            max(
                len(row[column_index].decode('utf-8') if isinstance(row[column_index], bytes)
                    else str(row[column_index])
                    )
                for row in excel_data[1:]) * 4 // 3 if len(excel_data) >= 2 else 0
        )

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # export info
    worksheet2 = workbook.create_sheet("Export Info")
    worksheet2.column_dimensions['A'].width = 14
    for n, export_info_item in enumerate(export_info, start=1):
        worksheet2['A{0}'.format(n)].value = export_info_item[0]
        worksheet2['B{0}'.format(n)].value = export_info_item[1]

    # saving file
    file_hash = uuid.uuid4().hex
    export_file = BytesIO()
    icds_file = IcdsFile(blob_id=file_hash, data_type=data_type)
    workbook.save(export_file)
    export_file.seek(0)
    icds_file.store_file_in_blobdb(export_file, expired=ONE_DAY)
    icds_file.save()
    return file_hash


def get_datatables_ordering_info(request):
    # retrive table ordering provided by datatables plugin upon clicking on column header
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    order_by_number_column = request.GET.get('order[0][column]')
    order_by_name_column = request.GET.get('columns[%s][data]' % order_by_number_column)
    order_dir = request.GET.get('order[0][dir]', 'asc')
    return start, length, order_by_number_column, order_by_name_column, order_dir


def phone_number_function(x):
    return "+{0}{1}".format('' if str(x).startswith('91') else '91', x) if x else x


def filter_cas_data_export(export_file, location):
    with TransientTempfile() as path:
        with open(path, 'wb') as temp_file:
            temp_file.write(export_file.get_file_from_blobdb().read())
        with open(path, 'r') as temp_file:
            fd, export_file_path = mkstemp()
            csv_file = os.fdopen(fd, 'w')
            reader = csv.reader(temp_file)
            writer = csv.writer(csv_file)
            headers = next(reader)
            for i, header in enumerate(headers):
                if header == f'{location.location_type.name}_name':
                    index_of_location_type_name_column = i
                    break
            else:
                raise InvalidLocationType(f'{location.location_type.name} is not a valid location option for cas data exports')
            writer.writerow(headers)
            for row in reader:
                if row[index_of_location_type_name_column] == location.name:
                    writer.writerow(row)
        return export_file_path


def prepare_rollup_query(columns_tuples):
    def _columns_and_calculations(column_tuple):
        column = column_tuple[0]

        if len(column_tuple) == 2:
            agg_col = column_tuple[1]
            if isinstance(agg_col, str):
                return column_tuple
            elif callable(agg_col):
                return column, agg_col(column)
        return column, 'SUM({})'.format(column)

    columns = list(map(_columns_and_calculations, columns_tuples))

    column_names = ", ".join([col[0] for col in columns])
    calculations = ", ".join([col[1] for col in columns])
    return column_names, calculations


def test_column_name(aggregation_level):
    test_location_column_names = ['district_is_test', 'block_is_test', 'supervisor_is_test', 'awc_is_test']
    return test_location_column_names[aggregation_level-1]


def create_group_by(aggregation_level):
    all_group_by_columns = ["state_id", "district_id", "block_id", "supervisor_id"]
    return all_group_by_columns[:aggregation_level]


def column_value_as_per_agg_level(aggregation_level, agg_level_threshold, true_value, false_value):
    return true_value if aggregation_level > agg_level_threshold else false_value


@attr.s
class AggLevelInfo(object):
    agg_level = attr.ib()
    col_name = attr.ib()


def _construct_replacement_map_from_awc_location(loc_level, replacement_location_ids):
    from custom.icds_reports.models.aggregate import AwcLocation

    def _full_hierarchy_name(loc):
        loc_names = [loc[f'{level}_name'] for level in levels.keys() if loc[f'{level}_name']]
        return ' > '.join(loc_names)

    levels = {
        'state': AggLevelInfo(AggregationLevels.STATE, 'state_id'),
        'district': AggLevelInfo(AggregationLevels.DISTRICT, 'district_id'),
        'block': AggLevelInfo(AggregationLevels.BLOCK, 'block_id'),
        'supervisor': AggLevelInfo(AggregationLevels.SUPERVISOR, 'supervisor_id'),
        'awc': AggLevelInfo(AggregationLevels.AWC, 'doc_id')
    }
    level_info = levels[loc_level]
    filters = {
        f'{level_info.col_name}__in': replacement_location_ids,
        'aggregation_level': level_info.agg_level
    }
    columns = [
        level_info.col_name,
        'state_name',
        'district_name',
        'block_name',
        'supervisor_name',
        'awc_name'
    ]
    replacement_locations = AwcLocation.objects.filter(**filters).values(*columns)
    replacement_names = {loc[level_info.col_name]: _full_hierarchy_name(loc) for loc in replacement_locations}
    return replacement_names


def _construct_replacement_map_from_sql_location(replacement_location_ids):

    def _full_hierarchy_name(loc):
        loc_names = []
        while loc is not None:
            loc_names.append(loc.name)
            loc = loc.parent
        loc_names.reverse()
        return ' > '.join(loc_names)

    # prefetch all possible parents
    replacement_locations = SQLLocation.objects.filter(location_id__in=replacement_location_ids).select_related('parent__parent__parent__parent')

    replacement_names = {loc.location_id: _full_hierarchy_name(loc) for loc in replacement_locations}
    return replacement_names


def get_deprecation_info(locations, show_test, multiple_levels=False):
    locations_list = []
    replacement_location_ids = []
    for loc in locations:
        loc_level = loc.location_type.name
        if show_test or loc.metadata.get('is_test_location', 'real') != 'test':
            locations_list.append(loc)
            if loc.metadata.get('deprecated_to'):
                replacement_location_ids.extend(loc.metadata['deprecated_to'].split(','))
            if loc.metadata.get('deprecates'):
                replacement_location_ids.extend(loc.metadata['deprecates'].split(','))

    if multiple_levels:
        replacement_names = _construct_replacement_map_from_sql_location(replacement_location_ids)

    else:
        replacement_names = _construct_replacement_map_from_awc_location(loc_level, replacement_location_ids)

    return locations_list, replacement_names


def get_location_replacement_name(location, field, replacement_names):
    locations = location.metadata.get(field)
    if locations:
        locations = locations.split(',')
    else:
        locations = []
    return [replacement_names.get(loc_id, '') for loc_id in locations]


@icds_quickcache(['filters', 'loc_name'], timeout=30 * 60)
def get_location_launched_status(filters, loc_name):
    from custom.icds_reports.models import AggAwcMonthly

    def select_location_filter(filters):
        location_filters = dict()
        location_filters['aggregation_level'] = filters['aggregation_level']
        location_filters['month'] = filters['month']

        if location_filters['aggregation_level'] == 1 and 'state_id' not in filters:
            return location_filters
        else:
            location_id_cols = ['state_id', 'district_id', 'block_id', 'supervisor_id', 'awc_id']
            # Subtracting 1 from agg level because agg level is 1 level deeper as it finds the
            # launched status of all sub locations
            reduced_loc_id_cols = location_id_cols[:location_filters['aggregation_level'] - 1]

            location_filters.update(
                {
                    col: filters[col]
                    for col in reduced_loc_id_cols
                }
            )

        return location_filters

    locations_launched_status = AggAwcMonthly.objects.filter(
        **select_location_filter(filters)
    ).values('%s_name' % loc_name, 'num_launched_awcs')

    return {loc['%s_name' % loc_name]: loc['num_launched_awcs'] for loc in locations_launched_status}


def timestamp_string_to_date_string(ts_string):
    if ts_string:
        return parse(ts_string).strftime(ISO_DATE_FORMAT)
    else:
        return None


def datetime_to_date_string(dtime):
    if dtime:
        return dtime.strftime(ISO_DATE_FORMAT)
    else:
        return None


def generate_quarter_months(quarter, year):
    months = []
    end_month = int(quarter) * 3
    for i in range(end_month - 2, end_month + 1):
        months.append(date(year, i, 1))
    return months


def calculate_percent(num, den, extra_number, truncate_out=True):
    if den == 0:
        ret = 0
    else:
        ret = (num / den) * 100

    if extra_number:
        ret = ret / extra_number
    if truncate_out is True:
        return "{}%".format("%.2f" % ret)
    else:
        return ret


def handle_average(val):
    if val is None:
        ret = 0
    else:
        ret = val / 3
    return ret


def get_filters_from_config_for_chart_view(config):
    config_filter = copy.deepcopy(config)
    if 'gender' in config_filter:
        config_filter['sex'] = config_filter['gender']
        del config_filter['gender']
    del config_filter['aggregation_level']
    return config_filter
