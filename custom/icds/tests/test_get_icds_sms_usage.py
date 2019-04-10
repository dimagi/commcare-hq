from __future__ import absolute_import
from __future__ import unicode_literals

import openpyxl
import os

from datetime import date
from django.core.management import call_command
from openpyxl.utils import get_column_letter

from corehq.apps.domain.shortcuts import create_domain
from corehq.apps.locations.models import LocationType, SQLLocation
from corehq.apps.sms.models import SMS
from corehq.apps.users.models import CommCareUser
from corehq.apps.locations.tests.util import make_loc, setup_location_types
from custom.icds.tests.base import BaseICDSTest
from corehq.messaging.smsbackends.airtel_tcl.models import AirtelTCLBackend
from six.moves import range


class GetICDSSmsUsageTest(BaseICDSTest):
    domain = 'domain'
    sms_list = []

    @classmethod
    def setUpClass(cls):
        super(GetICDSSmsUsageTest, cls).setUpClass()
        cls.domain_obj = create_domain(cls.domain)

        def make_user(name, location):
            user = CommCareUser.create(cls.domain, name, 'password')
            user.set_location(location)
            return user

        cls.loc_types = setup_location_types(cls.domain, ['state', 'district', 'block', 'supervisor', 'awc'])
        cls.states = (
            make_loc('state', type='state', domain=cls.domain),
            make_loc('state2', type='state', domain=cls.domain),
        )
        cls.districts = (
            make_loc('district', type='district', domain=cls.domain, parent=cls.states[0]),
            make_loc('district2', type='district', domain=cls.domain, parent=cls.states[1]),
        )
        cls.blocks = (
            make_loc('block', type='block', domain=cls.domain, parent=cls.districts[0]),
            make_loc('block2', type='block', domain=cls.domain, parent=cls.districts[1]),
        )
        cls.supervisors = (
            make_loc('supervisor', type='supervisor', domain=cls.domain, parent=cls.blocks[0]),
            make_loc('supervisor2', type='supervisor', domain=cls.domain, parent=cls.blocks[1]),
        )
        cls.awcs = (
            make_loc('awc', type='awc', domain=cls.domain, parent=cls.supervisors[0]),
            make_loc('awc2', type='awc', domain=cls.domain, parent=cls.supervisors[1]),
        )
        cls.users = (
            make_user('user', cls.awcs[0]),
            make_user('user2', cls.awcs[1]),
        )

        cls.sms_list.append(SMS.objects.create(
            domain=cls.domain,
            date=date(2017, 4, 10),
            backend_api=AirtelTCLBackend.get_api_id(),
            direction='O',
            processed=True,
            couch_recipient=cls.users[0]._id,
            custom_metadata={'icds_indicator': 'xxx'}
        ))
        cls.sms_list.append(SMS.objects.create(
            domain=cls.domain,
            date=date(2017, 4, 10),
            backend_api=AirtelTCLBackend.get_api_id(),
            direction='O',
            processed=True,
            couch_recipient=cls.users[0]._id,
            custom_metadata={'icds_indicator': 'xxx'}
        ))
        cls.sms_list.append(SMS.objects.create(
            domain=cls.domain,
            date=date(2017, 4, 10),
            backend_api=AirtelTCLBackend.get_api_id(),
            direction='O',
            processed=True,
            couch_recipient=cls.users[1]._id,
            custom_metadata={'icds_indicator': 'xxx'}
        ))
        cls.sms_list.append(SMS.objects.create(
            domain=cls.domain,
            date=date(2017, 4, 10),
            backend_api=AirtelTCLBackend.get_api_id(),
            direction='O',
            processed=True,
            couch_recipient=cls.users[1]._id,
            custom_metadata={'icds_indicator': 'aaa'}
        ))
        date_start = '2017-04-01'
        date_end = '2017-05-01'
        call_command("get_icds_sms_usage", cls.domain, date_start, date_end)
        cls.workbook_name = 'icds-sms-usage--{0}--{1}.xlsx'.format(date_start, date_end)
        cls.workbook = openpyxl.load_workbook(cls.workbook_name)

    @classmethod
    def tearDownClass(cls):
        cls.users[0].delete()
        cls.users[1].delete()
        SMS.objects.filter(
            domain=cls.domain,
        ).delete()
        SQLLocation.objects.filter(
            domain=cls.domain,
        ).delete()
        LocationType.objects.filter(
            domain=cls.domain,
        ).delete()
        cls.workbook.close()
        os.remove(cls.workbook_name)
        super(GetICDSSmsUsageTest, cls).tearDownClass()

    def test_sheet_names(self):
        self.assertEquals(self.workbook.get_sheet_names(), ['icds-sms-usage', 'icds-sms-usage-by-district'])

    def test_values_by_state(self):
        sheet = self.workbook.get_sheet_by_name('icds-sms-usage')
        self.assertEquals(sheet.max_column, 4)
        self.assertEquals(sheet.max_row, 4)
        data = []
        for row_number in range(1, sheet.max_row + 1):
            data.append([])
            for column_number in range(1, sheet.max_column + 1):
                data[row_number - 1].append(
                    sheet['{0}{1}'.format(get_column_letter(column_number), row_number)].value
                )
        self.assertEquals(
            data,
            [
                [
                    'State Code',
                    'State Name',
                    'Indicator',
                    'SMS Count'
                ],
                [
                    'state',
                    'state',
                    'xxx',
                    2
                ],
                [
                    'state2',
                    'state2',
                    'aaa',
                    1
                ],
                [
                    'state2',
                    'state2',
                    'xxx',
                    1
                ]
            ]
        )

    def test_values_by_district(self):
        sheet = self.workbook.get_sheet_by_name('icds-sms-usage-by-district')
        self.assertEquals(sheet.max_column, 6)
        self.assertEquals(sheet.max_row, 4)
        data = []
        for row_number in range(1, sheet.max_row + 1):
            data.append([])
            for column_number in range(1, sheet.max_column + 1):
                data[row_number - 1].append(
                    sheet['{0}{1}'.format(get_column_letter(column_number), row_number)].value
                )
        self.assertEquals(
            data,
            [
                [
                    'State Code',
                    'State Name',
                    'District Code',
                    'District Name',
                    'Indicator',
                    'SMS Count'
                ],
                [
                    'state',
                    'state',
                    'district',
                    'district',
                    'xxx',
                    2
                ],
                [
                    'state2',
                    'state2',
                    'district2',
                    'district2',
                    'aaa',
                    1
                ],
                [
                    'state2',
                    'state2',
                    'district2',
                    'district2',
                    'xxx',
                    1
                ]
            ]
        )
