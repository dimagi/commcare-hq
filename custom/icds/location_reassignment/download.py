import io
import itertools
import zipfile
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from couchexport.export import export_raw

from corehq.apps.es import UserES
from corehq.apps.locations.models import SQLLocation
from corehq.form_processor.interfaces.dbaccessors import CaseAccessors
from custom.icds.location_reassignment.const import (
    AWC_CODE_COLUMN,
    AWC_NAME_COLUMN,
    BLOCK_CODE,
    CASE_ID_COLUMN,
    CASE_NAME,
    CURRENT_LGD_CODE,
    CURRENT_NAME,
    CURRENT_PARENT_NAME,
    CURRENT_PARENT_SITE_CODE,
    CURRENT_PARENT_TYPE,
    CURRENT_SITE_CODE_COLUMN,
    CURRENT_SUB_DISTRICT_NAME,
    EXTRACT_OPERATION,
    HAVE_APPENDED_LOCATION_NAMES,
    HOUSEHOLD_ID_COLUMN,
    HOUSEHOLD_MEMBER_DETAILS_COLUMN,
    LGD_CODE,
    MAP_LOCATION_NAME,
    NEW_LGD_CODE,
    NEW_NAME,
    NEW_PARENT_SITE_CODE,
    NEW_SITE_CODE_COLUMN,
    NEW_SUB_DISTRICT_NAME,
    NEW_USERNAME_COLUMN,
    OPERATION_COLUMN,
    OPERATIONS_TO_IGNORE,
    PERSON_CASE_TYPE,
    SHEETS_TO_IGNORE,
    SPLIT_OPERATION,
    USERNAME_COLUMN,
    VALID_OPERATIONS,
)
from custom.icds.location_reassignment.utils import (
    get_case_ids_for_reassignment,
    get_household_case_ids,
    get_household_child_cases_by_owner,
    split_location_name_and_site_code,
)


class Download(object):
    def __init__(self, location):
        """
        Generates an Excel file stream
        With details of all locations related to a location
        and users assigned to these locations.
        Each sheet corresponds to a location type.
        """
        self.location = location

    def dump(self):
        self._init_location_details()
        self._populate_assigned_users()
        wb = self._create_workbook()
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    def _init_location_details(self):
        # setup details for the main location and all related locations
        self._location_details_by_location_id = {}
        for location in self._locations():
            location_name = location.name
            if location.location_type.code in HAVE_APPENDED_LOCATION_NAMES:
                location_name, _ = split_location_name_and_site_code(location.name)
            self._location_details_by_location_id[location.location_id] = {
                'name': location_name,
                'type_name': location.location_type.name,
                'site_code': location.site_code,
                'lgd_code': location.metadata.get(LGD_CODE, ''),
                'sub_district_name': location.metadata.get(MAP_LOCATION_NAME, ''),
                'parent_location_id': location.parent_location_id,
                'location_type': location.location_type.code,
                'assigned_users': [],
            }

    def _locations(self):
        # fetch all locations necessary for this download request
        ancestors = list(self.location.get_ancestors().select_related('location_type'))
        self_and_descendants = list(self.location.get_descendants(include_self=True)
                                    .filter(is_archived=False).select_related('location_type'))
        return ancestors + self_and_descendants

    def _populate_assigned_users(self):
        # allot assign user details to each location
        for username, assigned_location_ids in self._get_assigned_location_ids().items():
            for location_id in assigned_location_ids:
                if location_id in self._location_details_by_location_id:
                    self._location_details_by_location_id[location_id]['assigned_users'].append(username)

    def _get_assigned_location_ids(self):
        location_ids = list(self._location_details_by_location_id.keys())
        assigned_location_ids_per_username = {}
        user_details = UserES().location(location_ids).values_list('base_username', 'assigned_location_ids')
        for username, assigned_location_ids in user_details:
            if not isinstance(assigned_location_ids, list):
                assigned_location_ids = [assigned_location_ids]
            assigned_location_ids_per_username[username] = assigned_location_ids
        return assigned_location_ids_per_username

    def _create_workbook(self):
        wb = Workbook()
        # workbook adds an empty sheet for new workbook unless it is write only
        wb.remove(wb.active)
        for location_type, rows in self._create_rows().items():
            worksheet = wb.create_sheet(location_type)
            uniq_headers = self._extract_unique_headers(rows)
            worksheet.append(uniq_headers)
            for row in rows:
                worksheet.append([row.get(header) for header in uniq_headers])
            self._add_validation(worksheet)
        self._add_extra_sheets(wb)
        return wb

    def _create_rows(self):
        # location type code mapped to rows which is a list of dictionaries to pull headers later using keys
        rows = defaultdict(list)
        for location_id, location_details in self._location_details_by_location_id.items():
            location_type = location_details['location_type']
            if location_details['assigned_users']:
                for username in location_details['assigned_users']:
                    rows[location_type].append(self._generate_row(username, location_type, location_id))
            else:
                # set up one row for location that has no user assigned
                rows[location_type].append(self._generate_row('', location_type, location_id))
        return rows

    def _generate_row(self, username, location_type, location_id):
        location_details = self._location_details_by_location_id[location_id]
        location_parent_id = location_details['parent_location_id']
        row = {
            CURRENT_SITE_CODE_COLUMN: location_details['site_code'],
            NEW_SITE_CODE_COLUMN: '',
            CURRENT_NAME: location_details['name'],
            NEW_NAME: '',
        }
        if location_type == BLOCK_CODE:
            row.update({
                CURRENT_SUB_DISTRICT_NAME: location_details['sub_district_name'],
                NEW_SUB_DISTRICT_NAME: '',
            })
        row.update({
            USERNAME_COLUMN: username,
            NEW_USERNAME_COLUMN: '',
            CURRENT_LGD_CODE: location_details['lgd_code'],
            NEW_LGD_CODE: '',
        })
        if location_parent_id and location_parent_id in self._location_details_by_location_id:
            location_parent_details = self._location_details_by_location_id[location_parent_id]
            parent_name = location_parent_details['name']
            if location_parent_details['location_type'] in HAVE_APPENDED_LOCATION_NAMES:
                parent_name, _ = split_location_name_and_site_code(parent_name)
            row.update({
                CURRENT_PARENT_TYPE: location_parent_details['type_name'],
                CURRENT_PARENT_NAME: parent_name,
                CURRENT_PARENT_SITE_CODE: location_parent_details['site_code'],
                NEW_PARENT_SITE_CODE: ''
            })
        row[OPERATION_COLUMN] = ''
        return row

    @staticmethod
    def _extract_unique_headers(rows):
        # extract headers preserving the order of occurrence
        headers_for_all_rows = list(itertools.chain.from_iterable(rows))
        uniq_headers = []
        for h in headers_for_all_rows:
            if h not in uniq_headers:
                uniq_headers.append(h)
        return uniq_headers

    @staticmethod
    def _add_validation(worksheet):
        operations = [""] + VALID_OPERATIONS + OPERATIONS_TO_IGNORE
        operation_data_validation = DataValidation(type="list", formula1='"%s"' % (','.join(operations)))
        worksheet.add_data_validation(operation_data_validation)
        for header_cell in worksheet[1]:
            if header_cell.value == OPERATION_COLUMN:
                letter = header_cell.column_letter
                operation_data_validation.add(f"{letter}2:{letter}{worksheet.max_row}")

    @staticmethod
    def _add_extra_sheets(workbook):
        for title, headers in SHEETS_TO_IGNORE.items():
            worksheet = workbook.create_sheet(title)
            worksheet.append(headers)


class Households(object):
    """
    Download Household cases assigned the old locations undergoing transition
    """
    valid_operations = [SPLIT_OPERATION, EXTRACT_OPERATION]
    headers = [AWC_NAME_COLUMN, AWC_CODE_COLUMN, 'Name of Household', 'Date of Registration', 'Religion',
               'Caste', 'APL/BPL', 'Number of Household Members', HOUSEHOLD_MEMBER_DETAILS_COLUMN,
               HOUSEHOLD_ID_COLUMN]

    def __init__(self, domain):
        self.domain = domain

    def dump(self, transitions):
        """
        :return: excel workbook with each tab titled as old location's site code,
        which holds details all household cases assigned to it
        """
        rows = {}
        for transition in transitions:
            operation = transition.operation
            if operation in self.valid_operations:
                rows.update(self._get_rows_for_transition(transition))
        if rows:
            stream = io.BytesIO()
            headers = [[site_code, self.headers] for site_code in rows]
            export_raw(headers, rows.items(), stream)
            stream.seek(0)
            return stream

    def _get_rows_for_transition(self, transition):
        rows = {}
        for site_code in transition.old_site_codes:
            rows[site_code] = self._build_rows(site_code)
        return rows

    def _build_rows(self, site_code):
        rows = []
        location = SQLLocation.active_objects.get(domain=self.domain, site_code=site_code)
        case_ids = get_household_case_ids(self.domain, location.location_id)
        for case_id in case_ids:
            household_case = CaseAccessors(self.domain).get_case(case_id)
            person_cases = get_household_child_cases_by_owner(
                self.domain, case_id, location.location_id, [PERSON_CASE_TYPE])
            rows.append([
                '',
                '',
                household_case.name,
                household_case.get_case_property('hh_reg_date'),
                household_case.get_case_property('hh_religion'),
                household_case.get_case_property('hh_caste'),
                household_case.get_case_property('hh_bpl_apl'),
                len(person_cases),
                ", ".join([
                    "%s (%s/%s)" % (
                        case.name, case.get_case_property('age_at_reg'), case.get_case_property('sex'))
                    for case in person_cases
                ]),
                household_case.case_id
            ])
        return rows


class OtherCases(object):
    """
    Download other cases that are not households or child cases of households
    """
    valid_operations = [SPLIT_OPERATION, EXTRACT_OPERATION]
    headers = [NEW_NAME, NEW_SITE_CODE_COLUMN, CASE_NAME, CASE_ID_COLUMN]

    def __init__(self, domain):
        self.domain = domain

    def dump(self, transitions):
        """
        :return: zip file with
        one workbook per case type
        with each sheet title as old location's site code,
        which holds details of all cases assigned to it
        """
        data = self._generate_data(transitions)
        if not data:
            return None
        zipstream = io.BytesIO()
        with zipfile.ZipFile(zipstream, "w") as z:
            for case_type, sheets in data.items():
                excel_file = self._dump_to_excel(sheets)
                z.writestr(f"{case_type}.xlsx", excel_file.read(), zipfile.ZIP_DEFLATED)
        zipstream.seek(0)
        return zipstream

    def _generate_data(self, transitions):
        # each case type mapped to sheets
        # mapped to rows for cases of that case type assigned to that site code
        data = defaultdict(dict)
        for transition in transitions:
            operation = transition.operation
            if operation in self.valid_operations:
                for old_site_code in transition.old_site_codes:
                    for case_type, rows in self._get_rows_for_location(old_site_code).items():
                        data[case_type][old_site_code] = rows
        return data

    def _get_rows_for_location(self, site_code):
        rows_per_case_type = defaultdict(list)
        location = SQLLocation.active_objects.get(domain=self.domain, site_code=site_code)
        _, case_ids = get_case_ids_for_reassignment(self.domain, location.location_id)
        if not case_ids:
            return rows_per_case_type
        cases = CaseAccessors(self.domain).get_cases(case_ids)
        for case in cases:
            rows_per_case_type[case.type].append([
                '',
                '',
                case.name,
                case.case_id
            ])
        return rows_per_case_type

    def _dump_to_excel(self, sheets):
        stream = io.BytesIO()
        headers = [[site_code, self.headers] for site_code in sheets]
        export_raw(headers, sheets.items(), stream)
        stream.seek(0)
        return stream
