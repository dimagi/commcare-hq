from __future__ import absolute_import
from __future__ import print_function
from datetime import datetime
from openpyxl import Workbook
import csv

from django.core.management.base import BaseCommand
from corehq.form_processor.exceptions import CaseNotFound
from corehq.form_processor.interfaces.dbaccessors import CaseAccessors
from corehq.motech.repeaters.dbaccessors import get_repeat_records_by_payload_id
from corehq.pillows.mappings.case_search_mapping import CASE_SEARCH_MAX_RESULTS
from corehq.apps.es.case_search import CaseSearchES
from corehq.util.log import with_progress_bar

from casexml.apps.case.xml.parser import CaseUpdateAction
from casexml.apps.case.xform import get_case_updates
from custom.enikshay.case_utils import get_person_case_from_episode, get_open_episode_case_from_person
from custom.enikshay.const import DSTB_EPISODE_TYPE
from custom.enikshay.exceptions import ENikshayCaseNotFound
from custom.enikshay.integrations.utils import is_valid_person_submission
from six.moves import range

DOMAIN = "enikshay"


class Command(BaseCommand):
    help = """
    Iterate over a list of episode case ids and check if they have been
    notified to nikshay or not
    Pass in a csv file that has
    1. episode case ids under header episode_id or
    2. person enikshay ids(not uuid) under header enikshay_id (along with option --parse_as_person_enikshay_ids)

    Gives a report with:
    'Enikshay ID', 'Episode ID', 'Should be Notified', 'Nikshay Registered', 'Nikshay Private Registered',
    'Nikshay ID', 'External ID', 'Nikshay ID history', 'Nikshay Error', 'Private Nikshay Error',
    """

    def __init__(self, *args, **kwargs):
        super(Command, self).__init__(*args, **kwargs)
        self.person_enikshay_ids = []
        self.person_ids = []
        self.episode_ids = []
        self.max_attempts_in_sheet = 0
        self.search = CaseSearchES().domain('enikshay').size(CASE_SEARCH_MAX_RESULTS)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append([
            'Enikshay ID', 'Episode ID', 'Should be Notified', 'Nikshay Registered', 'Nikshay Private Registered',
            'Nikshay ID', 'External ID', 'Nikshay ID history', 'Nikshay Error', 'Private Nikshay Error',
            # Forwarding ID 1, Attempts 1, Forwarding ID 2, Attempts 2 ..
        ])

    def add_arguments(self, parser):
        parser.add_argument('file_path')
        parser.add_argument(
            '--parse_as_person_enikshay_ids',
            dest="parse_as_person_enikshay_ids",
            action='store_true',
            help='pass this to let parser know to look for enikshay ids instead of episode ids in file'
        )

    def _search_person_using_enikshay_id(self, enikshay_id):
        return self.search.case_property_query("person_id", enikshay_id).values_list('_id', flat=True)

    def _load_episode_ids_from_person(self, file_path):
        """
        1. Load enikshay ids from file
        2. Find corresponding uuid for person cases (This can be more than one match since enikshay id
           can belong to more than one person due to some bug)
        3. Find corresponding open episode ids for each person. If not log in for manual check later.
        """
        print("Loading person enikshay ids from file")
        with open(file_path, 'rU') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                self.person_enikshay_ids.append(row['enikshay_id'])
            print("Loaded {number} eniksay ids from file".format(number=len(self.person_enikshay_ids)))

        print("Searching for person uuids corresponding to each enikshay id")
        for person_enikshay_id in self.person_enikshay_ids:
            self.person_ids += self._search_person_using_enikshay_id(person_enikshay_id)
        print("Loaded {number} person ids".format(number=len(self.person_enikshay_ids)))

        for person_id in self.person_ids:
            try:
                episode_case = get_open_episode_case_from_person(DOMAIN, person_id)
                self.episode_ids.append(episode_case.get_id)
            except ENikshayCaseNotFound:
                print(("Could not find episode for person with enikshay ID:", person_id))

    def _load_episode_ids_from_file(self, file_path):
        with open(file_path, 'rU') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                self.episode_ids.append(row['episode_id'])

    def _load_case_ids(self, file_path, parse_as_person_enikshay_ids):
        if parse_as_person_enikshay_ids:
            self._load_episode_ids_from_person(file_path)
        else:
            print("Loading episode ids from file")
            self._load_episode_ids_from_file(file_path)

    @staticmethod
    def _find_fowarding_attempts(episode_id):
        return get_repeat_records_by_payload_id(DOMAIN, episode_id)

    def _add_report_record_attempts(self, row, episode_id):
        for repeat_record in self._find_fowarding_attempts(episode_id):
            attempts = []
            repeat_record_count = len(repeat_record.attempts)
            if repeat_record_count > self.max_attempts_in_sheet:
                self.max_attempts_in_sheet = repeat_record_count

            for attempt in repeat_record.attempts:
                if attempt.message:
                    attempts.append(attempt.message)
            row.append(repeat_record.get_id)
            row.append(repeat_record.state)
            row.append(';'.join(attempts))

    def _add_row(self, episode_id, episode_case, episode_case_properties, should_be_forwarded, error_message=None):
        person_case = get_person_case_from_episode(DOMAIN, episode_id)
        person_enikshay_id = person_case.dynamic_case_properties().get('person_id')
        if error_message:
            # add blanks for pending columns to append correctly for repeat record attempts later
            row = [person_enikshay_id, episode_id, should_be_forwarded, error_message, '', '', '', '']
        else:
            row = [
                person_enikshay_id, episode_id, should_be_forwarded,
                episode_case_properties.get('nikshay_registered'),
                episode_case_properties.get('private_nikshay_registered'),
                episode_case_properties.get('nikshay_id'),
                episode_case.external_id,
                '->'.join(self._nikshay_id_update_history(episode_case)),
                episode_case_properties.get('nikshay_error'),
                episode_case_properties.get('private_nikshay_error'),
            ]
        self._add_report_record_attempts(row, episode_id)
        self.ws.append(row)

    def _add_repeat_record_attempt_headers(self):
        max_columns = self.ws.get_highest_column()
        first_empty_index = max_columns
        for i in range(1, max_columns + 1):
            if not self.ws.cell(row=1, column=i).value:
                first_empty_index = i
                break
        column_num = first_empty_index
        for i in range(0, self.max_attempts_in_sheet):
            self.ws.cell(row=1, column=column_num).value = "Attempt {index} ID".format(index=i + 1)
            self.ws.cell(row=1, column=(column_num + 1)).value = "Attempt {index} State".format(index=i + 1)
            self.ws.cell(row=1, column=(column_num + 2)).value = "Attempt {index} Messages".format(index=i + 1)
            column_num = column_num + 3

    def _save_file(self):
        if self.max_attempts_in_sheet > 0:
            self._add_repeat_record_attempt_headers()
        file_name = "nikshay_notification_report_{timestamp}.xlsx".format(
            timestamp=datetime.now().strftime("%Y-%m-%d-%H-%M-%S"),
        )
        self.wb.save(file_name)
        return file_name

    @staticmethod
    def _should_be_forwarded(episode_case, episode_case_properties):
        """
        This would add False for cases that have already been forwarded but
        is a good to have for other cases that are expected to be notified
        """
        person_case = get_person_case_from_episode(DOMAIN, episode_case)
        return (
            not episode_case_properties.get('nikshay_registered', 'false') == 'true' and
            not episode_case_properties.get('nikshay_registered', 'false') == 'true' and
            not episode_case_properties.get('nikshay_id', False) and
            episode_case_properties.get('episode_type') == DSTB_EPISODE_TYPE and
            is_valid_person_submission(person_case)
        )

    def _nikshay_id_update_history(self, episode_case):
        nikshay_id_update_history = []
        all_case_update_transactions = [
            trans for trans in episode_case.actions
            if not trans.is_case_create and not trans.is_case_rebuild
        ]
        all_case_actions = []
        for case_update_transaction in all_case_update_transactions:
            all_case_actions += [
                update.get_update_action() for update in get_case_updates(case_update_transaction.form)
                if update.id == episode_case.case_id
            ]

        for action in all_case_actions:
            if isinstance(action, CaseUpdateAction) and "nikshay_id" in action.dynamic_properties:
                nikshay_id_update_history.append(action.dynamic_properties.get("nikshay_id"))
        return nikshay_id_update_history

    def handle(self, file_path, *args, **options):
        print("Loading episode ids or person_ids")
        self._load_case_ids(file_path, options.get('parse_as_person_enikshay_ids'))
        print("Okay! Loaded {number} episode ids".format(number=len(self.episode_ids)))
        case_accessor = CaseAccessors(DOMAIN)
        for episode_id in with_progress_bar(self.episode_ids):
            should_be_forwarded = "Could not be determined"
            try:
                episode_case = case_accessor.get_case(episode_id)
                episode_case_properties = episode_case.dynamic_case_properties()
                should_be_forwarded = self._should_be_forwarded(episode_case, episode_case_properties)
                self._add_row(episode_id, episode_case, episode_case_properties, should_be_forwarded)
            except CaseNotFound:
                self._add_row(episode_id, None, None, should_be_forwarded,
                              error_message="Could not find episode case")

        file_name = self._save_file()
        print("Report saved in file:{filename}".format(filename=file_name))
